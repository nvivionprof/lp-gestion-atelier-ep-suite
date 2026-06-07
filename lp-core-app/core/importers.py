from openpyxl import load_workbook
from .models import CoreUser, CoreFormation, CoreClass, CoreAuditLog, normalize_code

ALIASES = {
    'code': ['code', 'code_utilisateur', 'id', 'identifiant_unique'],
    'username': ['identifiant', 'username', 'login', 'user', 'utilisateur'],
    'password': ['mot_de_passe_initial', 'password_initial', 'mot de passe initial', 'password', 'mdp'],
    'last_name': ['nom', 'last_name', 'lastname'],
    'first_name': ['prenom', 'prénom', 'first_name', 'firstname'],
    'email': ['email', 'mail', 'courriel'],
    'class_name': ['classe', 'class', 'classe_name'],
    'formation': ['formation', 'filiere', 'filière', 'formation_code'],
    'group_name': ['groupe', 'group', 'group_name'],
    'role': ['role', 'rôle', 'role_principal'],
    'rights': ['droits', 'rights', 'permissions'],
    'active': ['actif', 'active'],
    'school_year': ['annee_scolaire', 'année_scolaire', 'school_year'],
}

ROLE_MAP = {
    'UTILISATEUR': 'utilisateur', 'USER': 'utilisateur', 'ELEVE': 'eleve', 'ÉLÈVE': 'eleve', 'ELEVES': 'eleve',
    'MAGASINIER': 'magasinier', 'PROF': 'professeur', 'PROFESSEUR': 'professeur',
    'RESPONSABLE': 'responsable', 'ADMIN': 'admin', 'ADMINISTRATEUR': 'admin', 'LECTURE_SEULE': 'lecture_seule',
}


def norm_header(v):
    return str(v or '').strip().lower().replace('-', '_').replace(' ', '_')


def header_map(headers):
    normalized = {norm_header(h): idx for idx, h in enumerate(headers)}
    result = {}
    for field, aliases in ALIASES.items():
        for alias in aliases:
            key = norm_header(alias)
            if key in normalized:
                result[field] = normalized[key]
                break
    return result


def cell(row, mapping, field, default=''):
    idx = mapping.get(field)
    if idx is None or idx >= len(row):
        return default
    value = row[idx]
    return '' if value is None else str(value).strip()


def active_value(raw):
    if raw == '':
        return True
    return str(raw).strip().lower() not in {'0', 'false', 'faux', 'non', 'no', 'inactif'}


def import_users_xlsx(path, *, actor=None, source='excel'):
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {'created': 0, 'updated': 0, 'errors': ['Fichier vide']}
    mapping = header_map(rows[0])
    required = ['last_name', 'first_name']
    missing = [field for field in required if field not in mapping]
    if missing:
        return {'created': 0, 'updated': 0, 'errors': [f'Colonnes manquantes : {", ".join(missing)}']}
    created = updated = 0
    errors = []
    for lineno, row in enumerate(rows[1:], start=2):
        last_name = cell(row, mapping, 'last_name').upper()
        first_name = cell(row, mapping, 'first_name')
        if not last_name and not first_name:
            continue
        formation_raw = cell(row, mapping, 'formation') or 'GEN'
        formation_code = normalize_code(formation_raw, 'GEN')
        formation, _ = CoreFormation.objects.get_or_create(code=formation_code, defaults={'name': formation_raw})
        class_name = cell(row, mapping, 'class_name')
        if class_name:
            CoreClass.objects.get_or_create(formation=formation, name=class_name, school_year=cell(row, mapping, 'school_year'))
        code = cell(row, mapping, 'code')
        username = cell(row, mapping, 'username')
        if not code:
            code = f'{formation_code}-{last_name[:3]}-{first_name[:3]}'
        code = normalize_code(code, 'USER')
        if not username:
            username = code
        role_raw = cell(row, mapping, 'role') or 'UTILISATEUR'
        role = ROLE_MAP.get(role_raw.strip().upper(), role_raw.strip().lower() or 'utilisateur')
        password = cell(row, mapping, 'password')
        defaults = {
            'username': username,
            'first_name': first_name,
            'last_name': last_name,
            'email': cell(row, mapping, 'email'),
            'formation': formation,
            'class_name': class_name,
            'group_name': cell(row, mapping, 'group_name'),
            'role_principal': role,
            'rights': cell(row, mapping, 'rights'),
            'active': active_value(cell(row, mapping, 'active')),
            'school_year': cell(row, mapping, 'school_year'),
            'initial_password_for_sync': password,
            'source': source,
        }
        try:
            obj, was_created = CoreUser.objects.update_or_create(code=code, defaults=defaults)
            if password:
                obj.set_password(password)
                obj.save(update_fields=['password_hash', 'initial_password_for_sync', 'updated_at'])
            if was_created:
                created += 1
            else:
                updated += 1
        except Exception as exc:
            errors.append(f'Ligne {lineno}: {exc}')
    CoreAuditLog.objects.create(actor=actor, action='IMPORT_XLSX', target=str(path), details=f'{created} créés, {updated} modifiés, erreurs={len(errors)}')
    return {'created': created, 'updated': updated, 'errors': errors}
