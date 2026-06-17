from decimal import Decimal
from pathlib import Path

from django.apps import apps
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.db import models
from django.utils.dateparse import parse_date, parse_datetime
from django.utils import timezone

from openpyxl import Workbook, load_workbook


APP_LABEL = "core"
MODULE_NAME = "LP Core"


def _bool(value):
    if value is None or value == "":
        return False
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() in {"1", "true", "vrai", "oui", "yes", "y"}


def _text(value):
    if value is None:
        return ""
    return str(value).strip()


def _field_to_cell(obj, field):
    value = getattr(obj, field.name)

    if value is None:
        return ""

    if isinstance(field, models.ForeignKey):
        return getattr(obj, field.attname) or ""

    if isinstance(field, (models.FileField, models.ImageField)):
        return value.name or ""

    if isinstance(field, models.DateTimeField):
        if value:
            return timezone.localtime(value).isoformat()
        return ""

    if isinstance(field, models.DateField):
        return value.isoformat() if value else ""

    if isinstance(field, models.BooleanField):
        return bool(value)

    return value


def _cell_to_field(field, value):
    if value == "":
        value = None

    if isinstance(field, models.ForeignKey):
        return field.attname, value

    if value is None:
        if field.blank or field.null:
            return field.name, None if field.null else ""
        return field.name, field.get_default()

    if isinstance(field, models.BooleanField):
        return field.name, _bool(value)

    if isinstance(field, (models.IntegerField, models.PositiveIntegerField, models.PositiveSmallIntegerField)):
        return field.name, int(value)

    if isinstance(field, models.DecimalField):
        return field.name, Decimal(str(value))

    if isinstance(field, models.FloatField):
        return field.name, float(value)

    if isinstance(field, models.DateTimeField):
        parsed = parse_datetime(str(value))
        if parsed and timezone.is_naive(parsed):
            parsed = timezone.make_aware(parsed)
        return field.name, parsed

    if isinstance(field, models.DateField):
        return field.name, parse_date(str(value))

    if isinstance(field, (models.FileField, models.ImageField)):
        return field.name, _text(value)

    return field.name, value


def _models():
    return [
        model for model in apps.get_app_config(APP_LABEL).get_models()
        if not model._meta.proxy and not model._meta.auto_created
    ]


def _sheet_name(model):
    name = model.__name__
    return name[:31]


def _concrete_fields(model):
    fields = []
    for field in model._meta.fields:
        if field.auto_created and field.name != "id":
            continue
        fields.append(field)
    return fields


def _m2m_fields(model):
    return [field for field in model._meta.many_to_many if not field.auto_created]


def _write_readme(wb, action):
    ws = wb.create_sheet("README", 0)
    ws.append(["Module", MODULE_NAME])
    ws.append(["App label", APP_LABEL])
    ws.append(["Action", action])
    ws.append([])
    ws.append(["Principe"])
    ws.append(["Chaque feuille correspond à un modèle Django du module."])
    ws.append(["Les clés étrangères sont exportées sous forme d'ID."])
    ws.append(["Les champs ManyToMany sont exportés en liste d'ID séparés par ;."])
    ws.append(["Les fichiers/images ne sont pas inclus dans le XLSX : seul le chemin média est conservé."])
    ws.append([])
    ws.append(["Important"])
    ws.append(["Conserver aussi une archive média et un dump JSON avant réinstallation complète."])


def export_xlsx(output_path, template=False):
    wb = Workbook()
    wb.remove(wb.active)
    _write_readme(wb, "template" if template else "export")

    for model in _models():
        ws = wb.create_sheet(_sheet_name(model))
        concrete = _concrete_fields(model)
        m2m = _m2m_fields(model)

        headers = [field.name for field in concrete]
        headers += [f"m2m__{field.name}" for field in m2m]
        ws.append(headers)

        if template:
            continue

        queryset = model.objects.all().order_by("pk")
        for obj in queryset:
            row = [_field_to_cell(obj, field) for field in concrete]
            for field in m2m:
                ids = list(getattr(obj, field.name).all().values_list("pk", flat=True))
                row.append(";".join(str(x) for x in ids))
            ws.append(row)

    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    return output


def import_xlsx(input_path, dry_run=False):
    wb = load_workbook(input_path, data_only=True)
    created = 0
    updated = 0
    m2m_later = []

    with transaction.atomic():
        for model in _models():
            sheet = _sheet_name(model)
            if sheet not in wb.sheetnames:
                continue

            ws = wb[sheet]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue

            headers = [str(h).strip() if h is not None else "" for h in rows[0]]
            concrete = {field.name: field for field in _concrete_fields(model)}
            m2m = {f"m2m__{field.name}": field for field in _m2m_fields(model)}

            for values in rows[1:]:
                if not values or not any(v not in [None, ""] for v in values):
                    continue

                raw = dict(zip(headers, values))
                pk = raw.get("id") or None

                defaults = {}
                for header, value in raw.items():
                    if header in concrete:
                        field = concrete[header]
                        if field.name == "id":
                            continue
                        key, parsed = _cell_to_field(field, value)
                        defaults[key] = parsed

                if pk:
                    obj, was_created = model.objects.update_or_create(pk=int(pk), defaults=defaults)
                else:
                    obj = model.objects.create(**defaults)
                    was_created = True

                if was_created:
                    created += 1
                else:
                    updated += 1

                for header, field in m2m.items():
                    value = raw.get(header)
                    if value is not None:
                        m2m_later.append((obj, field.name, value))

        for obj, field_name, value in m2m_later:
            ids = []
            for part in str(value or "").split(";"):
                part = part.strip()
                if part:
                    ids.append(int(float(part)))
            getattr(obj, field_name).set(ids)

        if dry_run:
            transaction.set_rollback(True)

    return created, updated, len(m2m_later)


class Command(BaseCommand):
    help = "Export, modèle et import XLSX générique pour le module."

    def add_arguments(self, parser):
        sub = parser.add_subparsers(dest="action", required=True)

        p_template = sub.add_parser("template")
        p_template.add_argument("--output", required=True)

        p_export = sub.add_parser("export")
        p_export.add_argument("--output", required=True)

        p_import = sub.add_parser("import")
        p_import.add_argument("--input", required=True)
        p_import.add_argument("--dry-run", action="store_true")

    def handle(self, *args, **options):
        action = options["action"]

        if action == "template":
            output = export_xlsx(options["output"], template=True)
            self.stdout.write(self.style.SUCCESS(f"Modèle XLSX généré : {output}"))
            return

        if action == "export":
            output = export_xlsx(options["output"], template=False)
            self.stdout.write(self.style.SUCCESS(f"Export XLSX généré : {output}"))
            return

        if action == "import":
            input_path = Path(options["input"])
            if not input_path.exists():
                raise CommandError(f"Fichier introuvable : {input_path}")

            created, updated, m2m = import_xlsx(input_path, dry_run=options["dry_run"])
            suffix = " — DRY RUN, aucune donnée écrite" if options["dry_run"] else ""
            self.stdout.write(self.style.SUCCESS(
                f"Import terminé : {created} créé(s), {updated} mis à jour, {m2m} relation(s) M2M{suffix}"
            ))
            return

        raise CommandError(f"Action inconnue : {action}")
