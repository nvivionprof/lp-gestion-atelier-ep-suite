"""Services métier PedaShop.

Ce fichier contient les règles métier qui ne doivent pas être dispersées dans
les vues. Les commentaires sont volontairement pédagogiques : le code doit aussi
pouvoir être relu par un enseignant ou un élève avancé qui découvre Django.
"""
from __future__ import annotations

import re
import unicodedata
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Dict, Iterable, List, Tuple
from django.db import transaction
from django.utils import timezone
from openpyxl import load_workbook


def next_monthly_code(model, prefix: str) -> str:
    """Génère un code métier du type ``DEM-EL-2604-01``.

    Le compteur est indépendant par préfixe et par mois. Cette stratégie est
    plus lisible qu'un ID technique et correspond aux bons papier d'atelier.
    """
    now = timezone.localdate()
    base = f'{prefix}-{now:%y%m}-'
    last = model.objects.filter(code__startswith=base).order_by('-code').first()
    if not last:
        return f'{base}01'
    try:
        number = int(last.code.rsplit('-', 1)[-1]) + 1
    except Exception:
        number = model.objects.filter(code__startswith=base).count() + 1
    return f'{base}{number:02d}'


# Ancien nom conservé pour compatibilité avec les modules déjà générés.
def next_code(model, prefix: str) -> str:
    return next_monthly_code(model, prefix)


def internal_barcode(reference_interne: str) -> str:
    """Crée un code-barres interne si aucun EAN fabricant n'est disponible."""
    cleaned = ''.join(ch for ch in (reference_interne or '').upper() if ch.isalnum())
    return f'PED-{cleaned or timezone.now().strftime("%Y%m%d%H%M%S")}'


def slug_code(value: str) -> str:
    """Nettoie une valeur pour créer un code produit stable."""
    value = unicodedata.normalize('NFKD', str(value or '')).encode('ascii', 'ignore').decode('ascii')
    value = re.sub(r'[^A-Za-z0-9]+', '-', value.upper()).strip('-')
    return value


def make_product_code(fabricant: str, reference: str) -> str:
    prefix = slug_code(fabricant)[:3] or 'PED'
    ref = slug_code(reference) or timezone.now().strftime('%Y%m%d%H%M%S')
    return f'{prefix}-{ref}'


def as_decimal(value, default='0') -> Decimal:
    """Convertit proprement les nombres issus d'Excel, y compris avec virgule."""
    if value is None or value == '':
        return Decimal(default)
    try:
        return Decimal(str(value).replace(',', '.').strip())
    except (InvalidOperation, AttributeError):
        return Decimal(default)


def as_bool(value) -> bool:
    return str(value or '').strip().lower() in {'1', 'oui', 'yes', 'true', 'vrai', 'x', 'o'}


def normalize_header(value: str) -> str:
    txt = unicodedata.normalize('NFKD', str(value or '')).encode('ascii', 'ignore').decode('ascii')
    return txt.strip().lower().replace('.', '').replace('_', ' ').replace('-', ' ')


COLUMN_ALIASES = {
    'fabricant': ['fabricant', 'marque'],
    'reference_fabricant': ['reference', 'référence', 'ref fabricant', 'reference fabricant', 'référence constructeur', 'ref constructeur'],
    'reference_interne': ['code produit', 'code', 'ref interne', 'reference interne', 'référence interne'],
    'code_ean': ['ean', 'code ean', 'code barre', 'code barres'],
    'photo': ['photo', 'image'],
    'designation': ['designation', 'désignation', 'description', 'nom article'],
    'emplacement': ['emplacement', 'localisation', 'casier'],
    'categorie': ['categorie', 'catégorie'],
    'sous_categorie': ['sous categorie', 'sous catégorie', 'sous-catégorie'],
    'stock_reel': ['qte stock', 'qté stock', 'stock', 'quantite', 'quantité'],
    'stock_reserve_demande': ['qte res', 'qté rés', 'qte reservee', 'qté réservée'],
    'stock_temporairement_sorti': ['qte ret', 'qté ret', 'a retourner', 'à retourner'],
    'qte_ok': ['qte ok', 'qté ok'],
    'qte_use': ['qte use', 'qté usé', 'qte usee', 'qté usée'],
    'stock_hs': ['qte hs', 'qté hs', 'hs'],
    'stock_minimum': ['stock mini', 'stock minimum', 'minimum', 'effective de minimum'],
    'substituable': ['substituable', 'equivalence', 'équivalence'],
    'fournisseur': ['fournisseur'],
    'marche': ['marche', 'marché', 'numero marche', 'numéro marché'],
    'unite': ['unite', 'unité'],
    'prix_coutant': ['prix coutant', 'prix coûtant', 'cout', 'coût'],
    'prix_vente': ['prix de vente', 'pris de vente', 'vente'],
    'tva': ['tva', 't v a'],
    'archive': ['archive', 'archiver', 'archivé'],
}


def detect_columns(headers: Iterable[str]) -> Dict[str, int]:
    normalized = [normalize_header(h) for h in headers]
    mapping = {}
    for field, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            alias_norm = normalize_header(alias)
            if alias_norm in normalized:
                mapping[field] = normalized.index(alias_norm)
                break
    return mapping


def load_import_rows(path: str, sheet_name: str | None = None, limit: int | None = None) -> Tuple[List[dict], dict]:
    """Lit un fichier Excel et renvoie des lignes normalisées.

    Le modèle attendu est celui de ``Base magasin PédaShop.xlsx``. Les alias
    permettent néanmoins d'accepter quelques variantes de nom de colonne.
    """
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], {'sheet': ws.title, 'mapping': {}, 'total_rows': 0, 'sheetnames': wb.sheetnames}
    headers = rows[0]
    mapping = detect_columns(headers)
    result = []
    for idx, row in enumerate(rows[1:], start=2):
        if limit and len(result) >= limit:
            break

        def val(field):
            pos = mapping.get(field)
            return row[pos] if pos is not None and pos < len(row) else ''

        fabricant = str(val('fabricant') or '').strip()
        ref_fab = str(val('reference_fabricant') or '').strip()
        code = str(val('reference_interne') or '').strip() or make_product_code(fabricant, ref_fab)
        designation = str(val('designation') or '').strip() or ref_fab or code
        if not any([fabricant, ref_fab, code, designation]):
            continue
        result.append({
            'line': idx,
            'fabricant': fabricant,
            'reference_fabricant': ref_fab,
            'reference_interne': code,
            'code_ean': str(val('code_ean') or '').strip(),
            'photo': str(val('photo') or '').strip(),
            'designation': designation,
            'description': designation,
            'emplacement': str(val('emplacement') or '').strip() or 'A_DEFINIR',
            'categorie': str(val('categorie') or '').strip(),
            'sous_categorie': str(val('sous_categorie') or '').strip(),
            'stock_reel': as_decimal(val('stock_reel')),
            'stock_reserve_demande': as_decimal(val('stock_reserve_demande')),
            'stock_temporairement_sorti': as_decimal(val('stock_temporairement_sorti')),
            'qte_ok': as_decimal(val('qte_ok')),
            'qte_use': as_decimal(val('qte_use')),
            'stock_hs': as_decimal(val('stock_hs')),
            'stock_minimum': as_decimal(val('stock_minimum')),
            'substituable': as_bool(val('substituable')),
            'fournisseur': str(val('fournisseur') or '').strip(),
            'marche': str(val('marche') or '').strip(),
            'unite': str(val('unite') or 'u').strip() or 'u',
            'prix_coutant': as_decimal(val('prix_coutant')),
            'prix_vente': as_decimal(val('prix_vente')),
            'tva': as_decimal(val('tva'), '20'),
            'archive': as_bool(val('archive')),
        })
    return result, {'sheet': ws.title, 'mapping': mapping, 'total_rows': max(len(rows) - 1, 0), 'sheetnames': wb.sheetnames}


@transaction.atomic
def commit_import(rows: List[dict], magasin, actor=None, check_stock_consistency: bool = False) -> dict:
    """Crée les articles et leurs stocks par magasin.

    La règle demandée est stricte : si le code produit existe déjà, la ligne est
    bloquée. On évite ainsi d'écraser silencieusement une fiche article existante.
    """
    from .models import Article, Emplacement, StockArticleMagasin, MouvementStock
    report = {'created_articles': 0, 'updated_articles': 0, 'created_stocks': 0, 'updated_stocks': 0, 'errors': [], 'warnings': []}
    for row in rows:
        ref = row.get('reference_interne')
        if not ref:
            report['errors'].append(f"Ligne {row.get('line')}: code produit manquant")
            continue
        if Article.objects.filter(reference_interne=ref).exists():
            report['errors'].append(f"Ligne {row.get('line')}: code produit {ref} déjà existant. Ligne non importée.")
            continue

        if check_stock_consistency:
            detail = row.get('qte_ok', 0) + row.get('qte_use', 0) + row.get('stock_hs', 0)
            if detail and detail != row.get('stock_reel'):
                report['warnings'].append(
                    f"Ligne {row.get('line')}: incohérence stock. Stock={row.get('stock_reel')} ; détail OK+Usé+HS={detail}."
                )

        article = Article.objects.create(
            reference_interne=ref,
            reference_fabricant=row.get('reference_fabricant', ''),
            fabricant=row.get('fabricant', ''),
            designation=row.get('designation') or ref,
            description=row.get('description', ''),
            code_ean=row.get('code_ean', ''),
            unite=row.get('unite') or 'u',
            categorie=row.get('categorie', ''),
            sous_categorie=row.get('sous_categorie', ''),
            prix_coutant=row.get('prix_coutant') or 0,
            prix_vente=row.get('prix_vente') or 0,
            tva=row.get('tva') or 20,
            substituable=bool(row.get('substituable')),
            fournisseur=row.get('fournisseur', ''),
            marche=row.get('marche', ''),
            archive=bool(row.get('archive')),
        )
        report['created_articles'] += 1
        location_code = row.get('emplacement') or 'A_DEFINIR'
        emplacement, _ = Emplacement.objects.get_or_create(
            magasin=magasin,
            code=location_code,
            defaults={'nom': location_code, 'description': 'Créé automatiquement lors de l’import Excel.'}
        )
        stock = StockArticleMagasin.objects.create(
            article=article,
            magasin=magasin,
            emplacement=emplacement,
            stock_reel=row.get('stock_reel') or 0,
            stock_minimum=row.get('stock_minimum') or 0,
            stock_reserve_demande=row.get('stock_reserve_demande') or 0,
            stock_temporairement_sorti=row.get('stock_temporairement_sorti') or 0,
            qte_ok=row.get('qte_ok') or 0,
            qte_use=row.get('qte_use') or 0,
            stock_hs=row.get('stock_hs') or 0,
        )
        report['created_stocks'] += 1
        MouvementStock.objects.create(
            article=article,
            magasin_destination=magasin,
            emplacement_destination=emplacement,
            type_mouvement='entree_initiale',
            quantite=stock.stock_reel,
            stock_avant=0,
            stock_apres=stock.stock_reel,
            utilisateur=actor,
            commentaire='Import Excel PedaShop V1.7',
        )
    recalculate_stock_alerts()
    return report


def stock_alert_payload(stock) -> tuple[str, str]:
    status = stock.alert_status
    messages = {
        'OK': 'Stock OK.',
        'SOUS_STOCK_MINI': 'Stock disponible sous le stock minimum.',
        'ZERO_PAR_RESERVATION': 'Stock disponible à zéro car engagé en réservation ou préparation.',
        'RUPTURE_TEMPORAIRE_RETOUR_PREVU': 'Stock magasin à zéro, mais un retour est prévu.',
        'RUPTURE_REELLE': 'Rupture réelle : aucun stock en magasin et aucun retour prévu.',
        'STOCK_NEGATIF_AVEC_RESERVATION': 'Sur-réservation : les engagements dépassent le stock disponible.',
    }
    return status, messages.get(status, '')


def recalculate_stock_alerts():
    """Recalcule la table des alertes à partir des stocks courants."""
    from .models import StockAlert, StockArticleMagasin
    count = 0
    for stock in StockArticleMagasin.objects.select_related('article', 'magasin'):
        status, message = stock_alert_payload(stock)
        StockAlert.objects.update_or_create(
            article=stock.article,
            magasin=stock.magasin,
            defaults={
                'statut_alerte': status,
                'type_alerte': status,
                'stock_disponible': stock.stock_disponible,
                'stock_reel': stock.stock_reel,
                'stock_reserve': stock.reserve_total,
                'stock_preparation': stock.stock_en_preparation,
                'stock_exterieur': stock.stock_temporairement_sorti,
                'stock_mini': stock.stock_minimum,
                'message': message,
            }
        )
        count += 1
    return count


def affect_line_to_projection(line):
    """Essaie de prélever une ligne de demande sur une projection prof/TP.

    Cette fonction évite la double réservation : la demande élève consomme le
    solde de pré-réservation du professeur si le professeur, le TP, le magasin et
    l'article correspondent.
    """
    from .models import LigneProjectionPedagogique, StockArticleMagasin
    if not line.bon.professeur_responsable or not line.bon.nom_tp:
        return Decimal('0')
    candidates = LigneProjectionPedagogique.objects.select_related('projection').filter(
        projection__professeur=line.bon.professeur_responsable,
        projection__nom_tp__iexact=line.bon.nom_tp,
        projection__magasin=line.bon.magasin,
        projection__statut__in=['validee', 'en_cours_utilisation', 'partiellement_consommee'],
        article=line.article,
        projection__date_debut__lte=timezone.localdate(),
        projection__date_fin__gte=timezone.localdate(),
    ).order_by('projection__date_fin')
    remaining = line.quantite_demandee
    taken = Decimal('0')
    for lp in candidates:
        if remaining <= 0:
            break
        available = lp.quantite_restante
        if available <= 0:
            continue
        q = min(remaining, available)
        lp.quantite_affectee_aux_demandes += q
        lp.save(update_fields=['quantite_affectee_aux_demandes'])
        line.projection_liee = lp.projection
        line.ligne_projection_liee = lp
        line.quantite_prelevee_sur_projection += q
        taken += q
        remaining -= q
    if taken:
        line.save(update_fields=['projection_liee', 'ligne_projection_liee', 'quantite_prelevee_sur_projection'])
        try:
            stock = StockArticleMagasin.objects.get(article=line.article, magasin=line.bon.magasin)
            stock.stock_reserve_projection = max(Decimal('0'), stock.stock_reserve_projection - taken)
            stock.stock_reserve_demande += remaining
            stock.save()
        except StockArticleMagasin.DoesNotExist:
            pass
    return taken



@transaction.atomic
def commit_import_advanced(rows: List[dict], magasin, actor=None, check_stock_consistency: bool = False,
                           mode: str = 'append_only', key_field: str = 'reference_interne',
                           ignore_blank: bool = True) -> dict:
    """Import PedaShop multi-mode.

    Modes :
    - append_only : ajoute uniquement les absents, ne modifie pas l'existant.
    - upsert : met à jour selon une clé et ajoute les absents.
    - replace_all : remplace la base articles/stocks PedaShop, à utiliser avec confirmation côté vue.
    - simulation : calcule le rapport sans écrire.
    """
    from .models import Article, Emplacement, StockArticleMagasin, MouvementStock
    allowed_keys = {'reference_interne', 'reference_fabricant', 'code_ean', 'designation'}
    if key_field not in allowed_keys:
        key_field = 'reference_interne'
    dry_run = mode == 'simulation'
    report = {
        'mode': mode,
        'key_field': key_field,
        'created_articles': 0,
        'updated_articles': 0,
        'skipped_articles': 0,
        'created_stocks': 0,
        'updated_stocks': 0,
        'deleted_articles': 0,
        'errors': [],
        'warnings': [],
    }
    if mode == 'replace_all' and not dry_run:
        # On ne supprime plus les articles : l'historique PedaShop référence les articles
        # via des clés protégées, notamment MouvementStock.article.
        # Le remplacement total archive les anciens articles, vide les stocks opérationnels,
        # puis réactive/met à jour les articles présents dans le fichier importé.
        report['deleted_articles'] = Article.objects.filter(archive=False).count()
        report['warnings'].append(
            "Remplacement total : historique conservé ; anciens articles archivés et stocks reconstruits depuis le fichier."
        )
        StockArticleMagasin.objects.all().delete()
        Article.objects.update(archive=True)

    article_fields = [
        'reference_fabricant', 'fabricant', 'designation', 'description', 'code_ean', 'unite',
        'categorie', 'sous_categorie', 'prix_coutant', 'prix_vente', 'tva', 'substituable',
        'fournisseur', 'marche', 'archive'
    ]
    for row in rows:
        ref = row.get('reference_interne')
        if not ref:
            report['errors'].append(f"Ligne {row.get('line')}: code produit manquant")
            continue
        key_value = row.get(key_field) or ref
        qs = Article.objects.filter(**{key_field: key_value})
        if qs.count() > 1:
            report['warnings'].append(f"Ligne {row.get('line')}: clé {key_field}={key_value} non unique ; premier article utilisé.")
        article = qs.first()
        exists = article is not None

        if exists and mode == 'append_only':
            report['skipped_articles'] += 1
            continue

        if check_stock_consistency:
            detail = row.get('qte_ok', 0) + row.get('qte_use', 0) + row.get('stock_hs', 0)
            if detail and detail != row.get('stock_reel'):
                report['warnings'].append(f"Ligne {row.get('line')}: incohérence stock. Stock={row.get('stock_reel')} ; OK+Usé+HS={detail}.")

        if not exists:
            if dry_run:
                report['created_articles'] += 1
                continue
            article = Article(reference_interne=ref)
            for f in article_fields:
                setattr(article, f, row.get(f, getattr(article, f, '')))
            article.designation = article.designation or ref
            article.archive = False
            article.save()
            report['created_articles'] += 1
        else:
            if dry_run:
                report['updated_articles'] += 1
            else:
                changed = []
                for f in article_fields:
                    value = row.get(f)
                    if ignore_blank and value in ['', None]:
                        continue
                    if value is not None and getattr(article, f) != value:
                        setattr(article, f, value)
                        changed.append(f)
                if article.archive:
                    article.archive = False
                    changed.append('archive')
                if changed:
                    article.save()
                report['updated_articles'] += 1

        if dry_run:
            continue
        location_code = row.get('emplacement') or 'A_DEFINIR'
        emplacement, _ = Emplacement.objects.get_or_create(
            magasin=magasin,
            code=location_code,
            defaults={'nom': location_code, 'description': 'Créé automatiquement lors de l’import Excel.'}
        )
        stock, created_stock = StockArticleMagasin.objects.get_or_create(article=article, magasin=magasin)
        before = stock.stock_reel
        stock.emplacement = emplacement
        stock.stock_reel = row.get('stock_reel') or 0
        stock.stock_minimum = row.get('stock_minimum') or 0
        stock.stock_reserve_demande = row.get('stock_reserve_demande') or 0
        stock.stock_temporairement_sorti = row.get('stock_temporairement_sorti') or 0
        stock.qte_ok = row.get('qte_ok') or 0
        stock.qte_use = row.get('qte_use') or 0
        stock.stock_hs = row.get('stock_hs') or 0
        stock.save()
        report['created_stocks' if created_stock else 'updated_stocks'] += 1
        if before != stock.stock_reel:
            MouvementStock.objects.create(
                article=article,
                magasin_destination=magasin,
                emplacement_destination=emplacement,
                type_mouvement='import_excel',
                quantite=stock.stock_reel - before,
                stock_avant=before,
                stock_apres=stock.stock_reel,
                utilisateur=actor,
                commentaire=f'Import Excel mode={mode} clé={key_field}',
            )
    return report
