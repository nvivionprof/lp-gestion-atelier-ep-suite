from __future__ import annotations

from decimal import Decimal
from pathlib import Path
from io import BytesIO
from django.conf import settings
from django.contrib import messages
from django.db import transaction
from django.db.models import Count, Q, Sum
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import csrf_exempt
from openpyxl import Workbook
from .exports import pdf_response_bytes, supplier_consultation_pdf_bytes
from .forms import (
    AlertFilterForm, ArticleForm, ArticleSearchForm, BonHeaderForm, EmplacementForm, ExcelImportForm,
    LigneBonForm, LigneProjectionForm, LoginForm, MagasinForm, ProjectionForm, ReclamationForm,
    StockForm, SupplierConsultationForm, TransferForm, StockEntryForm, InventoryAdjustmentForm, UserVisibilityForm,
)
from .models import (
    Article, Bon, BonHistorique, Emplacement, LigneBon, LigneProjectionPedagogique, Magasin,
    MouvementStock, PedaShopUser, ProjectionPedagogique, Reclamation, Reservation, RetourAttendu,
    StockAlert, StockArticleMagasin, SupplierConsultation, SupplierConsultationLine,
)
from .permissions import current_user, require_admin, require_login, require_storekeeper
from .services import affect_line_to_projection, commit_import, load_import_rows, recalculate_stock_alerts
from .sync import sync_users_from_lp_core


# ---------------------------------------------------------------------------
# Outils internes de vue
# ---------------------------------------------------------------------------

def _active_role(request) -> str:
    """Retourne le mode de connexion choisi : utilisateur ou magasinier."""
    return request.session.get('pedashop_active_role', 'utilisateur')


def _can_act_as_storekeeper(request, user) -> bool:
    """Un utilisateur peut voir beaucoup de choses, mais les actions magasin
    exigent le mode magasinier actif + le droit magasinier.
    """
    return bool(user and user.is_storekeeper_like and _active_role(request) == 'magasinier')


def _log_bon(bon, action, user=None, old='', new='', comment='', line=None):
    BonHistorique.objects.create(bon=bon, ligne_bon=line, action=action, ancien_statut=old, nouveau_statut=new, utilisateur=user, commentaire=comment)


def _filter_stocks(qs, data):
    if data.get('magasin'):
        qs = qs.filter(magasin=data['magasin'])
    if data.get('fabricant'):
        qs = qs.filter(article__fabricant__icontains=data['fabricant'])
    if data.get('categorie'):
        qs = qs.filter(article__categorie__icontains=data['categorie'])
    if data.get('sous_categorie'):
        qs = qs.filter(article__sous_categorie__icontains=data['sous_categorie'])
    if data.get('marche'):
        qs = qs.filter(article__marche__icontains=data['marche'])
    if data.get('q'):
        q = data['q']
        qs = qs.filter(Q(article__reference_interne__icontains=q) | Q(article__reference_fabricant__icontains=q) | Q(article__designation__icontains=q) | Q(article__code_ean__icontains=q))
    if data.get('disponible'):
        qs = [s for s in qs if s.stock_disponible > 0]
    if data.get('substituable'):
        qs = qs.filter(article__substituable=True)
    return qs


# ---------------------------------------------------------------------------
# Connexion / API internes
# ---------------------------------------------------------------------------

def health(request):
    return JsonResponse({'status': 'ok', 'module': 'pedashop', 'time': timezone.now().isoformat()})




# --- LP Suite SSO portal-login V0.1.0 ---
def _portal_token_payload(request):
    from django.core import signing
    try:
        return signing.loads(
            request.GET.get('token') or '',
            key=getattr(settings, 'LP_CORE_API_TOKEN', ''),
            salt='lp-suite-sso',
            max_age=120,
        )
    except Exception:
        return None

def portal_login(request):
    payload = _portal_token_payload(request)
    if not payload:
        messages.error(request, 'Connexion LP Core impossible ou expirée. Merci de te reconnecter.')
        return redirect('pedashop_login')
    code = (payload.get('code') or '').strip()
    username = (payload.get('username') or '').strip()
    user = PedaShopUser.objects.filter(Q(code=code) | Q(username=username), active=True).first()
    if not user:
        messages.error(request, 'Compte LP Core non synchronisé dans PedaShop.')
        return redirect('pedashop_login')
    request.session['pedashop_user_id'] = user.id
    request.session['pedashop_active_role'] = 'magasinier' if user.is_storekeeper_like else 'utilisateur'
    messages.success(request, f'Connexion PedaShop via LP Core : {user.first_name} {user.last_name}.')
    return redirect('pedashop_dashboard')

def login_view(request):
    if request.method == 'POST':
        form = LoginForm(request.POST)
        if form.is_valid():
            username = form.cleaned_data['username']
            password = form.cleaned_data['password']
            mode = form.cleaned_data['mode']
            user = PedaShopUser.objects.filter(Q(username=username) | Q(code=username), active=True).first()
            if user and user.check_password(password):
                if mode == 'magasinier' and not user.is_storekeeper_like:
                    messages.error(request, 'Ce compte n’a pas le droit magasinier PedaShop.')
                else:
                    request.session['pedashop_user_id'] = user.id
                    request.session['pedashop_active_role'] = mode
                    messages.success(request, f'Connexion PedaShop en mode {mode}.')
                    return redirect('pedashop_dashboard')
            else:
                messages.error(request, 'Identifiant ou mot de passe incorrect.')
    else:
        form = LoginForm()
    return render(request, 'pedashop/login.html', {'form': form})


def logout_view(request):
    request.session.pop('pedashop_user_id', None)
    request.session.pop('pedashop_active_role', None)
    return redirect('pedashop_login')


@csrf_exempt
@require_http_methods(['POST'])
def internal_sync_lp_core(request):
    """Endpoint interne appelé par LP Core pour synchroniser PedaShop.

    La route est exemptée de CSRF car elle est appelée entre conteneurs Docker,
    pas depuis un formulaire navigateur. La protection repose sur le jeton
    partagé LP_CORE_API_TOKEN.
    """
    expected = getattr(settings, 'LP_CORE_API_TOKEN', '')
    provided = request.headers.get('X-API-Key') or request.POST.get('token') or request.GET.get('token') or ''
    if expected and provided != expected:
        return JsonResponse({'ok': False, 'error': 'Jeton API interne invalide'}, status=403)
    force_password = request.POST.get('force_password') in {'1', 'true', 'True', 'oui', 'OUI'}
    core_user_id = request.POST.get('core_user_id') or request.GET.get('core_user_id')
    report = sync_users_from_lp_core(reset_passwords=force_password, timeout=90, core_user_id=core_user_id)
    return JsonResponse({'ok': True, 'report': report})


def sync_users_view(request):
    user = require_admin(request)
    if not user:
        return redirect('pedashop_login')
    report = sync_users_from_lp_core()
    messages.success(request, f"Synchronisation LP Core → PedaShop : {report}")
    return redirect('pedashop_dashboard')


ACTIVE_DISPLAY_STATUSES = ['en_demande', 'en_cours_traitement', 'en_preparation', 'demande_prete', 'commande_prete', 'en_attente_enlevement', 'retour_attendu', 'retour_partiel', 'reclamation_ouverte']


def _display_name(user):
    if not user:
        return '—'
    return f'{user.first_name} {user.last_name}'.strip() or user.username or user.code



def _visible_magasins(user):
    """Magasins visibles pour l'utilisateur connecté.

    Un compte sans restriction explicite voit tous les magasins actifs. Ce choix
    évite de bloquer une installation existante après migration ; les admins
    peuvent ensuite affecter des magasins visibles dans les fiches utilisateurs.
    """
    qs = Magasin.objects.filter(actif=True)
    if user and not user.is_admin_like:
        ids = list(user.magasins_visibles.values_list('id', flat=True))
        if ids:
            qs = qs.filter(id__in=ids)
    return qs


def _filter_visible_stocks(qs, user):
    if not user or user.is_admin_like:
        return qs
    ids = list(user.magasins_visibles.values_list('id', flat=True))
    return qs.filter(magasin_id__in=ids) if ids else qs


def _can_view_all_bons(request, user):
    if not user:
        return False
    return user.is_admin_like or user.is_teacher_like or (user.is_storekeeper_like and _active_role(request) == 'magasinier')


def _visible_bons(request, qs, user):
    """Filtre les bons selon les droits métier.

    - public : seulement les bons actifs visibles au tableau de bord ;
    - magasinier/prof/admin : tous les bons ;
    - utilisateur : ses bons ou ceux de sa classe/groupe.
    """
    if not user:
        return qs.filter(statut__in=ACTIVE_DISPLAY_STATUSES)
    if _can_view_all_bons(request, user):
        return qs
    filters = Q(demandeur=user)
    if user.class_name:
        filters |= Q(classe_ou_groupe__iexact=user.class_name)
    if user.group_name:
        filters |= Q(classe_ou_groupe__iexact=user.group_name)
    return qs.filter(filters)


def _can_show_all_actions(user):
    return bool(user and (user.is_teacher_like or user.is_admin_like))


def _can_edit_articles(request, user):
    """Modification d'article autorisée aux admins ou élèves magasiniers habilités.

    Le droit temporaire peut être donné depuis LP Core dans le champ rights :
    PEDASHOP_ARTICLE_EDIT ou PEDASHOP_TEMP_ARTICLE_EDIT.
    """
    if not user:
        return False
    rights = set(user.rights_list())
    return user.is_admin_like or bool({'PEDASHOP_ARTICLE_EDIT', 'PEDASHOP_TEMP_ARTICLE_EDIT'} & rights)


def _status_class(status):
    if status == 'reclamation_ouverte':
        return 'orange'
    if status in {'annulee'}:
        return 'red'
    if status in {'commande_prete', 'demande_prete', 'en_attente_enlevement'}:
        return 'green'
    if status in {'en_cours_traitement', 'en_preparation'}:
        return 'blue'
    if status in {'retour_attendu', 'retour_partiel'}:
        return 'yellow'
    if status in {'distribuee'}:
        return 'dark'
    return 'blue'


def _bon_has_return_remaining(bon):
    for retour in RetourAttendu.objects.filter(ligne_bon__bon=bon):
        processed = retour.quantite_retournee + retour.quantite_usee + retour.quantite_cassee + retour.quantite_perdue
        if processed < retour.quantite_attendue:
            return True
    return False


def _make_articles_template_workbook():
    wb = Workbook()
    ws = wb.active
    ws.title = 'Base PedaShop'
    headers = ['Fabricant', 'reference', 'code produit', 'EAN', 'Photo', 'Désignation', 'Emplacement', 'Catégorie', 'Sous Catégorie', 'Qté stock', 'Qté rés', 'Qté ret', 'Qté OK', 'Qté Usé', 'Qté HS', 'Stock Mini', 'Substituable', 'Fournisseur', 'Marché']
    ws.append(headers)
    ws.append(['Schneider', 'A9F74216', 'SCH-A9F74216', 'SCH-A9F74216', 'SCH-A9F74216.jpg', 'Disjoncteur 16A courbe C', 'A01-E02', 'Protection', 'Disjoncteurs', 10, 0, 0, 10, 0, 0, 5, 'Oui', 'Fournisseur exemple', 'MAPA-2026-01'])
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = min(max(len(str(col[0].value or '')) + 4, 14), 32)
    return wb


def _xlsx_response(wb, filename):
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ---------------------------------------------------------------------------
# Tableaux de bord et consultation globale
# ---------------------------------------------------------------------------

def dashboard(request):
    """Tableau de bord public.

    La page d'accueil PedaShop doit être consultable sans connexion. Les
    indicateurs et les bons visibles restent affichés, mais les boutons d'action
    sont masqués tant qu'aucun rôle n'est actif.
    """
    user = current_user(request)
    recalculate_stock_alerts()
    all_bons = Bon.objects.select_related('magasin', 'demandeur', 'professeur_responsable')
    visible_bons = _visible_bons(request, all_bons, user)
    cards = {
        'demandes': visible_bons.filter(statut='en_demande').count(),
        'preparation': visible_bons.filter(statut__in=['en_cours_traitement', 'en_preparation']).count(),
        'pretes': visible_bons.filter(statut__in=['demande_prete', 'commande_prete', 'en_attente_enlevement']).count(),
        'distribuees_a_inventorier': visible_bons.filter(statut='distribuee').count(),
        'retours': RetourAttendu.objects.exclude(statut__in=['retourne', 'perdu']).count(),
        'reclamations': Reclamation.objects.exclude(statut='cloturee').count(),
        'alertes_rouges': StockAlert.objects.filter(statut_alerte__in=['RUPTURE_REELLE', 'STOCK_NEGATIF_AVEC_RESERVATION']).count(),
        'alertes_orange': StockAlert.objects.filter(statut_alerte__in=['ZERO_PAR_RESERVATION', 'RUPTURE_TEMPORAIRE_RETOUR_PREVU']).count(),
        'alertes_jaunes': StockAlert.objects.filter(statut_alerte='SOUS_STOCK_MINI').count(),
    }
    context = {
        'cards': cards,
        'user': user,
        'role': _active_role(request),
        'is_public': user is None,
        'can_all_actions': _can_show_all_actions(user),
        'can_storekeep': _can_act_as_storekeeper(request, user),
        'can_simple_request': bool(user and not _can_act_as_storekeeper(request, user)),
        'bons_a_traiter': visible_bons.filter(statut__in=ACTIVE_DISPLAY_STATUSES + ['distribuee'])[:20],
        'bons_distribues': visible_bons.filter(statut='distribuee')[:10],
        'alertes': StockAlert.objects.exclude(statut_alerte='OK').select_related('article', 'magasin')[:12],
    }
    return render(request, 'pedashop/dashboard.html', context)

def affichage_dynamique(request):
    """Affichage atelier : uniquement les bons et le demandeur."""
    bons_qs = Bon.objects.filter(statut__in=ACTIVE_DISPLAY_STATUSES).select_related('magasin', 'demandeur', 'professeur_responsable')
    grouped = {
        'Demandes en attente': bons_qs.filter(statut='en_demande'),
        'En cours de traitement': bons_qs.filter(statut__in=['en_cours_traitement', 'en_preparation']),
        'Commandes prêtes': bons_qs.filter(statut__in=['commande_prete', 'demande_prete', 'en_attente_enlevement']),
        'Retours attendus': bons_qs.filter(statut__in=['retour_attendu', 'retour_partiel']),
        'Réclamations ouvertes': bons_qs.filter(statut='reclamation_ouverte'),
    }
    context = {'grouped_bons': grouped, 'cards': {key: value.count() for key, value in grouped.items()}}
    return render(request, 'pedashop/affichage.html', context)


# ---------------------------------------------------------------------------
# Articles / stocks / mouvements
# ---------------------------------------------------------------------------

def article_list(request):
    user = current_user(request)
    qs = Article.objects.all()
    q = request.GET.get('q', '')
    cat = request.GET.get('categorie', '')
    sous = request.GET.get('sous_categorie', '')
    if q:
        qs = qs.filter(Q(reference_interne__icontains=q) | Q(reference_fabricant__icontains=q) | Q(designation__icontains=q) | Q(fabricant__icontains=q) | Q(code_ean__icontains=q))
    if cat:
        qs = qs.filter(categorie__icontains=cat)
    if sous:
        qs = qs.filter(sous_categorie__icontains=sous)
    return render(request, 'pedashop/article_list.html', {'items': qs[:500], 'q': q, 'categorie': cat, 'sous_categorie': sous, 'can_edit_articles': _can_edit_articles(request, user)})

@require_http_methods(['GET', 'POST'])
def article_form(request, pk=None):
    user = require_login(request)
    if not user or not _can_edit_articles(request, user):
        messages.error(request, 'Modification article réservée aux droits PEDASHOP_ARTICLE_EDIT ou administrateur.')
        return redirect('pedashop_login' if not user else 'pedashop_article_list')
    obj = get_object_or_404(Article, pk=pk) if pk else None
    form = ArticleForm(request.POST or None, request.FILES or None, instance=obj)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'Article enregistré.')
        return redirect('pedashop_article_list')
    return render(request, 'pedashop/form.html', {'form': form, 'title': 'Article'})

def article_detail(request, pk):
    user = current_user(request)
    article = get_object_or_404(Article, pk=pk)
    stocks = _filter_visible_stocks(article.stocks.select_related('magasin', 'emplacement'), user)
    retours = RetourAttendu.objects.filter(article=article).exclude(statut__in=['retourne', 'perdu']).select_related('ligne_bon__bon', 'magasin_retour_prevu')
    return render(request, 'pedashop/article_detail.html', {'article': article, 'stocks': stocks, 'retours': retours, 'can_edit_articles': _can_edit_articles(request, user)})

def magasin_list(request):
    user = require_login(request)
    if not user:
        return redirect('pedashop_login')
    return render(request, 'pedashop/magasin_list.html', {'items': Magasin.objects.all()})


@require_http_methods(['GET', 'POST'])
def magasin_form(request, pk=None):
    user = require_admin(request)
    if not user:
        return redirect('pedashop_login')
    obj = get_object_or_404(Magasin, pk=pk) if pk else None
    form = MagasinForm(request.POST or None, instance=obj)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'Magasin enregistré.')
        return redirect('pedashop_magasin_list')
    return render(request, 'pedashop/form.html', {'form': form, 'title': 'Magasin'})


def emplacement_list(request):
    user = require_login(request)
    if not user:
        return redirect('pedashop_login')
    return render(request, 'pedashop/emplacement_list.html', {'items': Emplacement.objects.select_related('magasin')})


@require_http_methods(['GET', 'POST'])
def emplacement_form(request):
    user = require_admin(request)
    if not user:
        return redirect('pedashop_login')
    form = EmplacementForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'Emplacement enregistré.')
        return redirect('pedashop_emplacement_list')
    return render(request, 'pedashop/form.html', {'form': form, 'title': 'Emplacement'})


def stock_list(request):
    """Consultation du stock accessible sans connexion.

    Deux vues sont proposées : par magasin ou par référence cumulée tous stocks
    confondus. Le détail magasin reste disponible dans la fiche article.
    """
    user = current_user(request)
    allowed = _visible_magasins(user)
    form = ArticleSearchForm(request.GET or None, allowed_magasins=allowed)
    mode = request.GET.get('mode', 'magasin')
    qs = StockArticleMagasin.objects.select_related('article', 'magasin', 'emplacement').all()
    qs = _filter_visible_stocks(qs, user)
    if form.is_valid():
        qs = _filter_stocks(qs, form.cleaned_data)
    if mode == 'reference':
        totals = {}
        for s0 in qs:
            a = s0.article
            row = totals.setdefault(a.id, {'article': a, 'stock_reel': Decimal('0'), 'reserve_total': Decimal('0'), 'prepa': Decimal('0'), 'exterieur': Decimal('0'), 'dispo': Decimal('0'), 'mini': Decimal('0'), 'hs': Decimal('0'), 'perdu': Decimal('0')})
            row['stock_reel'] += s0.stock_reel
            row['reserve_total'] += s0.reserve_total
            row['prepa'] += s0.stock_en_preparation
            row['exterieur'] += s0.stock_temporairement_sorti
            row['dispo'] += s0.stock_disponible
            row['mini'] += s0.stock_minimum
            row['hs'] += s0.stock_hs
            row['perdu'] += getattr(s0, 'stock_perdu', Decimal('0'))
        return render(request, 'pedashop/stock_list.html', {'items': qs, 'totals': totals.values(), 'form': form, 'mode': mode, 'user': user})
    return render(request, 'pedashop/stock_list.html', {'items': qs, 'form': form, 'mode': mode, 'user': user})

def movement_list(request):
    user = require_login(request)
    if not user:
        return redirect('pedashop_login')
    qs = MouvementStock.objects.select_related('article', 'magasin_source', 'magasin_destination', 'utilisateur')[:500]
    return render(request, 'pedashop/movement_list.html', {'items': qs})


# ---------------------------------------------------------------------------
# Bons / panier multi-articles / préparation
# ---------------------------------------------------------------------------

def bon_list(request):
    user = require_login(request)
    if not user:
        return redirect('pedashop_login')
    qs = Bon.objects.select_related('magasin', 'demandeur', 'professeur_responsable').all()
    qs = _visible_bons(request, qs, user)
    statut = request.GET.get('statut')
    if statut:
        qs = qs.filter(statut=statut)
    for b in qs:
        b.status_class = _status_class(b.statut)
    return render(request, 'pedashop/bon_list.html', {'items': qs, 'statut': statut, 'can_create_bon': bool(user and (user.is_teacher_like or user.is_admin_like or _active_role(request) != 'magasinier'))})

def _get_cart(request):
    return request.session.setdefault('pedashop_bon_cart', [])


@require_http_methods(['GET', 'POST'])
def bon_create(request):
    user = require_login(request)
    if not user:
        return redirect('pedashop_login')
    allowed = _visible_magasins(user)
    cart = _get_cart(request)
    header_initial = request.session.get('pedashop_bon_header', {})
    initial_type = request.GET.get('type_bon') or header_initial.get('type_bon')
    header_form = BonHeaderForm(None, initial=header_initial, user=user, active_role=_active_role(request), allowed_magasins=allowed, initial_type=initial_type)
    search_form = ArticleSearchForm(request.GET or None, allowed_magasins=allowed)
    results = _filter_visible_stocks(StockArticleMagasin.objects.select_related('article', 'magasin', 'emplacement'), user)
    if search_form.is_valid():
        results = _filter_stocks(results, search_form.cleaned_data)
    results = results[:50] if hasattr(results, '__getitem__') else results

    if request.method == 'POST':
        action = request.POST.get('action')
        # Les champs d'en-tête sont conservés en session afin que le bouton
        # "Ajouter" ne vide pas le demandeur, le TP ou le professeur responsable.
        header_keys = ['type_bon', 'magasin', 'professeur_responsable', 'nom_tp', 'classe_ou_groupe', 'commentaire']
        posted_header = {k: request.POST.get(k, '') for k in header_keys if k in request.POST}
        if posted_header:
            request.session['pedashop_bon_header'] = posted_header
            header_initial = posted_header
        if action == 'add_line':
            article = get_object_or_404(Article, pk=request.POST.get('article_id'))
            qty = Decimal(request.POST.get('quantite') or '1')
            type_sortie = request.POST.get('type_sortie', 'definitive')
            date_retour = request.POST.get('date_retour_prevue') or ''
            if type_sortie == 'temporaire' and not date_retour:
                messages.error(request, 'La date de retour est obligatoire pour une sortie temporaire.')
            else:
                cart.append({
                    'article_id': article.id,
                    'article_label': f'{article.reference_interne} — {article.designation}',
                    'quantite': str(qty),
                    'type_sortie': type_sortie,
                    'date_retour_prevue': date_retour,
                    'commentaire': request.POST.get('commentaire_ligne') or '',
                })
                request.session.modified = True
                messages.success(request, 'Article ajouté au bon.')
            return redirect('pedashop_bon_create')
        if action == 'clear_cart':
            request.session['pedashop_bon_cart'] = []
            messages.info(request, 'Panier vidé.')
            return redirect('pedashop_bon_create')
        if action == 'submit_bon':
            header_form = BonHeaderForm(request.POST, user=user, active_role=_active_role(request), allowed_magasins=allowed, initial_type=initial_type)
            if header_form.is_valid():
                if not cart:
                    messages.error(request, 'Ajoute au moins un article au bon.')
                else:
                    with transaction.atomic():
                        bon = header_form.save(commit=False)
                        bon.demandeur = user
                        bon.statut = 'en_demande' if bon.type_bon in ['demande_eleve', 'demande_prof'] else 'en_attente_enlevement'
                        bon.save()
                        _log_bon(bon, 'creation', user, new=bon.statut, comment='Création du bon multi-articles.')
                        for row in cart:
                            article = Article.objects.get(pk=row['article_id'])
                            line = LigneBon.objects.create(
                                bon=bon,
                                article=article,
                                quantite_demandee=Decimal(row['quantite']),
                                type_sortie=row['type_sortie'],
                                date_retour_prevue=row['date_retour_prevue'] or None,
                                commentaire=row['commentaire'],
                            )
                            taken = affect_line_to_projection(line)
                            if taken == 0:
                                stock = StockArticleMagasin.objects.get(article=article, magasin=bon.magasin)
                                stock.stock_reserve_demande += line.quantite_demandee
                                stock.save()
                        recalculate_stock_alerts()
                        request.session['pedashop_bon_cart'] = []
                        request.session.pop('pedashop_bon_header', None)
                        messages.success(request, f'Bon créé : {bon.code}.')
                        return redirect('pedashop_bon_detail', pk=bon.pk)
            else:
                messages.error(request, 'En-tête du bon incomplet ou incompatible avec le rôle actif.')
    return render(request, 'pedashop/bon_create.html', {'header_form': header_form, 'search_form': search_form, 'results': results, 'cart': cart})

def bon_detail(request, pk):
    user = require_login(request)
    if not user:
        return redirect('pedashop_login')
    bon = get_object_or_404(Bon.objects.select_related('magasin', 'demandeur', 'professeur_responsable', 'preparateur', 'distributeur', 'receptionnaire'), pk=pk)
    can_storekeep = _can_act_as_storekeeper(request, user)
    return render(request, 'pedashop/bon_detail.html', {
        'bon': bon,
        'can_storekeep': can_storekeep,
        'can_return': can_storekeep and _bon_has_return_remaining(bon),
        'status_class': _status_class(bon.statut),
    })

@require_http_methods(['GET', 'POST'])
@transaction.atomic
def bon_prepare(request, pk):
    """Préparation magasinier ligne par ligne."""
    user = require_storekeeper(request)
    if not user or not _can_act_as_storekeeper(request, user):
        return redirect('pedashop_login')
    bon = get_object_or_404(Bon.objects.select_related('magasin'), pk=pk)
    if bon.statut == 'en_demande':
        old = bon.statut
        bon.statut = 'en_cours_traitement'
        bon.preparateur = user
        bon.date_preparation_debut = timezone.now()
        bon.save(update_fields=['statut', 'preparateur', 'date_preparation_debut', 'updated_at'])
        _log_bon(bon, 'prise_traitement', user, old, bon.statut, 'Demande prise en cours de traitement.')
    if request.method == 'POST':
        action = request.POST.get('action')
        all_ok = True
        for ligne in bon.lignes.select_related('article'):
            checked = request.POST.get(f'ligne_{ligne.id}_preparee') == 'on'
            qty = Decimal(request.POST.get(f'ligne_{ligne.id}_quantite_preparee') or '0')
            if checked:
                if not ligne.est_preparee:
                    stock = StockArticleMagasin.objects.get(article=ligne.article, magasin=bon.magasin)
                    transfer_qty = qty or ligne.quantite_demandee
                    stock.stock_reserve_demande = max(Decimal('0'), stock.stock_reserve_demande - transfer_qty)
                    stock.stock_en_preparation += transfer_qty
                    stock.save()
                ligne.quantite_preparee = qty or ligne.quantite_demandee
                ligne.est_preparee = True
                ligne.preparee_par = user
                ligne.date_preparation = timezone.now()
                ligne.statut_ligne = 'preparee'
                ligne.commentaire = request.POST.get(f'ligne_{ligne.id}_commentaire', ligne.commentaire)
                ligne.save()
                _log_bon(bon, 'ligne_preparee', user, line=ligne, comment=f'Ligne préparée : {ligne.article.reference_interne}')
            if not ligne.est_preparee:
                all_ok = False
        if action == 'ready':
            if all_ok:
                old = bon.statut
                bon.statut = 'commande_prete'
                bon.date_preparation_fin = timezone.now()
                bon.save(update_fields=['statut', 'date_preparation_fin', 'updated_at'])
                _log_bon(bon, 'commande_prete', user, old, bon.statut, 'Toutes les lignes préparées : commande prête.')
                messages.success(request, 'Commande passée en prête.')
                recalculate_stock_alerts()
                return redirect('pedashop_bon_detail', pk=pk)
            messages.error(request, 'Impossible de passer en prête : toutes les lignes ne sont pas cochées.')
        else:
            messages.success(request, 'Préparation enregistrée.')
            recalculate_stock_alerts()
            return redirect('pedashop_bon_prepare', pk=pk)
    return render(request, 'pedashop/bon_prepare.html', {'bon': bon})


@require_http_methods(['GET', 'POST'])
@transaction.atomic
def bon_distribute(request, pk):
    """Remise de la commande au bénéficiaire indiqué par le magasinier."""
    user = require_storekeeper(request)
    if not user or not _can_act_as_storekeeper(request, user):
        return redirect('pedashop_login')
    bon = get_object_or_404(Bon, pk=pk)
    users = PedaShopUser.objects.filter(active=True).order_by('last_name', 'first_name', 'code')
    if request.method == 'POST':
        remis_a = get_object_or_404(PedaShopUser, pk=request.POST.get('remis_a'))
        old = bon.statut
        for ligne in bon.lignes.select_related('article'):
            stock = StockArticleMagasin.objects.get(article=ligne.article, magasin=bon.magasin)
            before = stock.stock_reel
            q = ligne.quantite_preparee or ligne.quantite_demandee
            stock.stock_reel -= q
            stock.stock_en_preparation = max(Decimal('0'), stock.stock_en_preparation - q)
            if ligne.type_sortie == 'temporaire':
                stock.stock_temporairement_sorti += q
                ligne.quantite_retour_prevue = q
            ligne.quantite_distribuee = q
            ligne.statut_ligne = 'distribuee'
            ligne.save()
            stock.save()
            MouvementStock.objects.create(article=ligne.article, magasin_source=bon.magasin, emplacement_source=stock.emplacement, type_mouvement='sortie_temporaire' if ligne.type_sortie == 'temporaire' else 'sortie_definitive', quantite=q, stock_avant=before, stock_apres=stock.stock_reel, utilisateur=user, demandeur=bon.demandeur, preparateur=bon.preparateur, distributeur=user, receptionnaire=remis_a, bon=bon, commentaire=f'Remis à {remis_a}')
        bon.distributeur = user
        bon.receptionnaire = remis_a
        bon.date_enlevement = timezone.now()
        bon.statut = 'distribuee'
        bon.save(update_fields=['distributeur', 'receptionnaire', 'date_enlevement', 'statut', 'updated_at'])
        _log_bon(bon, 'distribution', user, old, bon.statut, f'Commande remise à {remis_a}.')
        recalculate_stock_alerts()
        messages.success(request, 'Distribution enregistrée. Le bon attend maintenant l’inventaire de réception utilisateur.')
        return redirect('pedashop_bon_detail', pk=pk)
    return render(request, 'pedashop/bon_distribute.html', {'bon': bon, 'users': users})


@require_http_methods(['GET', 'POST'])
@transaction.atomic
def bon_reception_inventory(request, pk):
    """Inventaire de réception par l'utilisateur après distribution."""
    user = require_login(request)
    if not user:
        return redirect('pedashop_login')
    bon = get_object_or_404(Bon, pk=pk)
    if request.method == 'POST':
        has_problem = False
        for ligne in bon.lignes.select_related('article'):
            present = request.POST.get(f'ligne_{ligne.id}_presente') == 'on'
            q_reelle = Decimal(request.POST.get(f'ligne_{ligne.id}_quantite_reelle') or '0')
            if (not present) or q_reelle != ligne.quantite_distribuee:
                has_problem = True
                Reclamation.objects.create(bon=bon, ligne_bon=ligne, type_reclamation='quantite_incorrecte' if present else 'materiel_manquant', description=f'Inventaire de réception non conforme. Quantité distribuée : {ligne.quantite_distribuee}, quantité constatée : {q_reelle}.', declaree_par=user, concerne=bon.demandeur)
                ligne.statut_ligne = 'reclamation'
                ligne.save(update_fields=['statut_ligne'])
        old = bon.statut
        if has_problem:
            bon.statut = 'reclamation_ouverte'
            messages.warning(request, 'Inventaire non conforme : réclamation créée.')
        else:
            if bon.has_temporary_lines:
                for ligne in bon.lignes.filter(type_sortie='temporaire').select_related('article'):
                    if ligne.date_retour_prevue and not ligne.retours_attendus.exists():
                        RetourAttendu.objects.create(ligne_bon=ligne, article=ligne.article, magasin_retour_prevu=bon.magasin, quantite_attendue=ligne.quantite_retour_prevue or ligne.quantite_distribuee, date_retour_prevue=ligne.date_retour_prevue)
                bon.statut = 'retour_attendu'
                messages.success(request, 'Inventaire validé. Bon de retour prévu généré pour les lignes temporaires.')
            else:
                bon.statut = 'cloturee'
                bon.date_cloture = timezone.now()
                messages.success(request, 'Inventaire validé. Demande clôturée.')
        bon.save()
        _log_bon(bon, 'inventaire_reception', user, old, bon.statut, 'Inventaire de réception utilisateur.')
        recalculate_stock_alerts()
        return redirect('pedashop_bon_detail', pk=pk)
    return render(request, 'pedashop/bon_reception_inventory.html', {'bon': bon})


@require_http_methods(['GET', 'POST'])
@transaction.atomic
def bon_return(request, pk):
    """Retour partiel ou complet du matériel temporairement sorti.

    Le retour ne doit pas être un bouton global aveugle : le magasinier saisit
    quantité OK, usée, cassée ou perdue par ligne de retour attendu. Les pertes
    et casses sont tracées et déduites du matériel à retourner.
    """
    user = require_storekeeper(request)
    if not user or not _can_act_as_storekeeper(request, user):
        return redirect('pedashop_login')
    bon = get_object_or_404(Bon, pk=pk)
    retours = RetourAttendu.objects.filter(ligne_bon__bon=bon).select_related('ligne_bon__article', 'magasin_retour_prevu')
    if request.method == 'POST':
        old = bon.statut
        for retour in retours:
            ligne = retour.ligne_bon
            remaining = retour.quantite_attendue - retour.quantite_retournee - retour.quantite_usee - retour.quantite_cassee - retour.quantite_perdue
            if remaining <= 0:
                continue
            q_ok = Decimal(request.POST.get(f'retour_{retour.id}_ok') or '0')
            q_use = Decimal(request.POST.get(f'retour_{retour.id}_use') or '0')
            q_hs = Decimal(request.POST.get(f'retour_{retour.id}_hs') or '0')
            q_perdu = Decimal(request.POST.get(f'retour_{retour.id}_perdu') or '0')
            total = q_ok + q_use + q_hs + q_perdu
            if total <= 0:
                continue
            if total > remaining:
                messages.error(request, f'Retour {ligne.article.reference_interne} : quantité saisie supérieure au reste à retourner.')
                return redirect('pedashop_bon_return', pk=pk)
            stock = StockArticleMagasin.objects.get(article=ligne.article, magasin=bon.magasin)
            before = stock.stock_reel
            if q_ok:
                stock.stock_reel += q_ok
            if q_use:
                stock.stock_reel += q_use
                stock.qte_use += q_use
            if q_hs:
                stock.stock_hs += q_hs
            if q_perdu:
                stock.stock_perdu += q_perdu
            stock.stock_temporairement_sorti = max(Decimal('0'), stock.stock_temporairement_sorti - total)
            stock.save()
            ligne.quantite_retournee += q_ok + q_use
            ligne.quantite_usee += q_use
            ligne.quantite_hs += q_hs
            ligne.quantite_perdue += q_perdu
            retour.quantite_retournee += q_ok
            retour.quantite_usee += q_use
            retour.quantite_cassee += q_hs
            retour.quantite_perdue += q_perdu
            if retour.quantite_retournee + retour.quantite_usee + retour.quantite_cassee + retour.quantite_perdue >= retour.quantite_attendue:
                retour.statut = 'retourne' if not (q_hs or q_perdu) else 'perdu'
                retour.date_retour_reelle = timezone.localdate()
                ligne.statut_ligne = 'retournee' if not (q_hs or q_perdu) else 'retour_anomalie'
            else:
                retour.statut = 'partiel'
                ligne.statut_ligne = 'retour_partiel'
            ligne.save(); retour.save()
            MouvementStock.objects.create(article=ligne.article, magasin_destination=bon.magasin, emplacement_destination=stock.emplacement, type_mouvement='retour_temporaire' if (q_ok or q_use) else ('casse' if q_hs else 'perte'), quantite=total, stock_avant=before, stock_apres=stock.stock_reel, utilisateur=user, demandeur=bon.demandeur, receptionnaire=user, bon=bon, commentaire=request.POST.get(f'retour_{retour.id}_commentaire', ''))
        if _bon_has_return_remaining(bon):
            bon.statut = 'retour_partiel'
            messages.success(request, 'Retour partiel enregistré.')
        else:
            bon.statut = 'cloturee'
            bon.date_cloture = timezone.now()
            messages.success(request, 'Tous les retours attendus sont traités. Bon clôturé.')
        bon.receptionnaire = user
        bon.save()
        _log_bon(bon, 'retour', user, old, bon.statut, 'Retour temporaire saisi ligne par ligne.')
        recalculate_stock_alerts()
        return redirect('pedashop_bon_detail', pk=pk)
    return render(request, 'pedashop/bon_return.html', {'bon': bon, 'retours': retours})



# ---------------------------------------------------------------------------
# Projections pédagogiques / réservations prof
# ---------------------------------------------------------------------------

def projection_list(request):
    user = require_login(request)
    if not user:
        return redirect('pedashop_login')
    qs = ProjectionPedagogique.objects.select_related('professeur', 'magasin').all()
    return render(request, 'pedashop/projection_list.html', {'items': qs})


@require_http_methods(['GET', 'POST'])
def projection_create(request):
    user = require_login(request)
    if not user:
        return redirect('pedashop_login')
    proj_form = ProjectionForm(request.POST or None)
    line_form = LigneProjectionForm(request.POST or None)
    if request.method == 'POST' and proj_form.is_valid() and line_form.is_valid():
        with transaction.atomic():
            projection = proj_form.save(commit=False)
            projection.created_by = user
            projection.save()
            line = line_form.save(commit=False)
            line.projection = projection
            line.save()
            stock = StockArticleMagasin.objects.get(article=line.article, magasin=projection.magasin)
            stock.stock_reserve_projection += line.quantite_reservee
            stock.save()
            MouvementStock.objects.create(article=line.article, magasin_source=projection.magasin, type_mouvement='reservation_projection', quantite=line.quantite_reservee, stock_avant=stock.stock_reel, stock_apres=stock.stock_reel, utilisateur=user, commentaire=f'Pré-réservation {projection.code}')
            recalculate_stock_alerts()
        messages.success(request, f'Projection pédagogique créée : {projection.code}.')
        return redirect('pedashop_projection_list')
    return render(request, 'pedashop/projection_create.html', {'proj_form': proj_form, 'line_form': line_form})


# Anciennes réservations conservées.
def reservation_list(request):
    user = require_login(request)
    if not user:
        return redirect('pedashop_login')
    qs = Reservation.objects.select_related('magasin', 'demandeur')
    return render(request, 'pedashop/reservation_list.html', {'items': qs})


# ---------------------------------------------------------------------------
# Réclamations
# ---------------------------------------------------------------------------

@require_http_methods(['GET', 'POST'])
def reclamation_create(request, bon_id=None, pk=None):
    user = require_login(request)
    if not user:
        return redirect('pedashop_login')
    bon = get_object_or_404(Bon, pk=(bon_id or pk))
    form = ReclamationForm(request.POST or None, request.FILES or None)
    form.fields['ligne_bon'].queryset = bon.lignes.all()
    if request.method == 'POST' and form.is_valid():
        rec = form.save(commit=False)
        rec.bon = bon
        rec.declaree_par = user
        rec.concerne = bon.demandeur
        rec.save()
        old = bon.statut
        bon.statut = 'reclamation_ouverte'
        bon.save(update_fields=['statut'])
        _log_bon(bon, 'reclamation', user, old, bon.statut, rec.description)
        messages.success(request, 'Réclamation créée. Elle devra être traitée.')
        return redirect('pedashop_bon_detail', pk=bon.pk)
    return render(request, 'pedashop/form.html', {'form': form, 'title': f'Réclamation — {bon.code}'})


def reclamation_list(request):
    user = require_login(request)
    if not user:
        return redirect('pedashop_login')
    return render(request, 'pedashop/reclamation_list.html', {'items': Reclamation.objects.select_related('bon', 'declaree_par')})


# ---------------------------------------------------------------------------
# Alertes / consultation fournisseur
# ---------------------------------------------------------------------------

def alert_list(request):
    user = require_login(request)
    if not user:
        return redirect('pedashop_login')
    recalculate_stock_alerts()
    form = AlertFilterForm(request.GET or None)
    qs = StockAlert.objects.exclude(statut_alerte='OK').select_related('article', 'magasin')
    if form.is_valid():
        data = form.cleaned_data
        if data.get('magasin'):
            qs = qs.filter(magasin=data['magasin'])
        if data.get('statut_alerte'):
            qs = qs.filter(statut_alerte=data['statut_alerte'])
        if data.get('fabricant'):
            qs = qs.filter(article__fabricant__icontains=data['fabricant'])
        if data.get('categorie'):
            qs = qs.filter(article__categorie__icontains=data['categorie'])
        if data.get('sous_categorie'):
            qs = qs.filter(article__sous_categorie__icontains=data['sous_categorie'])
        if data.get('marche'):
            qs = qs.filter(article__marche__icontains=data['marche'])
        if data.get('q'):
            q = data['q']
            qs = qs.filter(Q(article__reference_interne__icontains=q) | Q(article__reference_fabricant__icontains=q) | Q(article__designation__icontains=q))
        if data.get('substituable') is not None:
            qs = qs.filter(article__substituable=data['substituable'])
    if request.method == 'POST':
        ids = request.POST.getlist('selected_alerts')
        if not ids:
            messages.error(request, 'Aucune alerte sélectionnée.')
            return redirect('pedashop_alert_list')
        with transaction.atomic():
            first = StockAlert.objects.filter(pk__in=ids).select_related('magasin').first()
            consultation = SupplierConsultation.objects.create(created_by=user, magasin=first.magasin if first else None)
            for alert in StockAlert.objects.filter(pk__in=ids).select_related('article'):
                qty = max((alert.stock_mini - alert.stock_disponible), Decimal('1'))
                SupplierConsultationLine.objects.create(
                    consultation=consultation,
                    article=alert.article,
                    designation=alert.article.designation,
                    fabricant=alert.article.fabricant,
                    reference_constructeur=alert.article.reference_fabricant,
                    quantite_souhaitee=qty,
                    equivalence_possible=alert.article.substituable,
                    valeur_substituable_article=alert.article.substituable,
                )
                alert.consultation_generee = True
                alert.save(update_fields=['consultation_generee'])
        messages.success(request, f'Consultation fournisseur créée : {consultation.code}.')
        return redirect('pedashop_consultation_edit', pk=consultation.pk)
    return render(request, 'pedashop/alert_list.html', {'items': qs, 'form': form})


def consultation_list(request):
    user = require_login(request)
    if not user:
        return redirect('pedashop_login')
    return render(request, 'pedashop/consultation_list.html', {'items': SupplierConsultation.objects.select_related('magasin', 'created_by')})


@require_http_methods(['GET', 'POST'])
def consultation_edit(request, pk):
    user = require_login(request)
    if not user:
        return redirect('pedashop_login')
    consultation = get_object_or_404(SupplierConsultation, pk=pk)
    if request.method == 'POST':
        for line in consultation.lignes.all():
            prefix = f'line_{line.id}_'
            line.designation = request.POST.get(prefix + 'designation', line.designation)
            line.fabricant = request.POST.get(prefix + 'fabricant', line.fabricant)
            line.reference_constructeur = request.POST.get(prefix + 'reference_constructeur', line.reference_constructeur)
            line.quantite_souhaitee = Decimal(request.POST.get(prefix + 'quantite_souhaitee') or line.quantite_souhaitee)
            line.equivalence_possible = request.POST.get(prefix + 'equivalence_possible') == 'on'
            line.save()
        consultation.commentaire = request.POST.get('commentaire', consultation.commentaire)
        consultation.save(update_fields=['commentaire'])
        messages.success(request, 'Consultation fournisseur mise à jour.')
        return redirect('pedashop_consultation_edit', pk=pk)
    return render(request, 'pedashop/consultation_edit.html', {'consultation': consultation})


def consultation_pdf(request, pk):
    consultation = get_object_or_404(SupplierConsultation, pk=pk)
    data = supplier_consultation_pdf_bytes(consultation)
    consultation.statut = 'generee'
    consultation.date_generation_pdf = timezone.now()
    consultation.save(update_fields=['statut', 'date_generation_pdf'])
    response = HttpResponse(data, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{consultation.code}.pdf"'
    return response



# ---------------------------------------------------------------------------
# Inventaire, entrées stock et fiche utilisateurs
# ---------------------------------------------------------------------------

@require_http_methods(['GET', 'POST'])
@transaction.atomic
def stock_entry(request):
    """Entrée en magasin : réassort acheté, retour produit ou stock initial."""
    user = require_storekeeper(request)
    if not user or not _can_act_as_storekeeper(request, user):
        return redirect('pedashop_login')
    form = StockEntryForm(request.POST or None, allowed_magasins=_visible_magasins(user))
    if request.method == 'POST' and form.is_valid():
        article = form.cleaned_data['article']; magasin = form.cleaned_data['magasin']; q = form.cleaned_data['quantite']
        stock, _ = StockArticleMagasin.objects.get_or_create(article=article, magasin=magasin)
        before = stock.stock_reel
        stock.stock_reel += q
        if form.cleaned_data.get('emplacement'):
            stock.emplacement = form.cleaned_data['emplacement']
        stock.save()
        type_mv = form.cleaned_data['type_entree'] if form.cleaned_data['type_entree'] in dict(MouvementStock.TYPE_CHOICES) else 'reception_fournisseur'
        MouvementStock.objects.create(article=article, magasin_destination=magasin, emplacement_destination=stock.emplacement, type_mouvement=type_mv, quantite=q, stock_avant=before, stock_apres=stock.stock_reel, utilisateur=user, commentaire=form.cleaned_data.get('commentaire', ''))
        recalculate_stock_alerts(); messages.success(request, 'Entrée stock enregistrée.')
        return redirect('pedashop_stock_list')
    return render(request, 'pedashop/form.html', {'form': form, 'title': 'Entrée en magasin / réassort'})

@require_http_methods(['GET', 'POST'])
@transaction.atomic
def inventory_adjustment(request):
    """Inventaire ou réassort depuis une page unique.

    En mode inventaire, seul ``stock_reel`` est remplacé : les réservations,
    préparations et sorties temporaires restent intactes. En mode réassort, la
    quantité saisie s'ajoute au stock réel.
    """
    user = require_storekeeper(request)
    if not user or not _can_act_as_storekeeper(request, user):
        return redirect('pedashop_login')
    form = InventoryAdjustmentForm(request.POST or None, allowed_magasins=_visible_magasins(user))
    recent = MouvementStock.objects.filter(type_mouvement__in=['correction_inventaire', 'reception_fournisseur']).select_related('article', 'magasin_destination', 'utilisateur')[:25]
    if request.method == 'POST' and form.is_valid():
        article = form.cleaned_data['article']
        ean = form.cleaned_data.get('ean')
        if ean:
            article = Article.objects.filter(Q(code_ean=ean) | Q(reference_interne=ean) | Q(code_barres_interne=ean)).first() or article
        magasin = form.cleaned_data['magasin']
        stock, _ = StockArticleMagasin.objects.get_or_create(article=article, magasin=magasin)
        before = stock.stock_reel
        qty = form.cleaned_data['stock_reel_compte']
        if form.cleaned_data['operation_type'] == 'reassort':
            stock.stock_reel += qty
            mv_type = 'reception_fournisseur'
            mv_qty = qty
        else:
            stock.stock_reel = qty
            mv_type = 'correction_inventaire'
            mv_qty = qty - before
        if form.cleaned_data.get('stock_mini') is not None:
            stock.stock_minimum = form.cleaned_data['stock_mini']
        if form.cleaned_data.get('emplacement'):
            stock.emplacement = form.cleaned_data['emplacement']
        stock.save()
        MouvementStock.objects.create(article=article, magasin_destination=magasin, type_mouvement=mv_type, quantite=mv_qty, stock_avant=before, stock_apres=stock.stock_reel, utilisateur=user, commentaire=form.cleaned_data.get('commentaire', ''))
        recalculate_stock_alerts(); messages.success(request, 'Opération stock enregistrée.')
        return redirect('pedashop_inventory_adjustment')
    return render(request, 'pedashop/inventory_adjustment.html', {'form': form, 'recent': recent})


def user_list(request):
    user = require_login(request)
    if not user: return redirect('pedashop_login')
    q = request.GET.get('q', '')
    qs = PedaShopUser.objects.filter(active=True)
    if q:
        qs = qs.filter(Q(code__icontains=q) | Q(username__icontains=q) | Q(first_name__icontains=q) | Q(last_name__icontains=q) | Q(class_name__icontains=q))
    return render(request, 'pedashop/user_list.html', {'items': qs[:300], 'q': q})

def user_detail(request, pk):
    user = require_login(request)
    if not user: return redirect('pedashop_login')
    obj = get_object_or_404(PedaShopUser, pk=pk)
    can_edit_visibility = bool(user and user.is_admin_like)
    form = UserVisibilityForm(instance=obj) if can_edit_visibility else None
    context = {'obj': obj, 'form': form, 'can_edit_visibility': can_edit_visibility, 'bons_demandes': Bon.objects.filter(demandeur=obj)[:20], 'bons_prepares': Bon.objects.filter(preparateur=obj)[:20], 'bons_distribues': Bon.objects.filter(distributeur=obj)[:20], 'mouvements': MouvementStock.objects.filter(Q(utilisateur=obj) | Q(demandeur=obj) | Q(preparateur=obj) | Q(distributeur=obj) | Q(receptionnaire=obj)).select_related('article')[:40], 'reclamations': Reclamation.objects.filter(Q(declaree_par=obj) | Q(concerne=obj) | Q(traitee_par=obj))[:30]}
    return render(request, 'pedashop/user_detail.html', context)


@require_http_methods(['POST'])
def user_visibility_update(request, pk):
    admin = require_admin(request)
    if not admin:
        return redirect('pedashop_login')
    obj = get_object_or_404(PedaShopUser, pk=pk)
    form = UserVisibilityForm(request.POST, instance=obj)
    if form.is_valid():
        form.save()
        messages.success(request, 'Magasins visibles et droits PedaShop mis à jour.')
    else:
        messages.error(request, 'Impossible de mettre à jour les droits de visibilité.')
    return redirect('pedashop_user_detail', pk=pk)


# ---------------------------------------------------------------------------
# Transfert / import / exports
# ---------------------------------------------------------------------------

@require_http_methods(['GET', 'POST'])
@transaction.atomic
def transfer_create(request):
    user = require_storekeeper(request)
    if not user or not _can_act_as_storekeeper(request, user):
        return redirect('pedashop_login')
    form = TransferForm(request.POST or None, allowed_magasins=_visible_magasins(user))
    if request.method == 'POST' and form.is_valid():
        article = form.cleaned_data['article']
        src = form.cleaned_data['magasin_source']
        dst = form.cleaned_data['magasin_destination']
        q = form.cleaned_data['quantite']
        if src == dst:
            messages.error(request, 'Le magasin source et destination doivent être différents.')
        else:
            src_stock = StockArticleMagasin.objects.get(article=article, magasin=src)
            dst_stock, _ = StockArticleMagasin.objects.get_or_create(article=article, magasin=dst)
            if src_stock.stock_disponible < q:
                messages.error(request, 'Stock disponible insuffisant pour transfert.')
            else:
                before_src = src_stock.stock_reel
                before_dst = dst_stock.stock_reel
                src_stock.stock_reel -= q
                dst_stock.stock_reel += q
                src_stock.save(); dst_stock.save()
                MouvementStock.objects.create(article=article, magasin_source=src, magasin_destination=dst, emplacement_source=src_stock.emplacement, emplacement_destination=dst_stock.emplacement, type_mouvement='transfert_interne', quantite=q, stock_avant=before_src, stock_apres=src_stock.stock_reel, utilisateur=user, commentaire=form.cleaned_data.get('commentaire', ''))
                MouvementStock.objects.create(article=article, magasin_source=src, magasin_destination=dst, type_mouvement='transfert_interne', quantite=q, stock_avant=before_dst, stock_apres=dst_stock.stock_reel, utilisateur=user, commentaire='Entrée transfert interne')
                recalculate_stock_alerts()
                messages.success(request, 'Transfert interne enregistré.')
                return redirect('pedashop_stock_list')
    return render(request, 'pedashop/form.html', {'form': form, 'title': 'Transfert interne'})


@require_http_methods(['GET', 'POST'])
def import_excel(request):
    user = require_admin(request)
    if not user:
        return redirect('pedashop_login')
    preview = None
    report = None
    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'commit':
            staged = request.session.get('pedashop_import_file')
            magasin_id = request.session.get('pedashop_import_magasin')
            sheet = request.session.get('pedashop_import_sheet')
            check = request.session.get('pedashop_import_check_stock', False)
            if staged and magasin_id:
                rows, info = load_import_rows(staged, sheet)
                report = commit_import(rows, get_object_or_404(Magasin, pk=magasin_id), actor=user, check_stock_consistency=check)
                messages.success(request, f"Import traité : {report['created_articles']} articles créés, {len(report['errors'])} erreurs.")
            else:
                messages.error(request, 'Aucun aperçu d’import à valider.')
        else:
            form = ExcelImportForm(request.POST, request.FILES, allowed_magasins=_visible_magasins(user))
            if form.is_valid():
                upload = form.cleaned_data['fichier']
                suffix = Path(upload.name).suffix or '.xlsx'
                target_dir = Path(settings.MEDIA_ROOT) / 'pedashop_imports'
                target_dir.mkdir(parents=True, exist_ok=True)
                target = target_dir / f'import_{timezone.now():%Y%m%d_%H%M%S}{suffix}'
                with target.open('wb') as f:
                    for chunk in upload.chunks():
                        f.write(chunk)
                rows, info = load_import_rows(str(target), form.cleaned_data.get('feuille'))
                preview = {'rows': rows[:20], 'info': info, 'count': len(rows)}
                request.session['pedashop_import_file'] = str(target)
                request.session['pedashop_import_magasin'] = form.cleaned_data['magasin'].id
                request.session['pedashop_import_sheet'] = form.cleaned_data.get('feuille') or ''
                request.session['pedashop_import_check_stock'] = bool(form.cleaned_data.get('verifier_coherence_stock'))
            else:
                messages.error(request, 'Formulaire import invalide.')
    else:
        form = ExcelImportForm(allowed_magasins=_visible_magasins(user))
    return render(request, 'pedashop/import_excel.html', {'form': locals().get('form', ExcelImportForm(allowed_magasins=_visible_magasins(user))), 'preview': preview, 'report': report})


def export_bon_pdf(request, pk):
    bon = get_object_or_404(Bon, pk=pk)
    rows = [['Référence', 'Désignation', 'Qté demandée', 'Qté distribuée', 'Retour prévu']]
    for l in bon.lignes.select_related('article'):
        rows.append([l.article.reference_interne, l.article.designation, str(l.quantite_demandee), str(l.quantite_distribuee), str(l.date_retour_prevue or '')])
    data = pdf_response_bytes(f'Bon PedaShop {bon.code}', rows, f'Magasin : {bon.magasin} — Statut : {bon.get_statut_display()}')
    response = HttpResponse(data, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{bon.code}.pdf"'
    return response


def export_stock_pdf(request):
    rows = [['Magasin', 'Référence', 'Désignation', 'Stock réel', 'Réservé', 'Prépa', 'Extérieur', 'Disponible', 'Stock mini']]
    for s in StockArticleMagasin.objects.select_related('article', 'magasin'):
        rows.append([s.magasin.code, s.article.reference_interne, s.article.designation, str(s.stock_reel), str(s.reserve_total), str(s.stock_en_preparation), str(s.stock_temporairement_sorti), str(s.stock_disponible), str(s.stock_minimum)])
    data = pdf_response_bytes('État de stock PedaShop', rows)
    response = HttpResponse(data, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="pedashop_etat_stock.pdf"'
    return response



def articles_template_xlsx(request):
    return _xlsx_response(_make_articles_template_workbook(), 'modele_base_pedashop.xlsx')


def articles_export_xlsx(request):
    wb = _make_articles_template_workbook()
    ws = wb.active
    # On conserve la première ligne exemple du modèle uniquement dans le modèle ; export = données réelles.
    ws.delete_rows(2, ws.max_row)
    for a in Article.objects.all().order_by('reference_interne'):
        ws.append([a.fabricant, a.reference_fabricant, a.reference_interne, a.code_ean, f'{a.reference_interne}.jpg', a.designation, '', a.categorie, a.sous_categorie, '', '', '', '', '', '', '', 'Oui' if a.substituable else 'Non', a.fournisseur, a.marche])
    return _xlsx_response(wb, 'export_articles_pedashop.xlsx')


def stock_export_xlsx(request):
    user = current_user(request)
    qs = _filter_visible_stocks(StockArticleMagasin.objects.select_related('article', 'magasin', 'emplacement'), user)
    wb = Workbook(); ws = wb.active; ws.title = 'Stock PedaShop'
    ws.append(['Magasin', 'Emplacement', 'Code produit', 'EAN', 'Fabricant', 'Référence constructeur', 'Désignation', 'Catégorie', 'Sous-catégorie', 'Stock réel', 'Réservé', 'Prépa', 'Extérieur', 'Disponible', 'Stock mini', 'Usé', 'HS', 'Perdu', 'Marché'])
    for s0 in qs:
        ws.append([s0.magasin.code, s0.emplacement.code if s0.emplacement else '', s0.article.reference_interne, s0.article.code_ean, s0.article.fabricant, s0.article.reference_fabricant, s0.article.designation, s0.article.categorie, s0.article.sous_categorie, s0.stock_reel, s0.reserve_total, s0.stock_en_preparation, s0.stock_temporairement_sorti, s0.stock_disponible, s0.stock_minimum, s0.qte_use, s0.stock_hs, getattr(s0, 'stock_perdu', 0), s0.article.marche])
    return _xlsx_response(wb, 'export_stock_pedashop.xlsx')


# --- Administration SQL base module ---
def sql_database_admin(request):
    from django.shortcuts import redirect
    from .permissions import require_admin
    from .db_sql_admin import render_sql_admin
    user = require_admin(request)
    if not user:
        return redirect('pedashop_dashboard')
    return render_sql_admin(request, 'pedashop/sql_database.html', 'PedaShop')


def sql_database_export(request):
    from django.shortcuts import redirect
    from .permissions import require_admin
    from .db_sql_admin import export_sql_response
    user = require_admin(request)
    if not user:
        return redirect('pedashop_dashboard')
    return export_sql_response(request, 'pedashop')


def sql_database_import(request):
    from django.shortcuts import redirect
    from .permissions import require_admin
    from .db_sql_admin import import_sql_response
    user = require_admin(request)
    if not user:
        return redirect('pedashop_dashboard')
    return import_sql_response(request, 'pedashop/sql_database.html', 'PedaShop', 'pedashop')

def help_view(request):
    return render(request, 'pedashop/help.html')


def about_view(request):
    return render(request, 'pedashop/about.html')
