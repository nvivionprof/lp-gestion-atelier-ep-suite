import json
import re
from pathlib import Path
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.utils import timezone
from openpyxl import load_workbook
from pfmp_manager.models import Company, CompanyContact, Formation, CompanyTag, ImportBatch


def clean(value):
    if value is None:
        return ''
    return str(value).strip()


def yes(value):
    return clean(value).lower() in {'oui','yes','true','1','x','vrai'}


def dec(value):
    if value in (None, ''):
        return None
    try:
        return str(value).replace(',', '.').strip()
    except Exception:
        return None


def norm_code(value):
    value = clean(value).upper()
    value = re.sub(r'[^A-Z0-9]+', '_', value).strip('_')
    return value[:80] or 'TAG'


class Command(BaseCommand):
    help = 'Importe la base entreprises/contacts PFMP depuis le XLSX RC16.'

    def add_arguments(self, parser):
        parser.add_argument('--file', required=True, help='Chemin XLSX dans le conteneur')
        parser.add_argument('--mode', default='simulation', choices=['simulation','append_only','upsert','replace_all','delete_all_then_import'])
        parser.add_argument('--key', default='code_entreprise', choices=['code_entreprise','siret','nom_code_postal_ville'])
        parser.add_argument('--confirm', default='')

    def handle(self, *args, **options):
        path = Path(options['file'])
        if not path.exists():
            raise CommandError(f'Fichier introuvable : {path}')
        mode = options['mode']
        key_strategy = options['key']
        destructive = mode in {'replace_all','delete_all_then_import'}
        if destructive and options.get('confirm') != 'CONFIRMER IMPORT DESTRUCTIF':
            raise CommandError('Mode destructif refusé : ajoute --confirm "CONFIRMER IMPORT DESTRUCTIF"')

        wb = load_workbook(path, read_only=True, data_only=True)
        for required in ['Entreprises_Import','Contacts_Import']:
            if required not in wb.sheetnames:
                raise CommandError(f'Feuille manquante : {required}')

        batch = ImportBatch.objects.create(file_name=path.name, mode=mode, key_strategy=key_strategy)
        report = {'companies': [], 'contacts': [], 'errors': []}
        stats = {'created':0,'updated':0,'deleted':0,'ignored':0,'errors':0}

        with transaction.atomic():
            if destructive and mode != 'simulation':
                stats['deleted'] += CompanyContact.objects.count()
                CompanyContact.objects.all().delete()
                stats['deleted'] += Company.objects.count()
                Company.objects.all().delete()

            company_index = {}
            rows = self.rows(wb['Entreprises_Import'])
            for row in rows:
                try:
                    key = self.company_key(row, key_strategy)
                    if not key:
                        stats['ignored'] += 1
                        continue
                    payload = self.company_payload(row, batch)
                    formations = self.ensure_formations(clean(row.get('domaines')))
                    tags = self.ensure_tags(clean(row.get('domaines')), clean(row.get('sous_domaines')), clean(row.get('secteur_activite_source')))
                    existing = self.find_company(key_strategy, row, key)
                    if mode == 'simulation':
                        report['companies'].append({'key': key, 'action': 'update' if existing else 'create', 'name': payload['name']})
                        continue
                    if mode == 'append_only' and existing:
                        company = existing
                        stats['ignored'] += 1
                    else:
                        if existing:
                            for k, v in payload.items():
                                setattr(existing, k, v)
                            existing.save()
                            company = existing
                            stats['updated'] += 1
                        else:
                            company = Company.objects.create(**payload)
                            stats['created'] += 1
                        company.formations.set(formations)
                        company.tags.set(tags)
                    company_index[clean(row.get('code_entreprise'))] = company
                except Exception as exc:
                    stats['errors'] += 1
                    report['errors'].append({'sheet':'Entreprises_Import','row':row,'error':str(exc)})

            rows = self.rows(wb['Contacts_Import'])
            for row in rows:
                try:
                    code = clean(row.get('code_entreprise'))
                    company = company_index.get(code) or Company.objects.filter(external_key=code).first()
                    if not company:
                        stats['ignored'] += 1
                        report['contacts'].append({'company_key': code, 'action': 'ignored_company_not_found', 'email': clean(row.get('email'))})
                        continue
                    email = clean(row.get('email'))
                    full_name = clean(row.get('nom_complet')) or email or 'Contact PFMP'
                    qs = CompanyContact.objects.filter(company=company, email=email) if email else CompanyContact.objects.filter(company=company, full_name=full_name)
                    existing = qs.first()
                    payload = self.contact_payload(row, company, batch)
                    formations = self.ensure_formations(clean(row.get('domaine_principal')))
                    if mode == 'simulation':
                        report['contacts'].append({'company_key': code, 'action': 'update' if existing else 'create', 'email': email, 'name': full_name})
                        continue
                    if mode == 'append_only' and existing:
                        stats['ignored'] += 1
                        contact = existing
                    else:
                        if existing:
                            for k, v in payload.items():
                                setattr(existing, k, v)
                            existing.save()
                            contact = existing
                            stats['updated'] += 1
                        else:
                            contact = CompanyContact.objects.create(**payload)
                            stats['created'] += 1
                    contact.formations.set(formations)
                except Exception as exc:
                    stats['errors'] += 1
                    report['errors'].append({'sheet':'Contacts_Import','row':row,'error':str(exc)})

            if mode == 'simulation':
                # Annule les créations de batch et d'objets référentiels éventuels créés pendant la simulation.
                transaction.set_rollback(True)

        if mode == 'simulation':
            self.stdout.write(self.style.WARNING('SIMULATION : aucune donnée entreprise/contact modifiée.'))
            self.stdout.write(json.dumps(report, ensure_ascii=False, indent=2)[:12000])
            return

        batch.created_count = stats['created']
        batch.updated_count = stats['updated']
        batch.deleted_count = stats['deleted']
        batch.ignored_count = stats['ignored']
        batch.error_count = stats['errors']
        batch.report_json = report
        batch.finished_at = timezone.now()
        batch.save()
        self.stdout.write(self.style.SUCCESS(f"Import terminé : créés={stats['created']} mis_a_jour={stats['updated']} supprimés={stats['deleted']} ignorés={stats['ignored']} erreurs={stats['errors']}"))

    def rows(self, ws):
        headers = [clean(v) for v in next(ws.iter_rows(max_row=1, values_only=True))]
        for values in ws.iter_rows(min_row=2, values_only=True):
            yield dict(zip(headers, values))

    def company_key(self, row, strategy):
        if strategy == 'siret':
            return clean(row.get('siret'))
        if strategy == 'nom_code_postal_ville':
            return '|'.join([clean(row.get('nom')).lower(), clean(row.get('code_postal')), clean(row.get('ville')).lower()])
        return clean(row.get('code_entreprise'))

    def find_company(self, strategy, row, key):
        if strategy == 'siret' and key:
            return Company.objects.filter(siret=key).first()
        if strategy == 'nom_code_postal_ville':
            return Company.objects.filter(name__iexact=clean(row.get('nom')), postal_code=clean(row.get('code_postal')), city__iexact=clean(row.get('ville'))).first()
        return Company.objects.filter(external_key=key).first()

    def company_payload(self, row, batch):
        status = 'active' if yes(row.get('actif')) else 'inactive'
        full_address = clean(row.get('adresse_complete')) or ', '.join([v for v in [clean(row.get('adresse_1')), clean(row.get('code_postal')), clean(row.get('ville')), clean(row.get('pays')) or 'France'] if v])
        return {
            'external_key': clean(row.get('code_entreprise')) or None,
            'name': clean(row.get('nom'))[:180] or 'Entreprise sans nom',
            'status': status,
            'student_visible': yes(row.get('visible_eleves')),
            'student_visible_notes': clean(row.get('description')),
            'siret': clean(row.get('siret'))[:20],
            'naf_ape': clean(row.get('naf_ape'))[:20],
            'activity': clean(row.get('secteur_activite_source'))[:220],
            'source_activity': clean(row.get('secteur_activite_source'))[:260],
            'domains_text': clean(row.get('domaines'))[:260],
            'subdomains_text': clean(row.get('sous_domaines'))[:260],
            'address': ' '.join([clean(row.get('adresse_1')), clean(row.get('adresse_2')), clean(row.get('adresse_3'))]).strip()[:240],
            'postal_code': clean(row.get('code_postal'))[:20],
            'city': clean(row.get('ville'))[:120],
            'country': clean(row.get('pays'))[:80] or 'France',
            'full_address': full_address[:360],
            'latitude': dec(row.get('latitude')),
            'longitude': dec(row.get('longitude')),
            'geocoding_status': clean(row.get('statut_geocodage'))[:40] or 'A_GEOCODER',
            'osm_search_url': clean(row.get('osm_search_url')),
            'phone': clean(row.get('telephone'))[:40],
            'email': clean(row.get('email_general')),
            'website': clean(row.get('site_web')),
            'transport_access': clean(row.get('Transport'))[:220],
            'global_rating': int(float(row.get('Notation') or 0)) if clean(row.get('Notation')) else 0,
            'internal_comment': clean(row.get('commentaire_admin')),
            'import_source': 'xlsx_rc16',
            'import_batch': batch,
        }

    def contact_payload(self, row, company, batch):
        contact_type_raw = clean(row.get('type_contact')).lower()
        if 'dirigeant' in contact_type_raw:
            contact_type = 'dirigeant'
        elif 'rh' in contact_type_raw:
            contact_type = 'rh'
        elif 'tech' in contact_type_raw:
            contact_type = 'technique'
        else:
            contact_type = 'pfmp'
        visible_eleve = yes(row.get('visible_eleve')) or yes(row.get('contact_eleve'))
        return {
            'company': company,
            'full_name': clean(row.get('nom_complet'))[:160] or clean(row.get('email')) or 'Contact PFMP',
            'role': clean(row.get('fonction'))[:120],
            'service': clean(row.get('sous_domaine'))[:120],
            'email': clean(row.get('email')),
            'phone': clean(row.get('telephone'))[:40],
            'mobile_phone': clean(row.get('telephone_mobile'))[:40],
            'contact_type': contact_type,
            'visibility': 'students' if visible_eleve else 'professors',
            'student_visible': visible_eleve,
            'teacher_visible': yes(row.get('visible_prof')) or True,
            'can_help_transport': yes(row.get('Transport stagiaire ')),
            'personal_address': clean(row.get('Adresse perso'))[:240],
            'use_personal_location_for_student_search': bool(clean(row.get('Adresse perso'))),
            'note': clean(row.get('commentaire')),
            'import_source': 'xlsx_rc16',
            'import_batch': batch,
            'active': True,
        }

    def ensure_formations(self, text):
        result = []
        for part in re.split(r'[;,/]+', text or ''):
            code = norm_code(part)
            if not code:
                continue
            f, _ = Formation.objects.get_or_create(code=code, defaults={'nom': clean(part)[:160] or code})
            result.append(f)
        return result

    def ensure_tags(self, *texts):
        result = []
        for text in texts:
            for part in re.split(r'[;,/]+', text or ''):
                label = clean(part)
                if not label:
                    continue
                code = norm_code(label)
                tag, _ = CompanyTag.objects.get_or_create(code=code, defaults={'label': label[:140], 'category': 'activite'})
                result.append(tag)
        return result
