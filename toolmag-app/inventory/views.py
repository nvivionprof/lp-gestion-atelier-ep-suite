from django.conf import settings
from django.contrib import messages
from django.core import signing
from django.db import transaction, IntegrityError
from django.db.models import Count, Q
from django.http import FileResponse, HttpResponse, JsonResponse, HttpResponseForbidden
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.utils.dateparse import parse_date
from django.views.decorators.csrf import csrf_exempt
import base64
import io
import qrcode
import json
import os
import subprocess
import uuid
from datetime import datetime, time
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from .forms import (CheckoutForm, ComponentEditForm, EquipmentDocumentEditForm, EquipmentDuplicateForm, EquipmentForm, EvaluationFilterForm, EvaluationValidationForm,
    ForceLockerForm, MaterialEditGrantForm, PedagogicalSessionForm, PersonCreateForm, PromotionActionForm,
    PromotionFilterForm, ReturnForm, StorekeeperSessionForm, TerminalRegistrationForm, UserImportForm, UserSessionForm,
    ChangePasswordForm, ResetPasswordForm, ManualBackupForm, RestoreBackupForm, RepairForm, InterventionForm)
from .models import AuthorizedTerminal, Category, Component, ComponentCheck, Competence, CompetenceMapping, EnrollmentHistory, Equipment, EquipmentDocument, EvaluationRecord, Formation, SchoolClass, InterventionLog, Loan, LockerOpenLog, LockerSettings, Location, MaterialEditGrant, PedagogicalSession, Person, RepairLog, SessionRoleAssignment, UserInventory, UserInventoryItem, next_equipment_code_for_category
from .backup_utils import create_backup, list_backups, restore_backup_from_archive, safe_backup_path

USER_COLUMNS = [
    'code_utilisateur', 'nom', 'prenom', 'identifiant', 'email', 'formation', 'classe', 'groupe',
    'niveau', 'actif', 'archive', 'role_principal', 'roles_autorises', 'rfid_uid', 'mot_de_passe_initial'
]
ROLE_EXPORT = {
    Person.Role.USER: 'UTILISATEUR',
    Person.Role.STOREKEEPER: 'MAGASINIER',
    Person.Role.TECH_INVENTORY: 'TECH_INVENTAIRE',
    Person.Role.RESPONSIBLE: 'RESPONSABLE',
    Person.Role.READ_ONLY: 'LECTURE_SEULE',
    Person.Role.ADMIN: 'ADMIN',
}
ROLE_IMPORT = {v: k for k, v in ROLE_EXPORT.items()}
ROLE_IMPORT.update(Person.ROLE_IMPORT_ALIASES)


def make_qr_data_uri(payload):
    """Return a PNG data URI QR code for a URL or payload."""
    img = qrcode.make(payload)
    buffer = io.BytesIO()
    img.save(buffer, format='PNG')
    encoded = base64.b64encode(buffer.getvalue()).decode('ascii')
    return f'data:image/png;base64,{encoded}'


def absolute_uri(request, view_name, **kwargs):
    return request.build_absolute_uri(reverse(view_name, kwargs=kwargs))






def get_client_ip(request):
    forwarded = request.META.get('HTTP_X_FORWARDED_FOR', '')
    if forwarded:
        return forwarded.split(',')[0].strip()
    return request.META.get('REMOTE_ADDR', '') or ''


def current_terminal(request):
    token = request.COOKIES.get('toolmag_terminal_token') or request.session.get('terminal_token')
    if not token:
        return None
    terminal = AuthorizedTerminal.objects.filter(token=token, active=True).first()
    if terminal:
        terminal.last_seen_at = timezone.now()
        terminal.last_ip = get_client_ip(request)
        terminal.user_agent = request.META.get('HTTP_USER_AGENT', '')[:2000]
        terminal.save(update_fields=['last_seen_at', 'last_ip', 'user_agent', 'updated_at'])
    return terminal


def is_super_admin_person(person):
    """Droits professeur / responsable stricts.
    Important : les rôles autorisés MAGASINIER ne doivent pas donner les droits professeur.
    Seul le rôle principal RESPONSABLE ou ADMIN ouvre les fonctions sensibles.
    """
    return bool(person and person.role in [Person.Role.RESPONSIBLE, Person.Role.ADMIN])


def locker_availability(request, equipment=None, context='detail', force_mode=False):
    settings_obj = LockerSettings.get_solo()
    storekeeper = current_storekeeper(request)
    terminal = current_terminal(request)
    ip = get_client_ip(request)
    allowed_ips = settings_obj.allowed_ip_set()
    reasons = []
    if not settings_obj.module_enabled:
        reasons.append('module armoires désactivé')
    if not storekeeper:
        reasons.append('magasinier non connecté')
    if equipment is not None:
        if not equipment.secure_storage:
            reasons.append('matériel non stocké en armoire sécurisée')
        if not equipment.secure_cabinet or not equipment.secure_locker:
            reasons.append('armoire/casier non renseigné')
    if settings_obj.require_authorized_terminal:
        bypass = force_mode and settings_obj.allow_superadmin_force_without_terminal and is_super_admin_person(storekeeper)
        if not bypass and not (terminal and terminal.can_open_lockers and terminal.active):
            reasons.append('terminal non autorisé')
    if settings_obj.require_allowed_public_ip:
        bypass = force_mode and settings_obj.allow_superadmin_force_without_ip and is_super_admin_person(storekeeper)
        if not bypass and allowed_ips and ip not in allowed_ips:
            reasons.append(f'IP non autorisée : {ip or "inconnue"}')
    return {
        'settings': settings_obj,
        'storekeeper': storekeeper,
        'terminal': terminal,
        'client_ip': ip,
        'allowed': len(reasons) == 0,
        'reasons': reasons,
    }


def dispatch_locker_script(payload, timeout=5):
    """Exécute un script local côté serveur. Le script reçoit le JSON sur stdin.
    Variable .env attendue : LOCKER_POST_SCRIPT=/chemin/script_ouverture.py
    Si aucun script n'est configuré, l'action est journalisée mais aucune commande physique n'est envoyée.
    """
    script = os.getenv('LOCKER_POST_SCRIPT', '').strip()
    if not script:
        return False, 'Aucun script LOCKER_POST_SCRIPT configuré : ouverture simulée / journalisée uniquement.'
    try:
        result = subprocess.run(
            [script],
            input=json.dumps(payload, ensure_ascii=False),
            text=True,
            capture_output=True,
            timeout=timeout,
            check=False,
        )
        output = (result.stdout or '') + (('\nERR: ' + result.stderr) if result.stderr else '')
        return result.returncode == 0, output.strip()[:4000]
    except Exception as exc:
        return False, f'Erreur script ouverture casier : {exc}'


def create_locker_log(request, *, equipment=None, cabinet='', locker='', context='detail', force_reason=''):
    avail = locker_availability(request, equipment=equipment, context=context, force_mode=(context == LockerOpenLog.Context.FORCE))
    storekeeper = avail['storekeeper']
    terminal = avail['terminal']
    settings_obj = avail['settings']
    payload = {
        'armoire': cabinet or (equipment.secure_cabinet if equipment else ''),
        'casier': locker or (equipment.secure_locker if equipment else ''),
        'materiel_code': equipment.code if equipment else '',
        'materiel_nom': equipment.name if equipment else '',
        'contexte': context,
        'motif': force_reason,
        'magasinier_code': storekeeper.code if storekeeper else '',
        'magasinier_nom': storekeeper.last_name if storekeeper else '',
        'magasinier_prenom': storekeeper.first_name if storekeeper else '',
        'terminal': terminal.name if terminal else '',
        'ip_client': avail['client_ip'],
        'horodatage': timezone.localtime().isoformat(),
    }
    log = LockerOpenLog.objects.create(
        equipment=equipment,
        storekeeper=storekeeper,
        terminal=terminal,
        cabinet=payload['armoire'],
        locker=payload['casier'],
        context=context,
        client_ip=avail['client_ip'],
        user_agent=request.META.get('HTTP_USER_AGENT', '')[:2000],
        payload=payload,
    )
    if not avail['allowed']:
        log.refused = True
        log.refusal_reason = '; '.join(avail['reasons'])
        log.controller_response = 'Ouverture refusée par ToolMag.'
        log.save(update_fields=['refused', 'refusal_reason', 'controller_response', 'updated_at'])
        return log
    ok, response = dispatch_locker_script(payload, timeout=settings_obj.script_timeout_seconds)
    log.success = ok
    log.controller_response = response
    log.save(update_fields=['success', 'controller_response', 'updated_at'])
    return log



def _portal_token_payload(request):
    token = request.GET.get('token') or request.POST.get('token') or ''
    if not token:
        return None
    try:
        return signing.loads(token, key=getattr(settings, 'LP_CORE_API_TOKEN', ''), salt='lp-suite-sso', max_age=600)
    except Exception:
        return None


def _safe_next_or_dashboard(request):
    nxt = request.GET.get('next') or request.POST.get('next') or ''
    prefix = getattr(settings, 'APP_URL_PREFIX', '').rstrip('/')
    if nxt.startswith('http://') or nxt.startswith('https://'):
        return reverse('dashboard')
    if prefix and nxt.startswith(prefix + '/'):
        return nxt
    if nxt.startswith('/') and not nxt.startswith('//'):
        return nxt
    return reverse('dashboard')


def portal_login(request):
    payload = _portal_token_payload(request)
    if not payload:
        messages.error(request, 'Connexion LP Core impossible ou expirée. Merci de te reconnecter.')
        return redirect('storekeeper_login')
    code = (payload.get('code') or '').strip()
    username = (payload.get('username') or '').strip()
    person = Person.objects.filter(Q(code=code) | Q(username=username), active=True, archived=False).first()
    if not person:
        messages.error(request, 'Compte LP Core non synchronisé dans ToolMag. Lance une synchronisation LP Core → ToolMag.')
        return redirect('storekeeper_login')
    is_prof = is_super_admin_person(person)
    is_mag = person.has_role(Person.Role.STOREKEEPER, Person.Role.RESPONSIBLE, Person.Role.ADMIN)
    if is_prof:
        request.session['storekeeper_code'] = person.code
        request.session['borrower_code'] = person.code
        messages.success(request, f'Connexion ToolMag via LP Core : {person.first_name} {person.last_name} — mode professeur/magasinier + utilisateur.')
        return redirect(_safe_next_or_dashboard(request))
    if is_mag and person.has_role(Person.Role.USER):
        request.session['tm_pending_role_code'] = person.code
        return redirect('role_choice')
    if is_mag:
        request.session['storekeeper_code'] = person.code
        messages.success(request, f'Connexion ToolMag via LP Core : {person.first_name} {person.last_name} — mode magasinier.')
        return redirect(_safe_next_or_dashboard(request))
    request.session['borrower_code'] = person.code
    messages.success(request, f'Connexion ToolMag via LP Core : {person.first_name} {person.last_name} — mode utilisateur.')
    return redirect(_safe_next_or_dashboard(request))


def role_choice(request):
    code = request.session.get('tm_pending_role_code')
    person = Person.objects.filter(code=code, active=True, archived=False).first() if code else None
    if not person:
        return redirect('dashboard')
    if request.method == 'POST':
        role = request.POST.get('role')
        if role == 'storekeeper' and person.has_role(Person.Role.STOREKEEPER, Person.Role.RESPONSIBLE, Person.Role.ADMIN):
            request.session['storekeeper_code'] = person.code
            request.session.pop('borrower_code', None)
        elif role == 'both' and is_super_admin_person(person):
            request.session['storekeeper_code'] = person.code
            request.session['borrower_code'] = person.code
        else:
            request.session['borrower_code'] = person.code
            request.session.pop('storekeeper_code', None)
        request.session.pop('tm_pending_role_code', None)
        return redirect(_safe_next_or_dashboard(request))
    return render(request, 'inventory/role_choice.html', {'person': person})

def current_storekeeper(request):
    code = request.session.get('storekeeper_code')
    if not code:
        return None
    person = Person.objects.filter(code=code, active=True, archived=False).first()
    if person and person.has_role(Person.Role.STOREKEEPER, Person.Role.RESPONSIBLE, Person.Role.ADMIN):
        return person
    request.session.pop('storekeeper_code', None)
    return None


def storekeeper_login(request):
    form = StorekeeperSessionForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        storekeeper = form.cleaned_data['storekeeper']
        request.session['storekeeper_code'] = storekeeper.code
        if is_super_admin_person(storekeeper):
            request.session['borrower_code'] = storekeeper.code
        messages.success(request, f'Magasinier connecté : {storekeeper.first_name} {storekeeper.last_name}.')
        if storekeeper.must_change_password:
            return redirect('change_password')
        next_url = request.GET.get('next') or request.POST.get('next') or reverse('dashboard')
        return redirect(next_url)
    return render(request, 'inventory/storekeeper_login.html', {'form': form, 'next': request.GET.get('next', '')})


def storekeeper_logout(request):
    request.session.pop('storekeeper_code', None)
    messages.success(request, 'Magasinier déconnecté.')
    return redirect('dashboard')


def current_borrower(request):
    code = request.session.get('borrower_code')
    if not code:
        return None
    person = Person.objects.filter(code=code, active=True, archived=False).first()
    if person:
        return person
    request.session.pop('borrower_code', None)
    return None


def user_login(request):
    form = UserSessionForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        borrower = form.cleaned_data['borrower']
        request.session['borrower_code'] = borrower.code
        if is_super_admin_person(borrower):
            request.session['storekeeper_code'] = borrower.code
        messages.success(request, f'Utilisateur connecté : {borrower.first_name} {borrower.last_name}.')
        if borrower.must_change_password:
            return redirect('change_password')
        next_url = request.GET.get('next') or request.POST.get('next') or reverse('dashboard')
        return redirect(next_url)
    return render(request, 'inventory/user_login.html', {'form': form, 'next': request.GET.get('next', '')})


def user_logout(request):
    request.session.pop('borrower_code', None)
    messages.success(request, 'Utilisateur déconnecté.')
    return redirect('dashboard')


def api_person_search(request):
    """Recherche rapide d'utilisateur/emprunteur pour préremplir un code depuis un nom."""
    q = (request.GET.get('q') or '').strip()
    people = Person.objects.filter(active=True, archived=False)
    if request.GET.get('role') == 'storekeeper':
        people = [p for p in people if p.has_role(Person.Role.STOREKEEPER, Person.Role.RESPONSIBLE, Person.Role.ADMIN)]
    if q:
        terms = [term for term in q.replace(',', ' ').split() if term]
        if isinstance(people, list):
            low_terms = [t.lower() for t in terms]
            people = [p for p in people if all(t in (' '.join([p.code or '', p.first_name or '', p.last_name or '', p.username or '', p.email or '', p.class_name or '', p.group_name or ''])).lower() for t in low_terms)]
        else:
            for term in terms:
                people = people.filter(
                    Q(code__icontains=term)
                    | Q(first_name__icontains=term)
                    | Q(last_name__icontains=term)
                    | Q(username__icontains=term)
                    | Q(email__icontains=term)
                    | Q(class_name__icontains=term)
                    | Q(group_name__icontains=term)
                )
    if isinstance(people, list):
        people = sorted(people, key=lambda p: (p.last_name, p.first_name))[:12]
    else:
        people = people.order_by('last_name', 'first_name')[:12]
    results = []
    for person in people:
        results.append({
            'code': person.code,
            'name': f'{person.first_name} {person.last_name}'.strip() or person.username or person.code,
            'username': person.username,
            'formation': person.formation.code if person.formation else '',
            'class_name': person.class_name,
            'group_name': person.group_name,
            'role': person.get_role_display(),
        })
    return JsonResponse({'results': results})



def _current_session_person(request):
    # Priorité au magasinier actif ; sinon emprunteur actif.
    return current_storekeeper(request) or current_borrower(request)


def change_password(request):
    person = _current_session_person(request)
    if not person:
        messages.error(request, 'Connecte-toi comme utilisateur ou magasinier avant de changer le mot de passe.')
        return redirect('user_login')
    form = ChangePasswordForm(request.POST or None, person=person)
    if request.method == 'POST' and form.is_valid():
        person.set_password(form.cleaned_data['new_password'])
        person.must_change_password = False
        person.save(update_fields=['password_hash', 'must_change_password', 'updated_at'])
        messages.success(request, 'Mot de passe modifié.')
        return redirect('dashboard')
    return render(request, 'inventory/change_password.html', {'form': form, 'person': person})


def reset_password(request):
    # Réinitialisation accessible à un magasinier responsable/admin connecté.
    storekeeper = current_storekeeper(request)
    if not is_super_admin_person(storekeeper):
        messages.error(request, 'Réinitialisation réservée à un professeur / responsable ToolMag connecté comme magasinier.')
        return redirect('storekeeper_login')
    selected_person_code = (request.GET.get('person_code') or '').strip()
    selected_person = find_active_person_by_code(selected_person_code) if selected_person_code else None
    form = ResetPasswordForm(
        request.POST or None,
        selected_person_code=selected_person.code if selected_person else selected_person_code,
    )
    if request.method == 'POST' and form.is_valid():
        person = form.cleaned_data['person']
        person.set_password(form.cleaned_data['new_password'])
        person.must_change_password = form.cleaned_data.get('force_change', True)
        person.save(update_fields=['password_hash', 'must_change_password', 'updated_at'])
        messages.success(request, f'Mot de passe réinitialisé pour {person.first_name} {person.last_name} ({person.code}).')
        return redirect('users_management')
    return render(request, 'inventory/reset_password.html', {
        'form': form,
        'current_storekeeper': storekeeper,
        'selected_person': selected_person,
        'selected_person_code': selected_person_code,
    })



def require_prof(request):
    storekeeper = current_storekeeper(request)
    if not is_super_admin_person(storekeeper):
        messages.error(request, 'Accès réservé aux professeurs / responsables ToolMag.')
        return False
    return True




def _active_people_for_permissions(request):
    """Retourne toutes les identités actives de la session ToolMag.

    Un professeur peut être connecté à la fois comme utilisateur et magasinier.
    Un élève peut aussi être connecté comme magasinier pendant une séance. Les
    droits ponctuels matériel doivent donc être vérifiés sur toutes les identités
    actives, et pas seulement sur la première trouvée.
    """
    people = []
    for person in (current_storekeeper(request), current_borrower(request)):
        if person and person.id not in [p.id for p in people]:
            people.append(person)
    return people


def _active_person_for_permissions(request):
    people = _active_people_for_permissions(request)
    return people[0] if people else None


def _matching_material_grants(person):
    if not person:
        return MaterialEditGrant.objects.none()
    return [grant for grant in MaterialEditGrant.objects.select_related('formation', 'granted_by').filter(active=True) if grant.applies_to(person)]


def _has_material_grant(person, attr):
    if is_super_admin_person(person):
        return True
    return any(getattr(grant, attr, False) for grant in _matching_material_grants(person))


def _has_any_material_grant(person):
    if is_super_admin_person(person):
        return True
    grant_fields = [
        'can_create_equipment',
        'can_edit_equipment',
        'can_add_photo',
        'can_add_document',
        'can_edit_components',
        'can_edit_location',
        'can_edit_description',
        'can_generate_qr',
    ]
    return any(any(getattr(grant, field, False) for field in grant_fields) for grant in _matching_material_grants(person))


def _request_has_material_grant(request, attr):
    return any(_has_material_grant(person, attr) for person in _active_people_for_permissions(request))


def _request_has_any_material_grant(request):
    return any(_has_any_material_grant(person) for person in _active_people_for_permissions(request))


def can_create_equipment_person(person):
    return _has_material_grant(person, 'can_create_equipment')


def can_edit_equipment_person(person):
    return _has_material_grant(person, 'can_edit_equipment')


def can_edit_components_person(person):
    return _has_material_grant(person, 'can_edit_components')


def can_add_document_person(person):
    return _has_material_grant(person, 'can_add_document')


def require_material_permission(request, attr, message='Droit de modification matériel insuffisant.'):
    people = _active_people_for_permissions(request)
    if not people:
        messages.warning(request, 'Connexion nécessaire.')
        return None, redirect('user_login')
    for person in people:
        if _has_material_grant(person, attr):
            return person, None
    messages.error(request, message)
    return people[0], redirect('equipment_list')

def _date_range_from_request(request):
    raw_from = request.GET.get('date_from') or ''
    raw_to = request.GET.get('date_to') or ''
    date_from = parse_date(raw_from) if raw_from else None
    date_to = parse_date(raw_to) if raw_to else None
    dt_from = timezone.make_aware(datetime.combine(date_from, time.min)) if date_from else None
    dt_to = timezone.make_aware(datetime.combine(date_to, time.max)) if date_to else None
    return date_from, date_to, dt_from, dt_to


def _filter_dt(qs, field_name, dt_from=None, dt_to=None):
    if dt_from:
        qs = qs.filter(**{f'{field_name}__gte': dt_from})
    if dt_to:
        qs = qs.filter(**{f'{field_name}__lte': dt_to})
    return qs

def dashboard(request):
    now = timezone.now()
    stats = {
        'available': Equipment.objects.filter(status=Equipment.Status.AVAILABLE).count(),
        'out': Equipment.objects.filter(status=Equipment.Status.OUT).count(),
        'late': Loan.objects.filter(status=Loan.LoanStatus.OPEN, due_at__lt=now).count(),
        'maintenance': Equipment.objects.filter(status=Equipment.Status.MAINTENANCE).count(),
        'incomplete': Equipment.objects.filter(status=Equipment.Status.INCOMPLETE).count(),
    }
    current_loans = list(Loan.objects.filter(status=Loan.LoanStatus.OPEN).select_related('equipment', 'borrower')[:8])
    # V25 : les statuts d'inventaire utilisateur sont aussi visibles sur le tableau de bord.
    for loan in current_loans:
        loan.user_inventory_status = _loan_user_inventory_status(loan)
    top_equipment = Equipment.objects.annotate(nb_loans=Count('loans')).order_by('-nb_loans')[:5]
    return render(request, 'inventory/dashboard.html', {'stats': stats, 'current_loans': current_loans, 'top_equipment': top_equipment, 'now': now})


def equipment_list(request):
    q = request.GET.get('q', '')
    status = request.GET.get('status', '')
    equipments = Equipment.objects.select_related('category', 'location')
    if q:
        equipments = equipments.filter(
            Q(code__icontains=q)
            | Q(name__icontains=q)
            | Q(description__icontains=q)
            | Q(brand__icontains=q)
            | Q(model__icontains=q)
            | Q(serial_number__icontains=q)
            | Q(location__name__icontains=q)
            | Q(category__name__icontains=q)
        )
    if status:
        equipments = equipments.filter(status=status)
    can_create = _request_has_material_grant(request, 'can_create_equipment')
    return render(request, 'inventory/equipment_list.html', {'equipments': equipments, 'q': q, 'status': status, 'statuses': Equipment.Status.choices, 'can_create_equipment': can_create})



def _can_create_lookup_item(request):
    # Ajout rapide depuis les listes déroulantes Catégorie / Emplacement.
    # Autorisé aux profs/admins et aux élèves/magasiniers ayant au moins un
    # droit ponctuel matériel actif.
    return _request_has_any_material_grant(request)


def _json_name_from_request(request):
    # Le bouton + Ajouter envoie normalement du FormData, mais cette fonction
    # accepte aussi du JSON afin d'éviter une réponse HTML inattendue si le
    # navigateur ou un proxy modifie le Content-Type.
    name = (request.POST.get('name') or '').strip()
    if not name and request.body:
        try:
            payload = json.loads(request.body.decode('utf-8'))
            name = (payload.get('name') or '').strip()
        except Exception:
            name = ''
    return name


def _create_lookup_response(request, model_cls, label='élément'):
    try:
        if request.method != 'POST':
            return JsonResponse({'ok': False, 'error': 'Méthode non autorisée.'}, status=405)
        if not _can_create_lookup_item(request):
            return JsonResponse({
                'ok': False,
                'error': 'Droit insuffisant : l’ajout est réservé aux professeurs/admins ou aux élèves disposant d’un droit ponctuel matériel actif.'
            }, status=403)
        name = _json_name_from_request(request)
        if not name:
            return JsonResponse({'ok': False, 'error': f'Nom de {label} obligatoire.'}, status=400)
        existing = model_cls.objects.filter(name__iexact=name).first()
        if existing:
            return JsonResponse({'ok': True, 'id': existing.id, 'name': existing.name, 'created': False, 'message': f'{label.capitalize()} déjà existant, sélectionné.'})
        obj = model_cls.objects.create(name=name)
        return JsonResponse({'ok': True, 'id': obj.id, 'name': obj.name, 'created': True, 'message': f'{label.capitalize()} ajouté.'})
    except Exception as exc:
        # Ne jamais renvoyer une page HTML au JavaScript : sinon l’interface affiche
        # simplement “réponse serveur invalide”. On renvoie une erreur JSON lisible.
        return JsonResponse({'ok': False, 'error': f'Erreur serveur lors de l’ajout de {label} : {exc}'}, status=500)


@csrf_exempt
def api_category_create(request):
    return _create_lookup_response(request, Category, label='catégorie')


@csrf_exempt
def api_location_create(request):
    return _create_lookup_response(request, Location, label='emplacement')



def equipment_create(request):
    actor, denied = require_material_permission(request, 'can_create_equipment', 'Création de matériel non autorisée pour ce profil ou cette période.')
    if denied:
        return denied
    form = EquipmentForm(request.POST or None, request.FILES or None)
    if request.method == 'POST' and form.is_valid():
        requested_code = (form.cleaned_data.get('code') or '').strip()

        if requested_code and Equipment.objects.filter(code__iexact=requested_code).exists():
            form.add_error('code', 'Un matériel avec ce code existe déjà.')
            messages.warning(request, 'Matériel déjà existant : vérifie le code inventaire avant de créer une nouvelle fiche.')
        else:
            try:
                equipment = form.save()
            except IntegrityError:
                form.add_error('code', 'Un matériel avec ce code existe déjà ou le code généré est déjà utilisé.')
                messages.warning(request, 'Matériel déjà existant : vérifie le code inventaire.')
            else:
                messages.success(request, f'Fiche matériel créée : {equipment.code} — {equipment.name}.')
                return redirect('equipment_detail', code=equipment.code)
    return render(request, 'inventory/equipment_form.html', {'form': form, 'mode': 'create'})



def _candidate_duplicate_code(source_equipment):
    """Code candidat indépendant pour une duplication.
    On laisse la logique métier ToolMag générer le préfixe par catégorie.
    """
    return next_equipment_code_for_category(source_equipment.category)


def _duplicate_preview_rows(source_equipment, form):
    return {
        'source_code': source_equipment.code,
        'source_name': source_equipment.name,
        'new_code': (form.cleaned_data.get('new_code') or _candidate_duplicate_code(source_equipment)).strip(),
        'new_name': form.cleaned_data.get('new_name') or f'{source_equipment.name} — copie',
        'components_count': source_equipment.components.count() if form.cleaned_data.get('copy_components') else 0,
        'documents_count': source_equipment.documents.filter(active=True).count() if form.cleaned_data.get('copy_documents') else 0,
        'status': 'À vérifier',
        'history_copied': False,
        'storage_copied': bool(form.cleaned_data.get('copy_storage')),
        'photo_copied': bool(form.cleaned_data.get('copy_photo') and source_equipment.photo),
    }


def equipment_duplicate(request, code):
    """Duplique un matériel composé ou un modèle avec prévisualisation obligatoire.

    La duplication crée un nouvel équipement indépendant :
    - informations générales copiées ;
    - composants attendus copiés si demandé ;
    - documents modèles copiés si demandé ;
    - consignes / notes / périodicités d'inventaire copiées ;
    - nouveau code inventaire généré ou saisi ;
    - statut remis à "À vérifier" ;
    - aucun historique d'emprunt, maintenance, casse, intervention ou contrôle n'est copié.
    """
    actor, denied = require_material_permission(request, 'can_create_equipment', 'Duplication de matériel non autorisée pour ce profil ou cette période.')
    if denied:
        return denied
    source = get_object_or_404(Equipment.objects.select_related('category', 'location').prefetch_related('components', 'documents'), code=code)
    initial = {
        'new_name': f'{source.name} — copie',
        'new_code': _candidate_duplicate_code(source),
        'copy_components': True,
        'copy_documents': True,
        'copy_photo': False,
        'copy_storage': False,
    }
    form = EquipmentDuplicateForm(request.POST or None, initial=initial)
    preview = None
    if request.method == 'POST' and form.is_valid():
        requested_code = (form.cleaned_data.get('new_code') or '').strip()
        if requested_code and Equipment.objects.filter(code=requested_code).exists():
            form.add_error('new_code', 'Ce code inventaire existe déjà. Laisse vide ou indique un code libre.')
        else:
            preview = _duplicate_preview_rows(source, form)
            if form.cleaned_data.get('confirm_creation'):
                with transaction.atomic():
                    new_equipment = Equipment(
                        code=requested_code or '',
                        name=form.cleaned_data['new_name'],
                        equipment_type=source.equipment_type,
                        category=source.category,
                        brand=source.brand,
                        model=source.model,
                        serial_number='',
                        location=source.location if form.cleaned_data.get('copy_storage') else None,
                        description=source.description,
                        status=Equipment.Status.TO_VERIFY,
                        current_condition=Equipment.Condition.WATCH,
                        inventory_required_out=source.inventory_required_out,
                        inventory_required_return=source.inventory_required_return,
                        sensitive=source.sensitive,
                        display_on_public_screen=source.display_on_public_screen,
                        secure_storage=source.secure_storage if form.cleaned_data.get('copy_storage') else False,
                        secure_cabinet=source.secure_cabinet if form.cleaned_data.get('copy_storage') else '',
                        secure_locker=source.secure_locker if form.cleaned_data.get('copy_storage') else '',
                        notes=((source.notes or '') + f"\n\nDupliqué depuis {source.code} — historique non copié.").strip(),
                    )
                    if form.cleaned_data.get('copy_photo') and source.photo:
                        new_equipment.photo = source.photo.name
                    new_equipment.save()
                    if form.cleaned_data.get('copy_components'):
                        for comp in source.components.all():
                            Component.objects.create(
                                equipment=new_equipment,
                                line_type=getattr(comp, 'line_type', Component.LineType.COMPONENT),
                                name=comp.name,
                                section_label=getattr(comp, 'section_label', ''),
                                required=comp.required if getattr(comp, 'line_type', Component.LineType.COMPONENT) == Component.LineType.COMPONENT else False,
                                expected_quantity=comp.expected_quantity if getattr(comp, 'line_type', Component.LineType.COMPONENT) == Component.LineType.COMPONENT else 1,
                                default_condition=Equipment.Condition.WATCH,
                                photo=comp.photo.name if comp.photo else '',
                                sort_order=comp.sort_order,
                                mobile_page_break=getattr(comp, 'mobile_page_break', True),
                                inventory_required=getattr(comp, 'inventory_required', True) if getattr(comp, 'line_type', Component.LineType.COMPONENT) == Component.LineType.COMPONENT else False,
                            )
                    if form.cleaned_data.get('copy_documents'):
                        for doc in source.documents.filter(active=True):
                            EquipmentDocument.objects.create(
                                equipment=new_equipment,
                                title=doc.title,
                                document_type=doc.document_type,
                                file=doc.file.name,
                                description=doc.description,
                                active=doc.active,
                                sort_order=doc.sort_order,
                            )
                messages.success(request, f'Matériel dupliqué : {new_equipment.code} — {new_equipment.name}. Statut initial : À vérifier.')
                return redirect('equipment_detail', code=new_equipment.code)
            else:
                messages.info(request, 'Prévisualisation générée. Coche la confirmation puis valide pour créer le nouveau matériel.')
    return render(request, 'inventory/equipment_duplicate.html', {'source': source, 'form': form, 'preview': preview})

def equipment_edit(request, code):
    actor, denied = require_material_permission(request, 'can_edit_equipment', 'Modification de fiche matériel non autorisée pour ce profil ou cette période.')
    if denied:
        return denied
    equipment = get_object_or_404(Equipment, code=code)
    form = EquipmentForm(request.POST or None, request.FILES or None, instance=equipment)
    if request.method == 'POST' and form.is_valid():
        equipment = form.save()
        messages.success(request, f'Fiche matériel mise à jour : {equipment.code}.')
        return redirect('equipment_detail', code=equipment.code)
    return render(request, 'inventory/equipment_form.html', {'form': form, 'equipment': equipment, 'mode': 'edit'})


def equipment_components_edit(request, code):
    actor, denied = require_material_permission(request, 'can_edit_components', 'Modification des composants non autorisée pour ce profil ou cette période.')
    if denied:
        return denied
    equipment = get_object_or_404(Equipment.objects.prefetch_related('components'), code=code)
    edit_id = request.GET.get('edit')
    component = None
    if edit_id:
        component = get_object_or_404(Component, id=edit_id, equipment=equipment)
    if request.method == 'POST' and request.POST.get('reorder_components'):
        raw_order = request.POST.get('component_order', '')
        ids = [item.strip() for item in raw_order.split(',') if item.strip()]
        with transaction.atomic():
            for index, comp_id in enumerate(ids, start=1):
                Component.objects.filter(id=comp_id, equipment=equipment).update(sort_order=index * 10)
        messages.success(request, 'Ordre des lignes mis à jour.')
        return redirect('equipment_components_edit', code=equipment.code)

    if request.method == 'POST' and request.POST.get('delete_id') and is_super_admin_person(actor):
        comp = get_object_or_404(Component, id=request.POST.get('delete_id'), equipment=equipment)
        comp.delete()
        messages.success(request, 'Composant supprimé.')
        return redirect('equipment_components_edit', code=equipment.code)
    form = ComponentEditForm(request.POST or None, request.FILES or None, instance=component)
    if request.method == 'POST' and not request.POST.get('delete_id') and form.is_valid():
        comp = form.save(commit=False)
        comp.equipment = equipment

        duplicate_qs = Component.objects.filter(
            equipment=equipment,
            name__iexact=(comp.name or '').strip()
        )

        if component:
            duplicate_qs = duplicate_qs.exclude(pk=component.pk)

        if duplicate_qs.exists():
            form.add_error('name', 'Une ligne avec ce nom existe déjà dans ce matériel.')
            messages.warning(
                request,
                'Composant, section ou note déjà existant dans ce matériel : vérifie la structure avant d’ajouter une nouvelle ligne.'
            )
        else:
            if comp.line_type in [Component.LineType.SECTION, Component.LineType.NOTE]:
                comp.required = False
                comp.expected_quantity = 1
                comp.inventory_required = False
                comp.default_condition = Equipment.Condition.GOOD

            try:
                comp.save()
            except IntegrityError:
                form.add_error('name', 'Une ligne avec ce nom existe déjà dans ce matériel.')
                messages.warning(
                    request,
                    'Composant, section ou note déjà existant dans ce matériel.'
                )
            else:
                messages.success(request, 'Composant enregistré.')
                return redirect('equipment_components_edit', code=equipment.code)
    return render(request, 'inventory/equipment_components_form.html', {'equipment': equipment, 'form': form, 'component': component, 'components': _all_component_lines_for_equipment(equipment), 'can_delete': is_super_admin_person(actor)})


def equipment_documents_edit(request, code):
    actor, denied = require_material_permission(request, 'can_add_document', 'Ajout de documents non autorisé pour ce profil ou cette période.')
    if denied:
        return denied
    equipment = get_object_or_404(Equipment.objects.prefetch_related('documents'), code=code)
    edit_id = request.GET.get('edit')
    document = None
    if edit_id:
        document = get_object_or_404(EquipmentDocument, id=edit_id, equipment=equipment)
    if request.method == 'POST' and request.POST.get('delete_id') and is_super_admin_person(actor):
        doc = get_object_or_404(EquipmentDocument, id=request.POST.get('delete_id'), equipment=equipment)
        doc.delete()
        messages.success(request, 'Document supprimé.')
        return redirect('equipment_documents_edit', code=equipment.code)
    form = EquipmentDocumentEditForm(request.POST or None, request.FILES or None, instance=document)
    if request.method == 'POST' and not request.POST.get('delete_id') and form.is_valid():
        doc = form.save(commit=False)
        doc.equipment = equipment
        doc.save()
        messages.success(request, 'Document enregistré.')
        return redirect('equipment_documents_edit', code=equipment.code)
    return render(request, 'inventory/equipment_documents_form.html', {'equipment': equipment, 'form': form, 'document': document, 'documents': equipment.documents.all(), 'can_delete': is_super_admin_person(actor)})


def equipment_detail(request, code):
    borrower = current_borrower(request)
    storekeeper = current_storekeeper(request)
    equipment = get_object_or_404(Equipment.objects.prefetch_related('components', 'documents'), code=code)
    documents = equipment.documents.filter(active=True)
    loans = equipment.loans.select_related('borrower', 'checkout_storekeeper', 'return_storekeeper')[:20]
    open_loan = Loan.objects.filter(equipment=equipment, status=Loan.LoanStatus.OPEN).select_related('borrower', 'checkout_storekeeper').first()
    loan_inventory_status = _loan_user_inventory_status(open_loan) if open_loan else None
    control_url = absolute_uri(request, 'equipment_control', code=equipment.code)
    inventory_entry_url = _user_inventory_entry_url(request, equipment)
    page_qr = make_qr_data_uri(inventory_entry_url)
    locker_info = locker_availability(request, equipment=equipment, context=LockerOpenLog.Context.DETAIL)
    locker_logs = equipment.locker_logs.select_related('storekeeper', 'terminal')[:10]
    repairs = equipment.repairs.select_related('storekeeper')[:20]
    interventions = equipment.interventions.select_related('storekeeper')[:20]
    user_inventories = equipment.user_inventories.select_related('borrower', 'applied_by', 'loan')[:20]
    return render(request, 'inventory/equipment_detail.html', {
        'equipment': equipment,
        'documents': documents,
        'loans': loans,
        'open_loan': open_loan,
        'loan_inventory_status': loan_inventory_status,
        'has_inventory_components': _equipment_has_inventory_components(equipment),
        'locker_info': locker_info,
        'locker_logs': locker_logs,
        'repairs': repairs,
        'interventions': interventions,
        'user_inventories': user_inventories,
        'current_storekeeper': storekeeper,
        'control_url': control_url,
        'inventory_entry_url': inventory_entry_url,
        'page_qr': page_qr,
        'current_borrower': borrower,
        'can_edit_equipment': can_edit_equipment_person(storekeeper or borrower),
        'can_edit_components': can_edit_components_person(storekeeper or borrower),
        'can_add_document': can_add_document_person(storekeeper or borrower),
    })


def equipment_control(request, code):
    borrower = current_borrower(request)
    storekeeper = current_storekeeper(request)
    equipment = get_object_or_404(Equipment.objects.prefetch_related('components', 'documents'), code=code)
    open_loan = Loan.objects.filter(equipment=equipment, status=Loan.LoanStatus.OPEN).select_related('borrower', 'checkout_storekeeper').first()
    control_url = absolute_uri(request, 'equipment_control', code=equipment.code)
    inventory_entry_url = _user_inventory_entry_url(request, equipment)
    page_qr = make_qr_data_uri(inventory_entry_url)
    documents = equipment.documents.filter(active=True)
    locker_info = locker_availability(request, equipment=equipment, context=LockerOpenLog.Context.DETAIL)
    loan_inventory_status = _loan_user_inventory_status(open_loan) if open_loan else None
    return render(request, 'inventory/equipment_control.html', {
        'equipment': equipment,
        'documents': documents,
        'locker_info': locker_info,
        'current_storekeeper': storekeeper,
        'open_loan': open_loan,
        'loan_inventory_status': loan_inventory_status,
        'control_url': control_url,
        'inventory_entry_url': inventory_entry_url,
        'page_qr': page_qr,
        'has_inventory_components': _equipment_has_inventory_components(equipment),
        'current_borrower': borrower,
    })


def _inventory_actor_for_check(request, storekeeper=None):
    """Personne qui réalise physiquement l'inventaire d'un kit.
    Si un emprunteur est connecté sur la tablette, il est l'acteur de l'inventaire.
    Sinon, c'est le magasinier actif.
    """
    borrower = current_borrower(request)
    if borrower:
        return borrower, 'utilisateur'
    if storekeeper:
        return storekeeper, 'magasinier'
    return None, ''



def _equipment_has_inventory_components(equipment):
    """Un inventaire utilisateur n'est utile que si le matériel possède réellement
    des composants déclarés. On ne s'appuie pas uniquement sur le type matériel,
    car certaines fiches peuvent être typées « kit » sans composants ou inversement.
    """
    if not equipment:
        return False
    try:
        return equipment.components.exists()
    except Exception:
        return False


def _user_inventory_entry_url(request, equipment):
    return absolute_uri(request, 'user_inventory_auto', code=equipment.code)


def _latest_user_inventory(equipment, borrower, inventory_type, loan=None):
    if not equipment or not borrower:
        return None
    qs = UserInventory.objects.filter(
        equipment=equipment,
        borrower=borrower,
        inventory_type=inventory_type,
        status__in=[UserInventory.InventoryStatus.SUBMITTED, UserInventory.InventoryStatus.APPLIED],
    ).prefetch_related('items__component')
    if loan is not None:
        qs = qs.filter(loan=loan)
    else:
        qs = qs.filter(loan__isnull=True)
    return qs.order_by('-submitted_at').first()


def _latest_submitted_user_inventory(equipment, borrower, inventory_type, loan=None):
    if not equipment or not borrower:
        return None
    qs = UserInventory.objects.filter(
        equipment=equipment,
        borrower=borrower,
        inventory_type=inventory_type,
        status=UserInventory.InventoryStatus.SUBMITTED,
    ).prefetch_related('items__component')
    if loan is not None:
        qs = qs.filter(loan=loan)
    else:
        qs = qs.filter(loan__isnull=True)
    return qs.order_by('-submitted_at').first()


def _loan_user_inventory_status(loan):
    """Return compact status flags for dynamic display and storekeeper preview.

    V25: for simple equipment without components, inventory is not applicable (S.O.).
    """
    if not loan:
        return {'out_done': False, 'return_done': False, 'out_inventory': None, 'return_inventory': None, 'not_applicable': False}
    if not _equipment_has_inventory_components(loan.equipment):
        return {'out_done': None, 'return_done': None, 'out_inventory': None, 'return_inventory': None, 'not_applicable': True}
    out_inventory = _latest_user_inventory(loan.equipment, loan.borrower, UserInventory.InventoryType.OUT, loan=loan)
    return_inventory = _latest_user_inventory(loan.equipment, loan.borrower, UserInventory.InventoryType.RETURN, loan=loan)
    return {
        'out_done': bool(out_inventory),
        'return_done': bool(return_inventory),
        'out_inventory': out_inventory,
        'return_inventory': return_inventory,
        'not_applicable': False,
    }


def _inventory_items_map(inventory):
    if not inventory:
        return {}
    return {item.component_id: item for item in inventory.items.all()}


def _inventory_components_for_equipment(equipment):
    """Lignes réellement contrôlables en inventaire.
    Les sections et notes structurent l’affichage mais ne génèrent aucun check.
    """
    if not equipment:
        return Component.objects.none()
    return equipment.components.filter(
        line_type=Component.LineType.COMPONENT,
        inventory_required=True,
    ).order_by('sort_order', 'name')


def _all_component_lines_for_equipment(equipment):
    """Toutes les lignes de structure pour l’édition : sections, notes, composants."""
    if not equipment:
        return Component.objects.none()
    return equipment.components.all().order_by('sort_order', 'name')


def _component_rows(components, inventory=None):
    items = _inventory_items_map(inventory)
    rows = []
    for component in components:
        rows.append({
            'component': component,
            'item': items.get(component.id),
        })
    return rows


def _latest_component_checks_for_equipment(equipment, prefer_return=True):
    """Dernier inventaire validé côté magasinier pour préremplir les formulaires magasinier.

    Retourne un dictionnaire {component_id: ComponentCheck}. On privilégie par défaut
    le dernier contrôle de retour, puis à défaut le dernier contrôle de sortie.
    """
    if not equipment:
        return {}
    loans = list(equipment.loans.prefetch_related('component_checks__component').order_by('-checked_out_at')[:8])
    order = [ComponentCheck.CheckType.RETURN, ComponentCheck.CheckType.OUT] if prefer_return else [ComponentCheck.CheckType.OUT, ComponentCheck.CheckType.RETURN]
    for loan in loans:
        for check_type in order:
            checks = [c for c in loan.component_checks.all() if c.check_type == check_type]
            if checks:
                return {c.component_id: c for c in checks}
    return {}


def _component_rows_from_last_check(components, equipment=None, prefer_return=True):
    checks = _latest_component_checks_for_equipment(equipment, prefer_return=prefer_return)
    rows = []
    for component in components:
        rows.append({'component': component, 'item': checks.get(component.id)})
    return rows




def _component_condition_choices():
    """Choix de statut pour une ligne de composant, avec l'état Absent.
    Utilisé uniquement pour les inventaires de composants.
    """
    return Equipment.Condition.choices


def _global_condition_choices():
    """Choix d'état global du matériel, sans l'état Absent.
    L'absence se traite composant par composant.
    """
    return [(value, label) for value, label in Equipment.Condition.choices if value != Equipment.Condition.ABSENT]


def _component_post_values(request, component):
    """Normalise la saisie d'une ligne composant.
    - Si le statut est Absent, la présence devient False.
    - Si la présence n'est pas cochée, le statut devient Absent.
    """
    condition = request.POST.get(f'component_{component.id}_condition', component.default_condition) or component.default_condition
    present = request.POST.get(f'component_{component.id}_present') == 'on'
    if condition == Equipment.Condition.ABSENT:
        present = False
    elif not present:
        condition = Equipment.Condition.ABSENT
    comment = request.POST.get(f'component_{component.id}_comment', '')
    return present, condition, comment

def _apply_user_inventory_if_possible(user_inventory, storekeeper):
    if user_inventory and user_inventory.status == UserInventory.InventoryStatus.SUBMITTED:
        user_inventory.status = UserInventory.InventoryStatus.APPLIED
        user_inventory.applied_by = storekeeper
        user_inventory.applied_at = timezone.now()
        user_inventory.save(update_fields=['status', 'applied_by', 'applied_at', 'updated_at'])


def _submitted_inventory_from_post(request, equipment=None, borrower=None, inventory_type=None, loan=None):
    inv_id = request.POST.get('user_inventory_id')
    if not inv_id:
        return None
    qs = UserInventory.objects.filter(id=inv_id, status=UserInventory.InventoryStatus.SUBMITTED).prefetch_related('items__component')
    if equipment:
        qs = qs.filter(equipment=equipment)
    if borrower:
        qs = qs.filter(borrower=borrower)
    if inventory_type:
        qs = qs.filter(inventory_type=inventory_type)
    if loan is not None:
        qs = qs.filter(loan=loan)
    return qs.first()


def checkout(request):
    equipment = None
    components = []
    storekeeper = current_storekeeper(request)
    borrower = current_borrower(request)
    selected_user_inventory = None
    if not storekeeper:
        messages.warning(request, 'Seul un magasinier connecté peut valider une sortie de matériel.')
        return redirect(f"{reverse('storekeeper_login')}?next={request.path}")
    if request.method == 'POST':
        form = CheckoutForm(request.POST, current_storekeeper=storekeeper, current_borrower=borrower)
        if form.is_valid():
            equipment = form.cleaned_data['equipment']
            selected_user_inventory = _submitted_inventory_from_post(request, equipment=equipment, borrower=form.cleaned_data['borrower'], inventory_type=UserInventory.InventoryType.OUT, loan=None)
            loan = Loan.objects.create(
                equipment=equipment,
                borrower=form.cleaned_data['borrower'],
                checkout_storekeeper=form.cleaned_data['storekeeper'],
                due_at=form.cleaned_data['due_at'],
                condition_out=form.cleaned_data['condition_out'],
                comment_out=form.cleaned_data['comment_out'],
            )
            if selected_user_inventory:
                selected_user_inventory.loan = loan
                selected_user_inventory.save(update_fields=['loan', 'updated_at'])
                checked_by, checked_by_role = selected_user_inventory.borrower, 'utilisateur'
            else:
                checked_by, checked_by_role = _inventory_actor_for_check(request, storekeeper=storekeeper)
            for component in equipment.components.filter(line_type=Component.LineType.COMPONENT, inventory_required=True):
                present, condition, comment = _component_post_values(request, component)
                ComponentCheck.objects.create(
                    loan=loan,
                    component=component,
                    check_type=ComponentCheck.CheckType.OUT,
                    present=present,
                    quantity=component.expected_quantity if present else 0,
                    condition=condition,
                    comment=comment,
                    checked_by=checked_by,
                    checked_by_role=checked_by_role,
                )
            _apply_user_inventory_if_possible(selected_user_inventory, storekeeper)
            equipment.status = Equipment.Status.OUT
            equipment.current_condition = form.cleaned_data['condition_out']
            equipment.save()
            messages.success(request, 'Sortie enregistrée par le magasinier connecté.')
            return redirect('equipment_detail', code=equipment.code)
    else:
        initial = {}
        if request.GET.get('equipment'):
            initial['equipment_code'] = request.GET['equipment']
            try:
                equipment = Equipment.objects.get(code=request.GET['equipment'])
                components = _inventory_components_for_equipment(equipment)
                pending_inventory = UserInventory.objects.filter(
                    equipment=equipment,
                    inventory_type=UserInventory.InventoryType.OUT,
                    status=UserInventory.InventoryStatus.SUBMITTED,
                    loan__isnull=True,
                ).select_related('borrower').order_by('-submitted_at').first()
                if pending_inventory and not borrower:
                    initial['borrower_code'] = pending_inventory.borrower.code
                    initial['condition_out'] = pending_inventory.global_condition
            except Equipment.DoesNotExist:
                pass
        if storekeeper:
            initial['storekeeper_code'] = storekeeper.code
        if borrower:
            initial['borrower_code'] = borrower.code
        form = CheckoutForm(initial=initial, current_storekeeper=storekeeper, current_borrower=borrower)
    page_qr = None
    control_url = None
    selected_user_inventory = None
    if equipment and not components:
        components = _inventory_components_for_equipment(equipment)
    if equipment:
        form_borrower_code = None
        try:
            form_borrower_code = form['borrower_code'].value()
        except Exception:
            pass
        form_borrower = borrower or Person.objects.filter(code=form_borrower_code, active=True, archived=False).first()
        selected_user_inventory = _latest_user_inventory(equipment, form_borrower, UserInventory.InventoryType.OUT, loan=None)
        if not selected_user_inventory and not form_borrower:
            selected_user_inventory = UserInventory.objects.filter(equipment=equipment, inventory_type=UserInventory.InventoryType.OUT, status=UserInventory.InventoryStatus.SUBMITTED, loan__isnull=True).select_related('borrower').prefetch_related('items__component').order_by('-submitted_at').first()
        control_url = _user_inventory_entry_url(request, equipment)
        page_qr = make_qr_data_uri(control_url)
    inventory_actor, inventory_actor_role = _inventory_actor_for_check(request, storekeeper=storekeeper)
    return render(request, 'inventory/checkout.html', {
        'form': form,
        'equipment': equipment,
        'components': components,
        'component_rows': _component_rows(components, selected_user_inventory) if selected_user_inventory else _component_rows_from_last_check(components, equipment=equipment, prefer_return=True),
        'selected_user_inventory': selected_user_inventory,
        'conditions': _global_condition_choices(),
        'component_conditions': _component_condition_choices(),
        'page_qr': page_qr,
        'control_url': control_url,
        'current_storekeeper': storekeeper,
        'current_borrower': borrower,
        'inventory_actor': inventory_actor,
        'inventory_actor_role': inventory_actor_role,
        'locker_info': locker_availability(request, equipment=equipment, context=LockerOpenLog.Context.CHECKOUT) if equipment else None,
    })

def return_equipment(request):
    loan = None
    components = []
    storekeeper = current_storekeeper(request)
    borrower = current_borrower(request)
    selected_user_inventory = None
    if not storekeeper:
        messages.warning(request, 'Seul un magasinier connecté peut valider le retour de matériel. L’inventaire peut être réalisé par l’utilisateur connecté ou par le magasinier.')
        return redirect(f"{reverse('storekeeper_login')}?next={request.path}")
    if request.method == 'POST':
        form = ReturnForm(request.POST, current_storekeeper=storekeeper)
        if form.is_valid():
            loan = form.cleaned_data['loan']
            equipment = form.cleaned_data['equipment']
            selected_user_inventory = _submitted_inventory_from_post(request, equipment=equipment, borrower=loan.borrower, inventory_type=UserInventory.InventoryType.RETURN, loan=loan)
            if selected_user_inventory:
                checked_by, checked_by_role = selected_user_inventory.borrower, 'utilisateur'
            else:
                checked_by, checked_by_role = _inventory_actor_for_check(request, storekeeper=storekeeper)
            has_problem = False
            for component in equipment.components.filter(line_type=Component.LineType.COMPONENT, inventory_required=True):
                present, condition, comment = _component_post_values(request, component)
                ComponentCheck.objects.create(
                    loan=loan,
                    component=component,
                    check_type=ComponentCheck.CheckType.RETURN,
                    present=present,
                    quantity=component.expected_quantity if present else 0,
                    condition=condition,
                    comment=comment,
                    checked_by=checked_by,
                    checked_by_role=checked_by_role,
                )
                if component.required and not present:
                    has_problem = True
                if condition in [Equipment.Condition.DAMAGED, Equipment.Condition.DANGEROUS, Equipment.Condition.INCOMPLETE, Equipment.Condition.ABSENT]:
                    has_problem = True
            _apply_user_inventory_if_possible(selected_user_inventory, storekeeper)
            loan.return_storekeeper = form.cleaned_data['storekeeper']
            loan.returned_at = timezone.now()
            loan.condition_return = form.cleaned_data['condition_return']
            loan.comment_return = form.cleaned_data['comment_return']
            loan.status = Loan.LoanStatus.PROBLEM if has_problem else Loan.LoanStatus.CLOSED
            loan.save()
            equipment.current_condition = form.cleaned_data['condition_return']
            action = form.cleaned_data['action']
            equipment.status = {
                'available': Equipment.Status.AVAILABLE,
                'incomplete': Equipment.Status.INCOMPLETE,
                'maintenance': Equipment.Status.MAINTENANCE,
                'out_of_service': Equipment.Status.OUT_OF_SERVICE,
            }[action]
            equipment.save()
            messages.success(request, 'Retour enregistré par le magasinier connecté.')
            return redirect('equipment_detail', code=equipment.code)
    else:
        initial = {}
        if request.GET.get('equipment'):
            initial['equipment_code'] = request.GET['equipment']
            try:
                equipment = Equipment.objects.get(code=request.GET['equipment'])
                loan = Loan.objects.get(equipment=equipment, status=Loan.LoanStatus.OPEN)
                components = _inventory_components_for_equipment(equipment)
                pending_inventory = _latest_user_inventory(equipment, loan.borrower, UserInventory.InventoryType.RETURN, loan=loan)
                if pending_inventory:
                    initial['condition_return'] = pending_inventory.global_condition
            except Exception:
                pass
        if storekeeper:
            initial['storekeeper_code'] = storekeeper.code
        form = ReturnForm(initial=initial, current_storekeeper=storekeeper)
    page_qr = None
    control_url = None
    selected_user_inventory = None
    if loan:
        selected_user_inventory = _latest_user_inventory(loan.equipment, loan.borrower, UserInventory.InventoryType.RETURN, loan=loan)
        control_url = _user_inventory_entry_url(request, loan.equipment)
        page_qr = make_qr_data_uri(control_url)
    inventory_actor, inventory_actor_role = _inventory_actor_for_check(request, storekeeper=storekeeper)
    return render(request, 'inventory/return.html', {
        'form': form,
        'loan': loan,
        'equipment': loan.equipment if loan else None,
        'components': components,
        'component_rows': _component_rows(components, selected_user_inventory) if selected_user_inventory else _component_rows_from_last_check(components, equipment=loan.equipment if loan else None, prefer_return=False),
        'selected_user_inventory': selected_user_inventory,
        'conditions': _global_condition_choices(),
        'component_conditions': _component_condition_choices(),
        'page_qr': page_qr,
        'control_url': control_url,
        'current_storekeeper': storekeeper,
        'current_borrower': borrower,
        'inventory_actor': inventory_actor,
        'inventory_actor_role': inventory_actor_role,
        'locker_info': locker_availability(request, equipment=loan.equipment, context=LockerOpenLog.Context.RETURN) if loan else None,
    })



def user_inventory_auto(request, code):
    """Point d'entrée prévu pour le QR code collé sur le matériel.
    Il choisit automatiquement l'inventaire utilisateur de sortie ou de retour.
    Si l'utilisateur n'est pas connecté, il est d'abord envoyé vers la page de connexion.
    """
    equipment = get_object_or_404(Equipment.objects.prefetch_related('components', 'documents'), code=code)
    if not _equipment_has_inventory_components(equipment):
        messages.info(request, 'Ce matériel ne possède pas de composants : aucun inventaire utilisateur n’est nécessaire.')
        return redirect('equipment_control', code=equipment.code)
    borrower = current_borrower(request)
    if not borrower:
        messages.warning(request, 'Connecte l’utilisateur avant de faire l’inventaire du matériel.')
        return redirect(f"{reverse('user_login')}?next={request.path}")
    open_loan = Loan.objects.filter(equipment=equipment, status=Loan.LoanStatus.OPEN).select_related('borrower').first()
    if open_loan:
        if open_loan.borrower_id != borrower.id:
            messages.error(request, f"Ce matériel est actuellement sorti au nom de {open_loan.borrower}. Connecte cet emprunteur pour faire l’inventaire utilisateur.")
            # On renvoie vers la connexion utilisateur plutôt que vers la page de contrôle,
            # afin que le QR code puisse toujours servir de point d’entrée inventaire.
            return redirect(f"{reverse('user_login')}?next={request.path}")
        # Règle métier V22 : après validation de sortie par le magasinier,
        # le premier scan utilisateur sert à faire l’inventaire utilisateur de sortie.
        # Une fois cet inventaire fait, le scan suivant ouvre l’inventaire utilisateur de retour.
        out_inventory = _latest_user_inventory(equipment, borrower, UserInventory.InventoryType.OUT, loan=open_loan)
        if not out_inventory:
            return redirect('user_inventory_equipment', code=equipment.code, inventory_type=UserInventory.InventoryType.OUT)
        return redirect('user_inventory_equipment', code=equipment.code, inventory_type=UserInventory.InventoryType.RETURN)
    if equipment.status in [Equipment.Status.AVAILABLE, Equipment.Status.INCOMPLETE]:
        messages.error(request, 'Inventaire utilisateur de sortie impossible : la sortie doit d’abord être validée par un magasinier.')
        return redirect('equipment_control', code=equipment.code)
    messages.error(request, 'Inventaire utilisateur impossible : ce matériel n’est pas sorti par l’utilisateur connecté.')
    return redirect('equipment_detail', code=equipment.code)


def user_inventory_equipment(request, code, inventory_type):
    borrower = current_borrower(request)
    if not borrower:
        messages.warning(request, 'Connexion utilisateur obligatoire pour faire un inventaire utilisateur.')
        return redirect(f"{reverse('user_login')}?next={request.path}")
    equipment = get_object_or_404(Equipment.objects.prefetch_related('components', 'documents'), code=code)
    if not _equipment_has_inventory_components(equipment):
        messages.warning(request, 'Ce matériel n’a pas de composants à inventorier : l’inventaire utilisateur est réservé aux matériels composés.')
        return redirect('equipment_control', code=equipment.code)
    if inventory_type not in [UserInventory.InventoryType.OUT, UserInventory.InventoryType.RETURN]:
        return HttpResponseForbidden('Type d’inventaire invalide.')
    loan = None
    if inventory_type == UserInventory.InventoryType.OUT:
        open_loan = Loan.objects.filter(equipment=equipment, status=Loan.LoanStatus.OPEN).select_related('borrower').first()
        if open_loan:
            if open_loan.borrower_id != borrower.id:
                messages.error(request, f'Inventaire de sortie refusé : ce matériel est sorti au nom de {open_loan.borrower}. Connecte cet emprunteur.')
                return redirect(f"{reverse('user_login')}?next={request.path}")
            loan = open_loan
        else:
            messages.error(request, 'Inventaire utilisateur de sortie impossible : le matériel doit d’abord être sorti par un magasinier au nom de cet utilisateur.')
            return redirect('equipment_control', code=equipment.code)
    else:
        loan = Loan.objects.filter(equipment=equipment, status=Loan.LoanStatus.OPEN).select_related('borrower').first()
        if not loan:
            messages.error(request, 'Inventaire de retour impossible : aucune sortie ouverte pour ce matériel.')
            return redirect('equipment_control', code=equipment.code)
        if loan.borrower_id != borrower.id:
            messages.error(request, f'Inventaire de retour refusé : ce matériel est sorti au nom de {loan.borrower}. Connecte cet emprunteur.')
            return redirect(f"{reverse('user_login')}?next={request.path}")
        out_inventory = _latest_user_inventory(equipment, borrower, UserInventory.InventoryType.OUT, loan=loan)
        if not out_inventory:
            messages.warning(request, 'Il faut d’abord faire l’inventaire utilisateur de sortie avant l’inventaire de retour.')
            return redirect('user_inventory_equipment', code=equipment.code, inventory_type=UserInventory.InventoryType.OUT)
    components = list(_inventory_components_for_equipment(equipment))
    documents = equipment.documents.filter(active=True)
    # V25 : l'inventaire utilisateur doit toujours être vierge.
    # Le magasinier, lui, peut relire un inventaire utilisateur ou le dernier contrôle connu.
    previous_inventory = None
    if request.method == 'POST':
        # On annule les précédents inventaires utilisateur non validés pour garder une version de référence claire.
        UserInventory.objects.filter(
            equipment=equipment,
            borrower=borrower,
            inventory_type=inventory_type,
            status=UserInventory.InventoryStatus.SUBMITTED,
            loan=loan,
        ).update(status=UserInventory.InventoryStatus.CANCELLED, updated_at=timezone.now())
        inventory = UserInventory.objects.create(
            equipment=equipment,
            loan=loan,
            borrower=borrower,
            inventory_type=inventory_type,
            status=UserInventory.InventoryStatus.SUBMITTED,
            submitted_at=timezone.now(),
            global_condition=request.POST.get('global_condition') or Equipment.Condition.GOOD,
            comment=request.POST.get('comment', ''),
        )
        for component in components:
            present, condition, comment = _component_post_values(request, component)
            UserInventoryItem.objects.create(
                inventory=inventory,
                component=component,
                present=present,
                quantity=component.expected_quantity if present else 0,
                condition=condition,
                comment=comment,
            )
        if inventory_type == UserInventory.InventoryType.OUT:
            if loan:
                messages.success(request, 'Inventaire utilisateur de sortie enregistré. Le formulaire d’inventaire retour est maintenant disponible pour préparer la rentrée du matériel.')
                return redirect('user_inventory_equipment', code=equipment.code, inventory_type=UserInventory.InventoryType.RETURN)
            messages.success(request, 'Inventaire utilisateur de sortie enregistré. Le magasinier pourra le relire et valider la sortie.')
            return redirect(f"{reverse('checkout')}?equipment={equipment.code}")
        messages.success(request, 'Inventaire utilisateur de retour enregistré. Le magasinier pourra le relire et valider le retour.')
        return redirect(f"{reverse('return_equipment')}?equipment={equipment.code}")
    return render(request, 'inventory/user_inventory.html', {
        'equipment': equipment,
        'documents': documents,
        'equipment_detail_url': reverse('equipment_detail', kwargs={'code': equipment.code}),
        'loan': loan,
        'loan_inventory_status': _loan_user_inventory_status(loan) if loan else None,
        'inventory_type': inventory_type,
        'inventory_label': 'sortie' if inventory_type == UserInventory.InventoryType.OUT else 'retour',
        'components': components,
        'component_rows': _component_rows(components, previous_inventory),
        'previous_inventory': previous_inventory,
        'conditions': _global_condition_choices(),
        'component_conditions': _component_condition_choices(),
        'current_borrower': borrower,
    })



def intervention_equipment(request, code):
    storekeeper = current_storekeeper(request)
    if not storekeeper:
        messages.warning(request, 'Connexion magasinier obligatoire pour enregistrer un bon d’intervention.')
        return redirect(f"{reverse('storekeeper_login')}?next={request.path}")
    equipment = get_object_or_404(Equipment.objects.prefetch_related('components', 'documents'), code=code)
    open_loan = Loan.objects.filter(equipment=equipment, status=Loan.LoanStatus.OPEN).first()
    if open_loan:
        messages.error(request, 'Impossible d’enregistrer un bon d’intervention sur un matériel actuellement sorti. Il faut d’abord enregistrer le retour.')
        return redirect('equipment_detail', code=equipment.code)
    form = InterventionForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        intervention = form.save(commit=False)
        intervention.equipment = equipment
        intervention.storekeeper = storekeeper
        intervention.intervention_at = timezone.now()
        intervention.save()
        equipment.status = intervention.target_status
        equipment.current_condition = intervention.resulting_condition
        equipment.save(update_fields=['status', 'current_condition', 'updated_at'])
        messages.success(request, f'Bon d’intervention enregistré : {intervention.get_intervention_type_display()} — {intervention.get_result_display()}.')
        return redirect('equipment_detail', code=equipment.code)
    locker_info = locker_availability(request, equipment=equipment, context=LockerOpenLog.Context.MAINTENANCE)
    recent_interventions = equipment.interventions.select_related('storekeeper')[:10]
    return render(request, 'inventory/intervention.html', {
        'form': form,
        'equipment': equipment,
        'current_storekeeper': storekeeper,
        'locker_info': locker_info,
        'recent_interventions': recent_interventions,
    })

def repair_equipment(request, code):
    storekeeper = current_storekeeper(request)
    if not storekeeper:
        messages.warning(request, 'Connexion magasinier obligatoire pour enregistrer un bon de réparation.')
        return redirect(f"{reverse('storekeeper_login')}?next={request.path}")
    equipment = get_object_or_404(Equipment.objects.prefetch_related('components', 'documents'), code=code)
    open_loan = Loan.objects.filter(equipment=equipment, status=Loan.LoanStatus.OPEN).first()
    if open_loan:
        messages.error(request, 'Impossible de réparer un matériel avec une sortie encore ouverte. Il faut d’abord enregistrer le retour.')
        return redirect('equipment_detail', code=equipment.code)
    if equipment.status not in [Equipment.Status.MAINTENANCE, Equipment.Status.INCOMPLETE, Equipment.Status.OUT_OF_SERVICE]:
        messages.info(request, 'La réparation est normalement prévue pour les matériels en maintenance, incomplets ou hors service. Tu peux tout de même tracer un contrôle de remise en état.')
    form = RepairForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        repair = form.save(commit=False)
        repair.equipment = equipment
        repair.storekeeper = storekeeper
        repair.repaired_at = timezone.now()
        repair.save()
        equipment.status = repair.target_status
        equipment.current_condition = repair.resulting_condition
        equipment.save(update_fields=['status', 'current_condition', 'updated_at'])
        if repair.result == RepairLog.Result.REPAIRED_AVAILABLE:
            messages.success(request, 'Bon de réparation enregistré : matériel dépanné et remis disponible.')
        else:
            messages.success(request, f'Bon de réparation enregistré : nouveau statut {equipment.get_status_display()}.')
        return redirect('equipment_detail', code=equipment.code)
    locker_info = locker_availability(request, equipment=equipment, context=LockerOpenLog.Context.MAINTENANCE)
    recent_repairs = equipment.repairs.select_related('storekeeper')[:10]
    return render(request, 'inventory/repair.html', {
        'form': form,
        'equipment': equipment,
        'current_storekeeper': storekeeper,
        'locker_info': locker_info,
        'recent_repairs': recent_repairs,
    })


def display_current_loans(request):
    now = timezone.now()
    loans = list(Loan.objects.filter(status=Loan.LoanStatus.OPEN, equipment__display_on_public_screen=True).select_related('equipment', 'borrower').order_by('due_at'))
    # Ajout V22 : statut des inventaires utilisateur de sortie et de retour sur l'affichage dynamique.
    for loan in loans:
        loan.user_inventory_status = _loan_user_inventory_status(loan)
    # Par défaut, l'affichage dynamique est en mode interne et affiche l'emprunteur.
    # Utiliser ?mode=public pour masquer les noms sur un écran visible par tous.
    public = request.GET.get('mode', 'interne') == 'public'
    return render(request, 'inventory/display_current_loans.html', {'loans': loans, 'now': now, 'public': public})


def api_current_loans(request):
    loans = Loan.objects.filter(status=Loan.LoanStatus.OPEN).select_related('equipment', 'borrower')
    data = []
    for loan in loans:
        data.append({
            'code': loan.equipment.code,
            'name': loan.equipment.name,
            'status': 'en_retard' if loan.is_late else 'sorti',
            'borrower': f'{loan.borrower.first_name} {loan.borrower.last_name}',
            'due_at': loan.due_at.isoformat() if loan.due_at else None,
            'late': bool(loan.is_late),
            'user_inventory_out_done': _loan_user_inventory_status(loan)['out_done'],
            'user_inventory_return_done': _loan_user_inventory_status(loan)['return_done'],
            'user_inventory_not_applicable': _loan_user_inventory_status(loan)['not_applicable'],
        })
    return JsonResponse(data, safe=False)


def _yes(value):
    return str(value).strip().upper() in ['OUI', 'YES', '1', 'TRUE', 'VRAI', 'X']


def _safe(value):
    return '' if value is None else str(value).strip()


def _role_from_excel(value):
    raw = _safe(value).upper().replace(' ', '_')
    return ROLE_IMPORT.get(raw)


def _export_workbook_response(wb, filename):
    output = io.BytesIO()
    wb.save(output)
    response = HttpResponse(output.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def _style_header(ws):
    fill = PatternFill('solid', fgColor='D9EAF7')
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
    ws.freeze_panes = 'A2'


def users_template_excel(request):
    if not require_prof(request):
        return redirect('storekeeper_login')
    wb = Workbook()
    ws = wb.active
    ws.title = 'utilisateurs'
    ws.append(USER_COLUMNS)
    ws.append(['USR-0001', 'MARTIN', 'Lucas', 'lucas.martin', 'lucas.martin@lycee.fr', 'BAC_CIEL', '1CIEL', 'A', 'Première', 'OUI', 'NON', 'UTILISATEUR', 'UTILISATEUR;MAGASINIER', '', 'CIEL2026!'])
    _style_header(ws)
    formations = ','.join(Formation.objects.values_list('code', flat=True)) or 'CAP_ELEC,BAC_MELEC,BAC_CIEL,BTS_ET,BTS_FED'
    roles = 'UTILISATEUR,MAGASINIER,TECH_INVENTAIRE,RESPONSABLE,LECTURE_SEULE,ADMIN'
    yesno = 'OUI,NON'
    for col, formula in [('F', formations), ('J', yesno), ('K', yesno), ('L', roles)]:
        dv = DataValidation(type='list', formula1=f'"{formula}"', allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f'{col}2:{col}500')
    for column_cells in ws.columns:
        ws.column_dimensions[column_cells[0].column_letter].width = 22
    return _export_workbook_response(wb, 'modele_import_utilisateurs_toolmag.xlsx')


def users_export_excel(request):
    if not require_prof(request):
        return redirect('storekeeper_login')
    qs = Person.objects.select_related('formation').all()
    formation = request.GET.get('formation')
    class_name = request.GET.get('classe')
    if formation:
        qs = qs.filter(formation__code=formation)
    if class_name:
        qs = qs.filter(class_name=class_name)
    wb = Workbook()
    ws = wb.active
    ws.title = 'utilisateurs'
    ws.append(USER_COLUMNS)
    for person in qs:
        ws.append([
            person.code,
            person.last_name,
            person.first_name,
            person.username or '',
            person.email,
            person.formation.code if person.formation else '',
            person.class_name,
            person.group_name,
            person.level,
            'OUI' if person.active else 'NON',
            'OUI' if person.archived else 'NON',
            ROLE_EXPORT.get(person.role, person.role),
            person.allowed_roles,
            person.rfid_uid,
            '',
        ])
    _style_header(ws)
    for column_cells in ws.columns:
        ws.column_dimensions[column_cells[0].column_letter].width = 22
    return _export_workbook_response(wb, 'toolmag_utilisateurs_export.xlsx')


def _parse_users_xlsx(file_obj):
    wb = load_workbook(file_obj, data_only=True)
    ws = wb.active
    headers = [_safe(c.value) for c in ws[1]]
    index = {h: i for i, h in enumerate(headers)}
    missing = [h for h in ['code_utilisateur', 'nom', 'prenom'] if h not in index]
    if missing:
        return [], [{'line': 1, 'level': 'ERREUR', 'message': f'Colonnes manquantes : {", ".join(missing)}'}], []
    rows = []
    report = []
    seen = set()
    classes_to_create = {}
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        values = {h: row[index[h]] if index[h] < len(row) else None for h in headers}
        code = _safe(values.get('code_utilisateur'))
        if not code:
            continue
        if code in seen:
            report.append({'line': row_idx, 'level': 'ERREUR', 'message': f'Doublon dans le fichier : {code}'})
        seen.add(code)
        role = _role_from_excel(values.get('role_principal')) or Person.Role.USER
        formation = None
        formation_code = _safe(values.get('formation')).upper()
        if formation_code:
            try:
                formation = Formation.objects.get(code=formation_code)
            except Formation.DoesNotExist:
                report.append({'line': row_idx, 'level': 'ERREUR', 'message': f'Formation inconnue : {formation_code}'})
        class_name = _safe(values.get('classe'))
        if class_name and formation:
            exists = SchoolClass.objects.filter(formation=formation, name__iexact=class_name).exists()
            if not exists:
                key = (formation.id, class_name.strip().lower())
                classes_to_create.setdefault(key, {
                    'formation': formation,
                    'formation_code': formation.code,
                    'class_name': class_name,
                    'lines': [],
                })
                classes_to_create[key]['lines'].append(row_idx)
        elif class_name and not formation_code:
            # Classe sans formation : on signale seulement, car aucune formation ne permet de classer proprement.
            key = (None, class_name.strip().lower())
            classes_to_create.setdefault(key, {
                'formation': None,
                'formation_code': '',
                'class_name': class_name,
                'lines': [],
            })
            classes_to_create[key]['lines'].append(row_idx)
        allowed_roles_raw = _safe(values.get('roles_autorises'))
        normalized_roles = []
        if allowed_roles_raw:
            for raw in allowed_roles_raw.split(';'):
                imported = _role_from_excel(raw)
                if not imported:
                    report.append({'line': row_idx, 'level': 'ERREUR', 'message': f'Rôle inconnu : {raw}'})
                else:
                    normalized_roles.append(ROLE_EXPORT.get(imported, imported))
        if not normalized_roles:
            normalized_roles = [ROLE_EXPORT.get(role, role)]
        if not _safe(values.get('nom')) or not _safe(values.get('prenom')):
            report.append({'line': row_idx, 'level': 'ERREUR', 'message': 'Nom ou prénom manquant.'})
        rows.append({
            'line': row_idx,
            'code': code,
            'last_name': _safe(values.get('nom')).upper(),
            'first_name': _safe(values.get('prenom')),
            'username': _safe(values.get('identifiant')) or None,
            'email': _safe(values.get('email')),
            'formation': formation,
            'formation_code': formation_code,
            'class_name': class_name,
            'group_name': _safe(values.get('groupe')),
            'level': _safe(values.get('niveau')),
            'active': _yes(values.get('actif')),
            'archived': _yes(values.get('archive')),
            'role': role,
            'allowed_roles': ';'.join(normalized_roles),
            'rfid_uid': _safe(values.get('rfid_uid')),
            'initial_password': _safe(values.get('mot_de_passe_initial')),
        })
    class_creations = list(classes_to_create.values())
    for item in class_creations:
        label = f"{item['formation_code']} / {item['class_name']}" if item['formation_code'] else item['class_name']
        report.append({
            'line': ','.join(str(x) for x in item['lines'][:5]) + ('…' if len(item['lines']) > 5 else ''),
            'level': 'INFO',
            'message': f'Classe inconnue détectée : {label}. Elle sera créée automatiquement si tu appliques l’import.'
        })
    return rows, report, class_creations

def users_import_excel(request):
    if not require_prof(request):
        return redirect('storekeeper_login')
    form = UserImportForm(request.POST or None, request.FILES or None)
    report = []
    summary = None
    if request.method == 'POST' and form.is_valid():
        rows, report, class_creations = _parse_users_xlsx(request.FILES['file'])
        has_errors = any(item['level'] == 'ERREUR' for item in report)
        to_create = []
        to_update = []
        if not has_errors:
            for row in rows:
                existing = Person.objects.filter(code=row['code']).first()
                (to_update if existing else to_create).append(row)
            summary = {'create': len(to_create), 'update': len(to_update), 'total': len(rows), 'class_create': len(class_creations), 'classes_to_create': class_creations, 'applied': False}
            if form.cleaned_data.get('apply_changes'):
                for class_item in class_creations:
                    SchoolClass.objects.get_or_create(
                        formation=class_item['formation'],
                        name=class_item['class_name'],
                        defaults={'active': True},
                    )
                for row in rows:
                    person, created = Person.objects.get_or_create(code=row['code'])
                    old_formation = person.formation
                    old_class = person.class_name
                    old_group = person.group_name
                    person.last_name = row['last_name']
                    person.first_name = row['first_name']
                    person.username = row['username']
                    person.email = row['email']
                    person.formation = row['formation']
                    person.class_name = row['class_name']
                    person.group_name = row['group_name']
                    person.level = row['level']
                    person.active = row['active']
                    person.archived = row['archived']
                    person.role = row['role']
                    person.allowed_roles = row['allowed_roles']
                    person.rfid_uid = row['rfid_uid']
                    if row.get('initial_password'):
                        person.set_password(row['initial_password'])
                        person.must_change_password = True
                    person.save()
                    EnrollmentHistory.objects.create(
                        person=person,
                        event_type=EnrollmentHistory.EventType.CREATED if created else EnrollmentHistory.EventType.IMPORT_UPDATED,
                        old_formation=old_formation,
                        new_formation=person.formation,
                        old_class_name=old_class,
                        new_class_name=person.class_name,
                        old_group_name=old_group,
                        new_group_name=person.group_name,
                        comment='Import Excel utilisateurs',
                    )
                summary['applied'] = True
                messages.success(request, f'Import appliqué : {len(to_create)} créations, {len(to_update)} mises à jour, {len(class_creations)} classe(s) créée(s) si nécessaire.')
        elif form.cleaned_data.get('apply_changes'):
            messages.error(request, 'Import refusé : corrige les erreurs bloquantes avant application.')
    return render(request, 'inventory/users_import.html', {'form': form, 'report': report, 'summary': summary})


def users_management(request):
    # La liste des utilisateurs est visible par tous les profils connectés ; seuls les boutons de gestion sont filtrés.
    if not (current_storekeeper(request) or current_borrower(request)):
        messages.warning(request, 'Connexion utilisateur ou magasinier nécessaire pour consulter les utilisateurs.')
        return redirect('user_login')
    persons = Person.objects.select_related('formation').all()
    q = (request.GET.get('q') or '').strip()
    formation = request.GET.get('formation') or ''
    class_name = request.GET.get('classe') or ''
    if q:
        persons = persons.filter(
            Q(code__icontains=q)
            | Q(first_name__icontains=q)
            | Q(last_name__icontains=q)
            | Q(username__icontains=q)
            | Q(class_name__icontains=q)
            | Q(group_name__icontains=q)
        )
    if formation:
        persons = persons.filter(formation__code=formation)
    if class_name:
        persons = persons.filter(class_name__icontains=class_name)
    formations = Formation.objects.filter(active=True)
    can_manage_users = is_super_admin_person(current_storekeeper(request))
    return render(request, 'inventory/users_management.html', {
        'persons': persons[:300],
        'formations': formations,
        'q': q,
        'selected_formation': formation,
        'selected_class': class_name,
        'can_manage_users': can_manage_users,
    })



def user_create(request):
    if not require_prof(request):
        return redirect('users_management')
    form = PersonCreateForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        person = form.save()
        EnrollmentHistory.objects.create(person=person, event_type=EnrollmentHistory.EventType.CREATED, new_formation=person.formation, new_class_name=person.class_name, new_group_name=person.group_name, comment='Création depuis ToolMag')
        messages.success(request, f'Utilisateur créé : {person.code} — {person.first_name} {person.last_name}.')
        return redirect('user_detail', code=person.code)
    return render(request, 'inventory/user_create.html', {'form': form})


def material_rights_management(request):
    if not require_prof(request):
        return redirect('dashboard')
    edit_id = request.GET.get('edit')
    grant = get_object_or_404(MaterialEditGrant, id=edit_id) if edit_id else None
    form = MaterialEditGrantForm(request.POST or None, instance=grant)
    if request.method == 'POST' and form.is_valid():
        obj = form.save(commit=False)
        if not obj.granted_by:
            obj.granted_by = current_storekeeper(request)
        obj.save()
        messages.success(request, 'Droit ponctuel matériel enregistré.')
        return redirect('material_rights_management')
    grants = MaterialEditGrant.objects.select_related('formation', 'granted_by')[:100]
    return render(request, 'inventory/material_rights.html', {'form': form, 'grants': grants, 'editing': grant})


def user_detail(request, code):
    viewer_storekeeper = current_storekeeper(request)
    viewer_borrower = current_borrower(request)
    if not (viewer_storekeeper or viewer_borrower):
        messages.warning(request, 'Connexion nécessaire pour consulter une fiche utilisateur.')
        return redirect(f"{reverse('user_login')}?next={request.path}")
    person = get_object_or_404(Person.objects.select_related('formation'), code=code)
    is_prof = is_super_admin_person(viewer_storekeeper)
    is_self = (viewer_storekeeper and viewer_storekeeper.id == person.id) or (viewer_borrower and viewer_borrower.id == person.id)
    date_from, date_to, dt_from, dt_to = _date_range_from_request(request)

    borrowed_loans = _filter_dt(person.borrowed_loans.select_related('equipment', 'checkout_storekeeper', 'return_storekeeper'), 'checked_out_at', dt_from, dt_to)[:100]
    user_inventories = _filter_dt(person.user_inventories.select_related('equipment', 'loan'), 'submitted_at', dt_from, dt_to)[:100]

    checkout_loans = _filter_dt(person.checkout_loans.select_related('equipment', 'borrower'), 'checked_out_at', dt_from, dt_to)[:100]
    return_loans = _filter_dt(person.return_loans.select_related('equipment', 'borrower'), 'returned_at', dt_from, dt_to)[:100]
    component_checks = _filter_dt(person.component_checks_done.select_related('loan__equipment', 'loan__borrower', 'component'), 'created_at', dt_from, dt_to)[:100]
    interventions = _filter_dt(person.interventions_done.select_related('equipment'), 'intervention_at', dt_from, dt_to)[:100]
    repairs = _filter_dt(person.repairs_done.select_related('equipment'), 'repaired_at', dt_from, dt_to)[:100]
    locker_logs = _filter_dt(person.locker_open_logs.select_related('equipment', 'terminal'), 'created_at', dt_from, dt_to)[:100]

    user_counts = {
        'borrowed_loans': borrowed_loans.count() if hasattr(borrowed_loans, 'count') else len(borrowed_loans),
        'user_inventories': user_inventories.count() if hasattr(user_inventories, 'count') else len(user_inventories),
    }
    storekeeper_counts = {
        'checkout_loans': checkout_loans.count() if hasattr(checkout_loans, 'count') else len(checkout_loans),
        'return_loans': return_loans.count() if hasattr(return_loans, 'count') else len(return_loans),
        'component_checks': component_checks.count() if hasattr(component_checks, 'count') else len(component_checks),
        'interventions': interventions.count() if hasattr(interventions, 'count') else len(interventions),
        'repairs': repairs.count() if hasattr(repairs, 'count') else len(repairs),
        'locker_logs': locker_logs.count() if hasattr(locker_logs, 'count') else len(locker_logs),
    }
    enrollment_history = person.enrollment_history.select_related('old_formation', 'new_formation')[:50] if is_prof else []
    return render(request, 'inventory/user_detail.html', {
        'person': person,
        'is_prof': is_prof,
        'is_self': is_self,
        'date_from': date_from,
        'date_to': date_to,
        'borrowed_loans': borrowed_loans,
        'user_inventories': user_inventories,
        'checkout_loans': checkout_loans,
        'return_loans': return_loans,
        'component_checks': component_checks,
        'interventions': interventions,
        'repairs': repairs,
        'locker_logs': locker_logs,
        'user_counts': user_counts,
        'storekeeper_counts': storekeeper_counts,
        'enrollment_history': enrollment_history,
    })


def _format_date_fr(value):
    if not value:
        return '-'
    try:
        if timezone.is_aware(value):
            value = timezone.localtime(value)
        return value.strftime('%d/%m/%Y %H:%M')
    except Exception:
        return str(value)


def _plain(value):
    """Text safe for PDF cells."""
    if value is None:
        return '-'
    value = str(value).strip()
    return value if value else '-'


def _period_label(date_from, date_to):
    if date_from and date_to:
        return f"du {date_from.strftime('%d/%m/%Y')} au {date_to.strftime('%d/%m/%Y')}"
    if date_from:
        return f"depuis le {date_from.strftime('%d/%m/%Y')}"
    if date_to:
        return f"jusqu'au {date_to.strftime('%d/%m/%Y')}"
    return "toute la période disponible"



def user_detail_pdf(request, code):
    """Export PDF de la fiche utilisateur filtrée.

    Le PDF reprend la période actuellement filtrée et distingue :
    - activité réalisée en tant qu'utilisateur ;
    - activité réalisée en tant que magasinier/responsable.
    """
    viewer_storekeeper = current_storekeeper(request)
    viewer_borrower = current_borrower(request)
    if not (viewer_storekeeper or viewer_borrower):
        messages.warning(request, 'Connexion nécessaire pour exporter une fiche utilisateur.')
        return redirect(f"{reverse('user_login')}?next={request.path}")

    person = get_object_or_404(Person.objects.select_related('formation'), code=code)
    is_prof = is_super_admin_person(viewer_storekeeper)
    is_self = (viewer_storekeeper and viewer_storekeeper.id == person.id) or (viewer_borrower and viewer_borrower.id == person.id)
    if not (is_prof or is_self):
        return HttpResponseForbidden('Export réservé à la personne concernée ou à un responsable.')

    date_from, date_to, dt_from, dt_to = _date_range_from_request(request)

    borrowed_loans = list(_filter_dt(person.borrowed_loans.select_related('equipment', 'checkout_storekeeper', 'return_storekeeper'), 'checked_out_at', dt_from, dt_to)[:300])
    user_inventories = list(_filter_dt(person.user_inventories.select_related('equipment', 'loan'), 'submitted_at', dt_from, dt_to)[:300])
    checkout_loans = list(_filter_dt(person.checkout_loans.select_related('equipment', 'borrower'), 'checked_out_at', dt_from, dt_to)[:300])
    return_loans = list(_filter_dt(person.return_loans.select_related('equipment', 'borrower'), 'returned_at', dt_from, dt_to)[:300])
    component_checks = list(_filter_dt(person.component_checks_done.select_related('loan__equipment', 'loan__borrower', 'component'), 'created_at', dt_from, dt_to)[:300])
    interventions = list(_filter_dt(person.interventions_done.select_related('equipment'), 'intervention_at', dt_from, dt_to)[:300])
    repairs = list(_filter_dt(person.repairs_done.select_related('equipment'), 'repaired_at', dt_from, dt_to)[:300])
    locker_logs = list(_filter_dt(person.locker_open_logs.select_related('equipment', 'terminal'), 'created_at', dt_from, dt_to)[:300])

    from xml.sax.saxutils import escape
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak

    def P(value, style):
        return Paragraph(escape(_plain(value)).replace('\n', '<br/>'), style)

    def make_table(headers, rows, widths=None):
        data = [[P(h, styles['TableHeader']) for h in headers]]
        if rows:
            for row in rows:
                data.append([P(cell, styles['Cell']) for cell in row])
        else:
            data.append([P('Aucune donnée sur cette période.', styles['Cell'])] + ['' for _ in headers[1:]])
        table = Table(data, colWidths=widths, repeatRows=1, hAlign='LEFT')
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#163b5c')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 7),
            ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#cccccc')),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f7f9fb')]),
            ('LEFTPADDING', (0, 0), (-1, -1), 3),
            ('RIGHTPADDING', (0, 0), (-1, -1), 3),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ]))
        return table

    buffer = io.BytesIO()
    filename = f"fiche_{person.code}_{timezone.now().strftime('%Y%m%d_%H%M')}.pdf"
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=10 * mm,
        leftMargin=10 * mm,
        topMargin=10 * mm,
        bottomMargin=10 * mm,
        title=f"Fiche utilisateur {person.code}",
    )
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle('Small', parent=styles['Normal'], fontSize=8, leading=10))
    styles.add(ParagraphStyle('Section', parent=styles['Heading2'], fontSize=12, leading=14, spaceBefore=8, spaceAfter=5, textColor=colors.HexColor('#163b5c')))
    styles.add(ParagraphStyle('TableHeader', parent=styles['Normal'], fontSize=7, leading=8, textColor=colors.white, alignment=TA_CENTER))
    styles.add(ParagraphStyle('Cell', parent=styles['Normal'], fontSize=7, leading=8))

    story = []
    story.append(Paragraph(f"Fiche élève / utilisateur - {escape(person.first_name)} {escape(person.last_name)}", styles['Title']))
    story.append(Paragraph(f"Période filtrée : {_period_label(date_from, date_to)}", styles['Small']))
    story.append(Paragraph(f"Export généré le {timezone.localtime(timezone.now()).strftime('%d/%m/%Y à %H:%M')}", styles['Small']))
    story.append(Spacer(1, 4 * mm))

    identity_rows = [
        ['Code', person.code, 'Identifiant', person.username or '-'],
        ['Formation', f"{person.formation.code} - {person.formation.name}" if person.formation else '-', 'Classe / groupe', f"{person.class_name or '-'} / {person.group_name or '-'}"],
        ['Rôle principal', person.get_role_display(), 'Rôles autorisés', person.allowed_roles or '-'],
        ['Statut', 'Actif' if person.active else 'Inactif', 'Archive', 'Oui' if person.archived else 'Non'],
    ]
    story.append(make_table(['Champ', 'Valeur', 'Champ', 'Valeur'], identity_rows, [35*mm, 85*mm, 35*mm, 110*mm]))

    story.append(Paragraph('Synthèse sur la période', styles['Section']))
    summary_rows = [[
        len(borrowed_loans), len(user_inventories), len(checkout_loans), len(return_loans),
        len(component_checks), len(interventions), len(repairs), len(locker_logs)
    ]]
    story.append(make_table([
        'Emprunts utilisateur', 'Inventaires utilisateur', 'Sorties magasinier', 'Retours magasinier',
        'Contrôles composants', 'Interventions', 'Réparations', 'Ouvertures casiers'
    ], summary_rows))

    story.append(Paragraph('Activité en tant qu’utilisateur', styles['Section']))
    story.append(Paragraph('Matériels empruntés', styles['Small']))
    story.append(make_table(
        ['Date sortie', 'Matériel', 'Magasinier', 'Retour', 'Statut', 'Commentaires'],
        [[
            _format_date_fr(loan.checked_out_at),
            f"{loan.equipment.code} - {loan.equipment.name}",
            f"{loan.checkout_storekeeper.first_name} {loan.checkout_storekeeper.last_name}",
            _format_date_fr(loan.returned_at),
            loan.get_status_display(),
            f"Sortie : {loan.comment_out or '-'}\nRetour : {loan.comment_return or '-'}",
        ] for loan in borrowed_loans],
        [30*mm, 55*mm, 42*mm, 30*mm, 25*mm, 90*mm]
    ))

    story.append(Spacer(1, 3 * mm))
    story.append(Paragraph('Inventaires utilisateur', styles['Small']))
    story.append(make_table(
        ['Date', 'Type', 'Matériel', 'État global', 'Statut', 'Commentaire'],
        [[
            _format_date_fr(inv.submitted_at),
            inv.get_inventory_type_display(),
            f"{inv.equipment.code} - {inv.equipment.name}",
            inv.get_global_condition_display(),
            inv.get_status_display(),
            inv.comment or '-',
        ] for inv in user_inventories],
        [30*mm, 55*mm, 58*mm, 28*mm, 28*mm, 73*mm]
    ))

    story.append(PageBreak())
    story.append(Paragraph('Activité en tant que magasinier / responsable', styles['Section']))
    story.append(Paragraph('Sorties validées', styles['Small']))
    story.append(make_table(
        ['Date', 'Matériel', 'Emprunteur', 'État sortie', 'Commentaire'],
        [[
            _format_date_fr(loan.checked_out_at),
            f"{loan.equipment.code} - {loan.equipment.name}",
            f"{loan.borrower.first_name} {loan.borrower.last_name} ({loan.borrower.code})",
            loan.get_condition_out_display(),
            loan.comment_out or '-',
        ] for loan in checkout_loans],
        [30*mm, 60*mm, 58*mm, 30*mm, 92*mm]
    ))

    story.append(Spacer(1, 3 * mm))
    story.append(Paragraph('Retours validés', styles['Small']))
    story.append(make_table(
        ['Date', 'Matériel', 'Emprunteur', 'État retour', 'Commentaire'],
        [[
            _format_date_fr(loan.returned_at),
            f"{loan.equipment.code} - {loan.equipment.name}",
            f"{loan.borrower.first_name} {loan.borrower.last_name} ({loan.borrower.code})",
            loan.get_condition_return_display(),
            loan.comment_return or '-',
        ] for loan in return_loans],
        [30*mm, 60*mm, 58*mm, 30*mm, 92*mm]
    ))

    story.append(Spacer(1, 3 * mm))
    story.append(Paragraph('Contrôles de composants', styles['Small']))
    story.append(make_table(
        ['Date', 'Type', 'Matériel', 'Emprunteur', 'Composant', 'Présent', 'État', 'Commentaire'],
        [[
            _format_date_fr(check.created_at),
            check.get_check_type_display(),
            check.loan.equipment.code if check.loan and check.loan.equipment else '-',
            f"{check.loan.borrower.first_name} {check.loan.borrower.last_name}" if check.loan and check.loan.borrower else '-',
            check.component.name,
            'Oui' if check.present else 'Non',
            check.get_condition_display(),
            check.comment or '-',
        ] for check in component_checks],
        [25*mm, 30*mm, 25*mm, 38*mm, 45*mm, 18*mm, 25*mm, 62*mm]
    ))

    story.append(PageBreak())
    story.append(Paragraph('Maintenance, interventions et traçabilité', styles['Section']))
    story.append(Paragraph('Bons d’intervention', styles['Small']))
    story.append(make_table(
        ['Date', 'Matériel', 'Type', 'Résultat', 'Constat', 'Action / commentaire'],
        [[
            _format_date_fr(intervention.intervention_at),
            f"{intervention.equipment.code} - {intervention.equipment.name}",
            intervention.get_intervention_type_display(),
            intervention.get_result_display(),
            intervention.finding or '-',
            intervention.comment or intervention.action_done or '-',
        ] for intervention in interventions],
        [30*mm, 55*mm, 35*mm, 35*mm, 55*mm, 62*mm]
    ))

    story.append(Spacer(1, 3 * mm))
    story.append(Paragraph('Bons de réparation / maintenance', styles['Small']))
    story.append(make_table(
        ['Date', 'Matériel', 'Type', 'Résultat', 'Diagnostic', 'Action / pièces / commentaire'],
        [[
            _format_date_fr(repair.repaired_at),
            f"{repair.equipment.code} - {repair.equipment.name}",
            repair.get_repair_type_display(),
            repair.get_result_display(),
            repair.diagnosis or '-',
            (repair.comment or repair.action_done or '-') + (f"\nPièces : {repair.parts_replaced}" if repair.parts_replaced else ''),
        ] for repair in repairs],
        [30*mm, 55*mm, 35*mm, 35*mm, 55*mm, 62*mm]
    ))

    story.append(Spacer(1, 3 * mm))
    story.append(Paragraph('Ouvertures de casiers', styles['Small']))
    story.append(make_table(
        ['Date', 'Matériel', 'Terminal', 'Armoire', 'Casier', 'Contexte', 'Résultat'],
        [[
            _format_date_fr(log.created_at),
            log.equipment.code if log.equipment else '-',
            log.terminal.name if log.terminal else '-',
            log.cabinet,
            log.locker,
            log.get_context_display(),
            'Succès' if log.success else f"Refusé - {log.refusal_reason}" if log.refused else 'Erreur',
        ] for log in locker_logs],
        [30*mm, 30*mm, 40*mm, 30*mm, 25*mm, 45*mm, 70*mm]
    ))

    def add_page_number(canvas, doc):
        canvas.saveState()
        canvas.setFont('Helvetica', 7)
        canvas.drawRightString(285 * mm, 7 * mm, f"Page {doc.page}")
        canvas.restoreState()

    doc.build(story, onFirstPage=add_page_number, onLaterPages=add_page_number)
    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def _person_has_history(person):
    return (
        person.borrowed_loans.exists() or
        person.checkout_loans.exists() or
        person.return_loans.exists() or
        person.enrollment_history.exists()
    )


def promotion_management(request):
    if not require_prof(request):
        return redirect('storekeeper_login')
    filter_form = PromotionFilterForm(request.GET or None)
    persons = Person.objects.select_related('formation').filter(archived=False)
    if filter_form.is_valid():
        if filter_form.cleaned_data.get('formation'):
            persons = persons.filter(formation=filter_form.cleaned_data['formation'])
        if filter_form.cleaned_data.get('class_name'):
            persons = persons.filter(class_name__iexact=filter_form.cleaned_data['class_name'])
        if filter_form.cleaned_data.get('group_name'):
            persons = persons.filter(group_name__iexact=filter_form.cleaned_data['group_name'])
    action_choices = [(p.code, f'{p.last_name} {p.first_name} — {p.class_name} {p.group_name}'.strip()) for p in persons]
    action_form = PromotionActionForm(request.POST or None)
    action_form.fields['selected'].choices = action_choices
    report = []
    summary = None
    if request.method == 'POST' and action_form.is_valid():
        selected_codes = action_form.cleaned_data['selected']
        selected_people = list(Person.objects.select_related('formation').filter(code__in=selected_codes))
        action = action_form.cleaned_data['action']
        new_formation = action_form.cleaned_data.get('new_formation')
        new_class = action_form.cleaned_data.get('new_class_name')
        new_group = action_form.cleaned_data.get('new_group_name')
        school_year = action_form.cleaned_data.get('school_year')
        comment = action_form.cleaned_data.get('comment')
        apply_changes = action_form.cleaned_data.get('apply_changes')
        for person in selected_people:
            old_formation = person.formation
            old_class = person.class_name
            old_group = person.group_name
            event_type = EnrollmentHistory.EventType.GROUP_CHANGED
            message = ''
            if action == 'promote':
                event_type = EnrollmentHistory.EventType.PROMOTED
                person.formation = new_formation or person.formation
                person.class_name = new_class or person.class_name
                person.group_name = new_group or person.group_name
                message = f'{old_class} → {person.class_name}'
            elif action == 'repeat':
                event_type = EnrollmentHistory.EventType.REPEATED
                person.formation = new_formation or person.formation
                person.class_name = new_class or person.class_name
                person.group_name = new_group or person.group_name
                message = f'Maintien / redoublement : {person.class_name}'
            elif action == 'transfer':
                event_type = EnrollmentHistory.EventType.TRANSFERRED
                person.formation = new_formation or person.formation
                person.class_name = new_class or person.class_name
                person.group_name = new_group or person.group_name
                message = f'{old_formation} → {person.formation}'
            elif action == 'change_group':
                event_type = EnrollmentHistory.EventType.GROUP_CHANGED
                person.formation = new_formation or person.formation
                person.class_name = new_class or person.class_name
                person.group_name = new_group or person.group_name
                message = f'{old_class} {old_group} → {person.class_name} {person.group_name}'
            elif action == 'deactivate':
                event_type = EnrollmentHistory.EventType.DEACTIVATED
                person.active = False
                message = 'Compte désactivé'
            elif action == 'archive':
                event_type = EnrollmentHistory.EventType.ARCHIVED
                person.active = False
                person.archived = True
                message = 'Compte archivé'
            elif action == 'delete_if_no_history':
                event_type = EnrollmentHistory.EventType.DELETED_REQUESTED
                if _person_has_history(person):
                    person.active = False
                    person.archived = True
                    message = 'Historique détecté : archivage au lieu de suppression'
                else:
                    message = 'Suppression définitive possible'
                    if apply_changes:
                        code = person.code
                        person.delete()
                        report.append({'code': code, 'name': '', 'action': 'supprimé', 'message': message})
                        continue
            if apply_changes:
                person.save()
                EnrollmentHistory.objects.create(
                    person=person,
                    school_year=school_year,
                    event_type=event_type,
                    old_formation=old_formation,
                    new_formation=person.formation,
                    old_class_name=old_class,
                    new_class_name=person.class_name,
                    old_group_name=old_group,
                    new_group_name=person.group_name,
                    comment=comment or message,
                )
            report.append({'code': person.code, 'name': f'{person.last_name} {person.first_name}', 'action': action, 'message': message})
        summary = {'count': len(selected_people), 'applied': apply_changes, 'action': action}
        if apply_changes:
            messages.success(request, f'Opération appliquée pour {len(selected_people)} utilisateur(s).')
    return render(request, 'inventory/promotion.html', {
        'filter_form': filter_form,
        'action_form': action_form,
        'persons': persons,
        'report': report,
        'summary': summary,
    })



def _level_from_evidence(evidence_count, problem_count=0):
    score = max(0, int(evidence_count) - int(problem_count))
    if score <= 0:
        return 0
    if score <= 2:
        return 1
    if score <= 5:
        return 2
    if score <= 10:
        return 3
    return 4


def _filter_period_qs(qs, date_from=None, date_to=None, field='created_at'):
    if date_from:
        qs = qs.filter(**{f'{field}__date__gte': date_from})
    if date_to:
        qs = qs.filter(**{f'{field}__date__lte': date_to})
    return qs


def _evaluation_evidence_for_person(person, date_from=None, date_to=None):
    borrowed = _filter_period_qs(person.borrowed_loans.all(), date_from, date_to, field='checked_out_at')
    checkout_validated = _filter_period_qs(person.checkout_loans.all(), date_from, date_to, field='checked_out_at')
    returns_validated = _filter_period_qs(person.return_loans.all(), date_from, date_to, field='returned_at')
    checks = _filter_period_qs(person.component_checks_done.all(), date_from, date_to, field='created_at')
    user_checks = checks.filter(checked_by_role='utilisateur')
    storekeeper_checks = checks.filter(checked_by_role='magasinier')
    submitted_user_inventories = _filter_period_qs(person.user_inventories.filter(status__in=[UserInventory.InventoryStatus.SUBMITTED, UserInventory.InventoryStatus.APPLIED]), date_from, date_to, field='submitted_at')
    problem_loans = Loan.objects.filter(Q(borrower=person) | Q(checkout_storekeeper=person) | Q(return_storekeeper=person), status=Loan.LoanStatus.PROBLEM).distinct()
    problem_loans = _filter_period_qs(problem_loans, date_from, date_to, field='updated_at')
    return {
        'borrowed_count': borrowed.count(),
        'checkout_validated_count': checkout_validated.count(),
        'return_validated_count': returns_validated.count(),
        'user_check_count': user_checks.count() + submitted_user_inventories.count(),
        'storekeeper_check_count': storekeeper_checks.count(),
        'problem_count': problem_loans.count(),
    }


def _suggestions_for_person(person, evidence):
    if not person.formation:
        return []
    action_evidence = {
        (CompetenceMapping.ActionType.CHECKOUT, Person.Role.USER): evidence['borrowed_count'],
        (CompetenceMapping.ActionType.INVENTORY_OUT, Person.Role.USER): evidence['user_check_count'],
        (CompetenceMapping.ActionType.INVENTORY_RETURN, Person.Role.USER): evidence['user_check_count'],
        (CompetenceMapping.ActionType.CHECKOUT, Person.Role.STOREKEEPER): evidence['checkout_validated_count'],
        (CompetenceMapping.ActionType.RETURN, Person.Role.STOREKEEPER): evidence['return_validated_count'],
        (CompetenceMapping.ActionType.INVENTORY_OUT, Person.Role.STOREKEEPER): evidence['storekeeper_check_count'],
        (CompetenceMapping.ActionType.INVENTORY_RETURN, Person.Role.STOREKEEPER): evidence['storekeeper_check_count'],
    }
    aggregated = {}
    mappings = CompetenceMapping.objects.filter(formation=person.formation, active=True, competence__active=True).select_related('competence')
    for mapping in mappings:
        roles = [mapping.role] if mapping.role else [Person.Role.USER, Person.Role.STOREKEEPER, Person.Role.TECH_INVENTORY]
        for role in roles:
            count = action_evidence.get((mapping.action_type, role), 0)
            if count <= 0:
                continue
            key = (mapping.competence_id, role)
            item = aggregated.setdefault(key, {'competence': mapping.competence, 'role': role, 'evidence_count': 0, 'criteria': []})
            item['evidence_count'] += count * mapping.weight
            if mapping.criterion:
                item['criteria'].append(mapping.criterion)
    suggestions = []
    for item in aggregated.values():
        item['proposed_level'] = _level_from_evidence(item['evidence_count'], evidence.get('problem_count', 0))
        suggestions.append(item)
    suggestions.sort(key=lambda x: (x['role'], x['competence'].code))
    return suggestions


def evaluation_dashboard(request):
    if not require_prof(request):
        return redirect('storekeeper_login')
    form = EvaluationFilterForm(request.GET or None)
    persons = Person.objects.select_related('formation').filter(active=True, archived=False)
    date_from = date_to = None
    selected_session = None
    if form.is_valid():
        date_from = form.cleaned_data.get('date_from')
        date_to = form.cleaned_data.get('date_to')
        selected_session = form.cleaned_data.get('session')
        if form.cleaned_data.get('formation'):
            persons = persons.filter(formation=form.cleaned_data['formation'])
        if form.cleaned_data.get('class_name'):
            persons = persons.filter(class_name__iexact=form.cleaned_data['class_name'])
        if form.cleaned_data.get('group_name'):
            persons = persons.filter(group_name__iexact=form.cleaned_data['group_name'])
        if form.cleaned_data.get('person'):
            persons = persons.filter(pk=form.cleaned_data['person'].pk)
        if selected_session:
            persons = persons.filter(session_assignments__session=selected_session).distinct()
    rows = []
    if request.method == 'POST' and request.POST.get('generate_evaluations') == '1':
        # POST conserve les filtres via querystring. Génère les propositions persistantes.
        storekeeper = current_storekeeper(request)
        for person in persons:
            evidence = _evaluation_evidence_for_person(person, date_from, date_to)
            for suggestion in _suggestions_for_person(person, evidence):
                record, _ = EvaluationRecord.objects.update_or_create(
                    session=selected_session,
                    person=person,
                    competence=suggestion['competence'],
                    role=suggestion['role'],
                    defaults={
                        'proposed_level': suggestion['proposed_level'],
                        'evidence_count': suggestion['evidence_count'],
                        'comment': '; '.join(suggestion['criteria'][:4]),
                        'validated_by': storekeeper if is_super_admin_person(storekeeper) else None,
                    },
                )
        messages.success(request, 'Propositions d’évaluation générées / mises à jour.')
        return redirect(request.get_full_path().replace(request.path, reverse('evaluation_dashboard')))
    for person in persons[:200]:
        evidence = _evaluation_evidence_for_person(person, date_from, date_to)
        suggestions = _suggestions_for_person(person, evidence)
        rows.append({'person': person, 'evidence': evidence, 'suggestions': suggestions})
    records = EvaluationRecord.objects.select_related('person', 'competence', 'session').all()[:100]
    return render(request, 'inventory/evaluation_dashboard.html', {
        'form': form,
        'rows': rows,
        'records': records,
        'selected_session': selected_session,
    })


def evaluation_sessions(request):
    if not require_prof(request):
        return redirect('storekeeper_login')
    form = PedagogicalSessionForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        session = form.save()
        messages.success(request, f'Séance créée : {session.title}. Les affectations élèves peuvent être complétées dans /admin ou ajoutées ensuite.')
        return redirect('evaluation_sessions')
    sessions = PedagogicalSession.objects.select_related('formation').prefetch_related('targeted_competences')[:50]
    return render(request, 'inventory/evaluation_sessions.html', {'form': form, 'sessions': sessions})


def evaluation_export_excel(request):
    if not require_prof(request):
        return redirect('storekeeper_login')
    records = EvaluationRecord.objects.select_related('person', 'competence', 'competence__formation', 'session').all()
    formation = request.GET.get('formation')
    class_name = request.GET.get('classe')
    if formation:
        records = records.filter(competence__formation__code=formation)
    if class_name:
        records = records.filter(person__class_name=class_name)
    wb = Workbook()
    ws = wb.active
    ws.title = 'evaluations'
    ws.append(['eleve_code', 'nom', 'prenom', 'formation', 'classe', 'groupe', 'seance', 'role', 'competence', 'libelle', 'niveau_propose', 'niveau_valide', 'preuves', 'commentaire'])
    for r in records:
        ws.append([
            r.person.code, r.person.last_name, r.person.first_name,
            r.competence.formation.code if r.competence.formation else '',
            r.person.class_name, r.person.group_name,
            r.session.title if r.session else '', r.get_role_display() if r.role else '',
            r.competence.code, r.competence.title, r.proposed_level,
            '' if r.validated_level is None else r.validated_level,
            r.evidence_count, r.comment,
        ])
    _style_header(ws)
    for column_cells in ws.columns:
        ws.column_dimensions[column_cells[0].column_letter].width = 24
    return _export_workbook_response(wb, 'toolmag_evaluations.xlsx')


def terminals_management(request):
    storekeeper = current_storekeeper(request)
    if not is_super_admin_person(storekeeper):
        messages.error(request, 'Gestion des terminaux réservée à un responsable ou administrateur connecté comme magasinier.')
        return redirect('storekeeper_login')
    terminals = AuthorizedTerminal.objects.all().order_by('name')
    settings_obj = LockerSettings.get_solo()
    return render(request, 'inventory/terminals_management.html', {
        'terminals': terminals,
        'settings': settings_obj,
        'current_terminal': current_terminal(request),
        'current_storekeeper': storekeeper,
    })


def terminal_register(request):
    storekeeper = current_storekeeper(request)
    if not is_super_admin_person(storekeeper):
        messages.error(request, 'Enregistrement de terminal réservé à un responsable ou administrateur connecté comme magasinier.')
        return redirect('storekeeper_login')
    form = TerminalRegistrationForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        token = uuid.uuid4().hex + uuid.uuid4().hex
        terminal = AuthorizedTerminal.objects.create(
            name=form.cleaned_data['name'],
            terminal_type=form.cleaned_data['terminal_type'],
            token=token,
            can_open_lockers=form.cleaned_data['can_open_lockers'],
            active=True,
            created_by=storekeeper,
            last_ip=get_client_ip(request),
            user_agent=request.META.get('HTTP_USER_AGENT', '')[:2000],
            last_seen_at=timezone.now(),
        )
        request.session['terminal_token'] = token
        response = redirect('terminals_management')
        response.set_cookie('toolmag_terminal_token', token, max_age=60*60*24*365, httponly=True, samesite='Lax', secure=request.is_secure())
        messages.success(request, f'Terminal enregistré : {terminal.name}.')
        return response
    return render(request, 'inventory/terminal_register.html', {'form': form, 'current_storekeeper': storekeeper})


def api_locker_open(request):
    if request.method != 'POST':
        return JsonResponse({'ok': False, 'error': 'Méthode interdite'}, status=405)
    equipment_code = request.POST.get('equipment_code') or ''
    context = request.POST.get('context') or LockerOpenLog.Context.DETAIL
    equipment = None
    if equipment_code:
        equipment = Equipment.objects.filter(code=equipment_code).first()
        if not equipment:
            return JsonResponse({'ok': False, 'error': 'Matériel introuvable'}, status=404)
    cabinet = request.POST.get('cabinet') or ''
    locker = request.POST.get('locker') or ''
    reason = request.POST.get('reason') or ''
    log = create_locker_log(request, equipment=equipment, cabinet=cabinet, locker=locker, context=context, force_reason=reason)
    if log.success:
        return JsonResponse({'ok': True, 'message': 'Commande casier envoyée.', 'log_id': log.id})
    status = 403 if log.refused else 500
    return JsonResponse({'ok': False, 'message': log.refusal_reason or log.controller_response or 'Échec ouverture casier.', 'log_id': log.id}, status=status)


def locker_force(request):
    storekeeper = current_storekeeper(request)
    if not is_super_admin_person(storekeeper):
        messages.error(request, 'Forçage casier réservé à un responsable ou administrateur connecté comme magasinier.')
        return redirect('storekeeper_login')
    form = ForceLockerForm(request.POST or None)
    last_log = None
    if request.method == 'POST' and form.is_valid():
        last_log = create_locker_log(
            request,
            cabinet=form.cleaned_data['cabinet'],
            locker=form.cleaned_data['locker'],
            context=form.cleaned_data['context'],
            force_reason=form.cleaned_data['reason'],
        )
        if last_log.success:
            messages.success(request, f'Commande de forçage envoyée pour {last_log.cabinet}/{last_log.locker}.')
        elif last_log.refused:
            messages.error(request, f'Forçage refusé : {last_log.refusal_reason}')
        else:
            messages.error(request, f'Échec forçage : {last_log.controller_response}')
    recent_logs = LockerOpenLog.objects.select_related('equipment', 'storekeeper', 'terminal')[:20]
    return render(request, 'inventory/locker_force.html', {
        'form': form,
        'last_log': last_log,
        'recent_logs': recent_logs,
        'current_storekeeper': storekeeper,
        'current_terminal': current_terminal(request),
        'locker_settings': LockerSettings.get_solo(),
    })



def sync_core_users(request):
    """Synchronisation manuelle ToolMag depuis LP Core.
    Le module ToolMag garde une copie locale des utilisateurs pour continuer
    à fonctionner même si LP Core est arrêté.
    """
    storekeeper = current_storekeeper(request)
    if not is_super_admin_person(storekeeper):
        messages.error(request, 'Synchronisation LP Core réservée à un responsable ou administrateur ToolMag.')
        return redirect('storekeeper_login')
    from .core_sync import sync_users_from_lp_core
    try:
        report = sync_users_from_lp_core(force_password=False)
        if report['errors']:
            messages.warning(request, f"Synchronisation partielle : {report['created']} créés, {report['updated']} mis à jour, {len(report['errors'])} erreurs.")
        else:
            messages.success(request, f"Synchronisation LP Core réussie : {report['created']} créés, {report['updated']} mis à jour.")
    except Exception as exc:
        messages.error(request, f'Échec de synchronisation LP Core : {exc}')
    return redirect('users_management')


@csrf_exempt
def api_internal_sync_core_users(request):
    """Endpoint interne appelé par LP Core pour pousser la base élèves vers ToolMag."""
    token = request.headers.get('X-API-Key') or request.POST.get('token') or request.GET.get('token')
    expected = getattr(settings, 'LP_CORE_API_TOKEN', '') or ''
    if expected and token != expected:
        return JsonResponse({'ok': False, 'error': 'API token invalide'}, status=403)
    if request.method != 'POST':
        return JsonResponse({'ok': False, 'error': 'Méthode POST requise'}, status=405)
    from .core_sync import sync_users_from_lp_core
    try:
        force_password = request.POST.get('force_password') in {'1', 'true', 'True', 'oui', 'OUI'}
        core_user_id = request.POST.get('core_user_id') or request.GET.get('core_user_id')
        report = sync_users_from_lp_core(timeout=90, force_password=force_password, core_user_id=core_user_id)
        return JsonResponse({'ok': len(report.get('errors', [])) == 0, 'report': report})
    except Exception as exc:
        return JsonResponse({'ok': False, 'error': str(exc)}, status=500)

def backups_management(request):
    """Gestion web super admin des sauvegardes/restaurations.
    - sauvegarde manuelle : conservée sans purge automatique ;
    - restauration : confirmation RESTAURER + sauvegarde pre_restore obligatoire.
    """
    storekeeper = current_storekeeper(request)
    if not is_super_admin_person(storekeeper):
        messages.error(request, 'Gestion des sauvegardes réservée à un responsable ou administrateur connecté comme magasinier.')
        return redirect('storekeeper_login')

    manual_form = ManualBackupForm()
    restore_form = RestoreBackupForm()

    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'manual_backup':
            manual_form = ManualBackupForm(request.POST)
            if manual_form.is_valid():
                note = manual_form.cleaned_data.get('note') or f'Sauvegarde manuelle par {storekeeper.code}'
                archive = create_backup('manual', note=note, retain_days=7)
                messages.success(request, f'Sauvegarde manuelle créée : {archive.name}')
                return redirect('backups_management')
        elif action == 'restore':
            restore_form = RestoreBackupForm(request.POST)
            if restore_form.is_valid():
                backup_name = restore_form.cleaned_data['backup_name']
                try:
                    pre, restored = restore_backup_from_archive(
                        backup_name,
                        create_pre_restore=True,
                        actor_label=f'{storekeeper.code} {storekeeper.first_name} {storekeeper.last_name}',
                    )
                    if pre:
                        messages.success(request, f'Sauvegarde de sécurité pré-restauration créée : {pre.name}')
                    messages.success(request, f'Restauration effectuée depuis {backup_name} : {restored}. Recharge la page et redémarre le conteneur si nécessaire.')
                    return redirect('backups_management')
                except Exception as exc:
                    messages.error(request, f'Échec restauration : {exc}')
        elif action == 'delete':
            backup_name = request.POST.get('backup_name') or ''
            try:
                target = safe_backup_path(backup_name)
                # Sécurité : suppression web autorisée seulement pour sauvegardes manuelles/pre_restore.
                if not (target.name.startswith('manual-toolmag-') or target.name.startswith('pre-restore-toolmag-')):
                    messages.error(request, 'La suppression web est réservée aux sauvegardes manuelles ou pré-restore. Les auto sont gérées par la rétention 7 jours.')
                else:
                    target.unlink(missing_ok=True)
                    messages.success(request, f'Sauvegarde supprimée : {target.name}')
                return redirect('backups_management')
            except Exception as exc:
                messages.error(request, f'Échec suppression : {exc}')
        elif action == 'download':
            backup_name = request.POST.get('backup_name') or ''
            try:
                target = safe_backup_path(backup_name)
                return FileResponse(open(target, 'rb'), as_attachment=True, filename=target.name)
            except Exception as exc:
                messages.error(request, f'Téléchargement impossible : {exc}')

    backups = list_backups()
    return render(request, 'inventory/backups_management.html', {
        'backups': backups,
        'manual_form': manual_form,
        'restore_form': restore_form,
        'current_storekeeper': storekeeper,
        'backup_dir': settings.BASE_DIR / 'backups',
    })
GLOSSARY_ROWS = [
    ('Catégorie', 'Nom de catégorie', 'Appareil de mesure', 'Measuring device'),
    ('Catégorie', 'Nom de catégorie', 'Outillage à main', 'Hand tools'),
    ('Catégorie', 'Nom de catégorie', 'Outillage électroportatif', 'Power tools'),
    ('Catégorie', 'Nom de catégorie', 'Réseau informatique', 'Computer networking'),
    ('Catégorie', 'Nom de catégorie', 'Fibre optique', 'Optical fiber'),
    ('Catégorie', 'Nom de catégorie', 'Caméra / vidéo', 'Camera / video equipment'),
    ('Catégorie', 'Nom de catégorie', 'Équipement de sécurité', 'Safety equipment'),
    ('Catégorie', 'Nom de catégorie', 'Équipement électrique', 'Electrical equipment'),
    ('Emplacement', 'Nom d’emplacement', 'Magasin électrique', 'Electrical store room'),
    ('Emplacement', 'Nom d’emplacement', 'Atelier CIEL', 'CIEL workshop'),
    ('Emplacement', 'Nom d’emplacement', 'Atelier MELEC', 'MELEC workshop'),
    ('Emplacement', 'Nom d’emplacement', 'Armoire mesure', 'Measuring cabinet'),
    ('Emplacement', 'Nom d’emplacement', 'Armoire fibre', 'Fiber cabinet'),
    ('Emplacement', 'Nom d’emplacement', 'Réserve matériel', 'Equipment storage room'),
    ('Emplacement', 'Nom d’emplacement', 'Salle professeurs', 'Teachers’ room'),
    ('Matériel', 'Nom du matériel', 'Multimètre Fluke', 'Fluke multimeter'),
    ('Matériel', 'Nom du matériel', 'Analyseur réseau', 'Network analyzer / Power quality analyzer'),
    ('Matériel', 'Nom du matériel', 'Contrôleur d’installation', 'Installation tester'),
    ('Matériel', 'Nom du matériel', 'Caméra thermique', 'Thermal camera'),
    ('Matériel', 'Nom du matériel', 'Soudeuse fibre optique', 'Optical fiber fusion splicer'),
    ('Matériel', 'Nom du matériel', 'Oscilloscope', 'Oscilloscope'),
    ('Matériel', 'Descriptif matériel', 'Contrôleur d’installation électrique', 'Electrical installation tester'),
    ('Matériel', 'Descriptif matériel', 'Oscilloscope triphasé', 'Three-phase oscilloscope'),
    ('Matériel', 'Descriptif matériel', 'Kit de soudure fibre', 'Fiber splicing kit'),
    ('Matériel', 'Descriptif matériel', 'Appareil de mesure courant faible', 'Low-current measuring device'),
    ('Composant', 'Nom du composant', 'Cordon rouge', 'Red test lead'),
    ('Composant', 'Nom du composant', 'Cordon noir', 'Black test lead'),
    ('Composant', 'Nom du composant', 'Alimentation secteur', 'Power supply'),
    ('Composant', 'Nom du composant', 'Pince ampèremétrique', 'Current clamp'),
    ('Composant', 'Nom du composant', 'Câble USB', 'USB cable'),
    ('Composant', 'Nom du composant', 'Mallette', 'Carrying case'),
    ('Composant', 'Nom du composant', 'Notice rapide', 'Quick start guide'),
    ('Document', 'Type de document', 'Notice constructeur', 'Manufacturer manual'),
    ('Document', 'Type de document', 'Fiche de prise en main', 'Getting started guide'),
    ('Document', 'Type de document', 'Consigne de sécurité', 'Safety instruction sheet'),
    ('Document', 'Type de document', 'Fiche maintenance', 'Maintenance sheet'),
    ('Document', 'Type de document', 'Procédure de test', 'Test procedure'),
    ('Intervention', 'Type d’intervention', 'Contrôle', 'Inspection'),
    ('Intervention', 'Type d’intervention', 'Nettoyage', 'Cleaning'),
    ('Intervention', 'Type d’intervention', 'Vérification périodique', 'Periodic check'),
    ('Intervention', 'Type d’intervention', 'Reconditionnement', 'Reconditioning'),
    ('Intervention', 'Type d’intervention', 'Contrôle accessoires', 'Accessory check'),
    ('Intervention', 'Type d’intervention', 'Test fonctionnement', 'Functional test'),
    ('Intervention', 'Type d’intervention', 'Maintenance légère', 'Light maintenance'),
    ('Réparation', 'Type d’intervention', 'Dépannage', 'Troubleshooting'),
    ('Réparation', 'Type d’intervention', 'Remplacement accessoire', 'Accessory replacement'),
    ('Réparation', 'Type d’intervention', 'Remise en état', 'Repair / restoration'),
    ('Réparation', 'Type d’intervention', 'Diagnostic', 'Diagnosis'),
    ('Statut matériel', 'Statut', 'Disponible', 'Available'),
    ('Statut matériel', 'Statut', 'Sorti', 'Checked out'),
    ('Statut matériel', 'Statut', 'En maintenance', 'Under maintenance'),
    ('Statut matériel', 'Statut', 'Incomplet', 'Incomplete'),
    ('Statut matériel', 'Statut', 'Hors service', 'Out of service'),
    ('Statut matériel', 'Statut', 'Perdu', 'Lost'),
    ('État matériel', 'État', 'Bon état', 'Good condition'),
    ('État matériel', 'État', 'Usure normale', 'Normal wear'),
    ('État matériel', 'État', 'À surveiller', 'To be monitored'),
    ('État matériel', 'État', 'Abîmé', 'Damaged'),
    ('État matériel', 'État', 'Dangereux', 'Unsafe'),
    ('État matériel', 'État', 'Absent', 'Missing'),
    ('Rôle', 'Rôle', 'Utilisateur', 'User'),
    ('Rôle', 'Rôle', 'Magasinier', 'Storekeeper'),
    ('Rôle', 'Rôle', 'Responsable', 'Supervisor'),
    ('Rôle', 'Rôle', 'Administrateur', 'Administrator'),
    ('Rôle', 'Rôle ponctuel', 'Technicien inventaire', 'Inventory technician'),
]

DNL_PROMPT = '''You are a teacher in a technical workshop. Use the following bilingual vocabulary to help students describe the equipment checkout and return process in English.

Task for students:
Write 8 to 10 sentences explaining how a user borrows equipment, checks the components, reports a missing or damaged item, returns the equipment, and how the storekeeper ensures traceability.

Required words:
- equipment
- user
- storekeeper
- checkout checklist
- return checklist
- missing component
- damaged equipment
- service report
- secure cabinet
- traceability
'''


def set_interface_language(request, lang):
    """Bascule légère FR/EN pour l’interface ToolMag."""
    lang = 'en' if str(lang).lower().startswith('en') else 'fr'
    request.session['toolmag_lang'] = lang
    request.session['_language'] = 'en' if lang == 'en' else 'fr'
    next_url = request.GET.get('next') or request.META.get('HTTP_REFERER') or reverse('dashboard')
    response = redirect(next_url)
    response.set_cookie('toolmag_language', 'en' if lang == 'en' else 'fr', max_age=365*24*3600, samesite='Lax')
    return response


def dnl_glossary(request):
    """Tableau des champs/données métier non traduits automatiquement, pour activités DNL."""
    if request.GET.get('format') == 'csv':
        import csv
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = 'attachment; filename="glossaire_dnl_toolmag_fr_en.csv"'
        response.write('\ufeff')
        writer = csv.writer(response, delimiter=';')
        writer.writerow(['Zone ToolMag', 'Champ', 'Exemple FR', 'Traduction EN proposée'])
        writer.writerows(GLOSSARY_ROWS)
        return response
    return render(request, 'inventory/dnl_glossary.html', {
        'glossary_rows': GLOSSARY_ROWS,
        'dnl_prompt': DNL_PROMPT,
    })


# --- Administration SQL base module ---
def _toolmag_sql_admin_user(request):
    person = current_storekeeper(request)
    return person if is_super_admin_person(person) else None


def sql_database_admin(request):
    from django.contrib import messages
    from django.shortcuts import redirect
    from .db_sql_admin import render_sql_admin
    if not _toolmag_sql_admin_user(request):
        messages.error(request, 'Accès réservé responsable/admin ToolMag.')
        return redirect('dashboard')
    return render_sql_admin(request, 'inventory/sql_database.html', 'ToolMag')


def sql_database_export(request):
    from django.contrib import messages
    from django.shortcuts import redirect
    from .db_sql_admin import export_sql_response
    if not _toolmag_sql_admin_user(request):
        messages.error(request, 'Accès réservé responsable/admin ToolMag.')
        return redirect('dashboard')
    return export_sql_response(request, 'toolmag')


def sql_database_import(request):
    from django.contrib import messages
    from django.shortcuts import redirect
    from .db_sql_admin import import_sql_response
    if not _toolmag_sql_admin_user(request):
        messages.error(request, 'Accès réservé responsable/admin ToolMag.')
        return redirect('dashboard')
    return import_sql_response(request, 'inventory/sql_database.html', 'ToolMag', 'toolmag')

def help_view(request):
    return render(request, 'inventory/help.html')


def about_view(request):
    return render(request, 'inventory/about.html')


def dynamic_loans_display(request):
    """Affichage dynamique des sorties magasin, filtrable par zone LP Core."""
    zone_core_id = (request.GET.get('zone_core_id') or '').strip()
    zone_code = (request.GET.get('zone') or '').strip()
    status_filter = (request.GET.get('status') or 'open').strip()

    loans = Loan.objects.select_related('equipment', 'borrower', 'checkout_storekeeper').order_by('due_at', '-checked_out_at')

    if status_filter == 'open':
        loans = loans.filter(status=Loan.LoanStatus.OPEN)
    elif status_filter == 'late':
        loans = loans.filter(status=Loan.LoanStatus.OPEN, due_at__lt=timezone.now())
    elif status_filter == 'problem':
        loans = loans.filter(status=Loan.LoanStatus.PROBLEM)

    if zone_core_id:
        loans = loans.filter(destination_zone_core_id=zone_core_id)

    if zone_code:
        loans = loans.filter(Q(destination_zone_code=zone_code) | Q(destination_zone_label_snapshot__icontains=zone_code))

    return render(request, 'inventory/dynamic_loans_display.html', {
        'loans': loans[:300],
        'zone_core_id': zone_core_id,
        'zone_code': zone_code,
        'status_filter': status_filter,
        'now': timezone.now(),
    })
